"""
Data Manager (Internal)
Developed by: Sean
Primary users: Internal SE teams for committee list processing and competitive landscape analysis

Changelog:
- v3.2: Original committee data tool (single-table exports)
- v4.0: UI refresh + dashboard export 
- v5.8: multi-structure detection
- v6.4: RapidFuzz grouping + threshold slider (tuned on Q4 exports)
- v7.1: Dashboard analytics modals + export improvements + network viz
- v8.0: Community template support (SESA sheets)
- v8.8: Chart export + UI polish (Patrick's feedback)

Known Issues:
- Keep openpyxl read_only=True (large files spike memory otherwise, learned this the hard way)
- 'SESA' is treated as required identifier in Community mode (legacy template format)
- Fuzzy threshold of 0.85 was tuned on 2025-Q4 committee exports; don't change without re-testing

TODO:
- finalize_export rebuilds tables by re-reading all files (acceptable for now but slow on 50+ files)
- dedupe logic differs slightly between initial read and finalize_export (needs unification)
- Consider caching parsed files to avoid double-reads

Performance Notes:
- Tested with 15+ files (~10MB total) → ~10s load time on typical laptops
- Dashboard generation adds ~2s for large datasets
"""

import customtkinter as ctk
import os
import glob
import sys
import threading
import webbrowser
import json
import time
from datetime import datetime
import tkinter as tk
from tkinter import messagebox, scrolledtext, filedialog
import openpyxl
from rapidfuzz import fuzz  


# CONFIGURATION & CONSTANTS

DEBUG = False  # Set to True to see detailed header/column logs during file processing

MAX_COMPANIES_COMPARE = 5

# Tuned on 2025-Q4 committee exports:
# - 0.90 missed too many merges ("Schneider Electric" vs "Schneider Electric Japan")
# - 0.80 merged unrelated short names ("Schneider" merged with random "Schindler" entities)
# - 0.85 is the sweet spot but still requires manual review on edge cases
DEFAULT_FUZZY_THRESHOLD = 0.85

CONFIG = {
    # Colors (green theme approved by Martial & Patrick)
    "BRAND": {
        "MAIN_COLOR": "#3DCD58", 
        "ERROR_COLOR": "#c0392b",
    },

    # Role weights (approved by Martial & Patrick, last reviewed Jan 2025)
    "ROLE_WEIGHTS": {
        "Secretary": "75",
        "Assistant Sec": "50",
        "Chair": "60",
        "Vice-Chair": "40",
        "Liaison": "15",
        "Convenor": "20",
        "Nat. Part": "5",
        "Member": "5",
    },

    # Analytics settings
    "TOP_GROUPS": 8,  # Dashboard default limit I choose 8 because too for more than 8 it looks messy and unreadable
    "SIGNIFICANT_GAP_THRESHOLD": 15,  # When changed to 20 almost nothing shows up , when changed to 5 almost everything shows up, 15 is a good balance for now based on 2025-Q4 data

    # Fuzzy grouping - common internal shorthand patterns
    # Schneider has tons of variants: "Schneider Electric", "Schneider Electric Espana", "Schneider Electric Indonesia" etc.
    # Kept explicit because fuzzy matching alone is inconsistent with very short tokens
    "CANONICAL_ALIASES": {
        "schneider electric": ["schneider", "schneider electric", "schneider-elec", "schneider elect."],
    },

    "PRINT_TRACEBACKS": False,  # yes this is a bool, don't change, backend logging uses it
}

C_MAIN = CONFIG["BRAND"]["MAIN_COLOR"]
C_ERR = CONFIG["BRAND"]["ERROR_COLOR"]

# CUSTOM EXCEPTIONS

class DataManagerError(Exception):

    def __init__(self, user_message: str, detail: str | None = None):
        super().__init__(user_message)
        self.user_message = user_message
        self.detail = detail


class InvalidTemplateError(DataManagerError):
    pass


class ColumnMismatchError(DataManagerError):
    pass


class ExportError(DataManagerError):
    pass


# UI SETUP

ctk.set_appearance_mode("light")
ctk.set_default_color_theme("green")

# UTILITY HELPERS

def clean_cell_value(val):
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    v = str(val).strip().replace('\xa0', ' ').replace('\n', ' ')
    if v in ["#VALUE!", "#REF!", "#N/A", "#DIV/0!"]:
        return ""
    return v

def safe_int(value, default=0):
    try:
        return int(value)
    except (ValueError, TypeError):
        return default

def get_stripped_text(widget):
    return widget.get().strip()

def extract_headers(ws, header_rows=1):
    raw_headers = []
    for r in ws.iter_rows(min_row=1, max_row=header_rows, values_only=True):
        raw_headers.append(r)
    
    if not raw_headers:
        return []
    
    local_cols = []
    max_c = len(raw_headers[0])
    
    for c_idx in range(max_c):
        parts = []
        for r_idx in range(header_rows):
            if r_idx < len(raw_headers) and c_idx < len(raw_headers[r_idx]):
                v = raw_headers[r_idx][c_idx]
                if v:
                    parts.append(str(v).strip().replace('\n', ' '))
        local_cols.append(" - ".join(parts) if parts else f"Column_{c_idx+1}")
    
    if DEBUG:
        print(f"[DEBUG] Extracted {len(local_cols)} headers from {header_rows} row(s)")
    
    return local_cols


# UI COMPONENTS


class tooltip:

    def __init__(self, widget, msg):
        self.widget = widget
        self.msg = msg
        self.win = None
        self.widget.bind("<Enter>", self.on_enter)
        self.widget.bind("<Leave>", self.on_leave)

    def on_enter(self, event=None):
        try:
            x, y, _, _ = self.widget.bbox("insert")
            x += self.widget.winfo_rootx() + 25
            y += self.widget.winfo_rooty() + 25
            
            self.win = tk.Toplevel(self.widget)
            self.win.wm_overrideredirect(True)
            self.win.wm_geometry(f"+{x}+{y}")
            
            lbl = tk.Label(self.win, text=self.msg, justify='left',
                           background="#2d3436", fg="white", relief='solid', borderwidth=0,
                           font=("Segoe UI", 9))
            lbl.pack(ipadx=5, ipady=3)
        except Exception:
            return


    def on_leave(self, event=None):
        if self.win:

            self.win.destroy()
            self.win = None




class TableStructureNamerDialog(ctk.CTkToplevel):
    def __init__(self, parent, buckets_info, callback):
        super().__init__(parent)
        self.callback = callback
        self.buckets_info = buckets_info 
        self.result = None
        
        self.title("Identify Tables")
        self.geometry("600x500")
        self.grab_set()
        self.protocol("WM_DELETE_WINDOW", self.on_cancel)
        self.bind("<Escape>", lambda e: self.on_cancel())
        

        ctk.CTkLabel(self, text="Different file formats detected", font=("Segoe UI", 16, "bold")).pack(pady=10)
        ctk.CTkLabel(self, text="Give each structure a tab name (what you want to see in the dashboard).", font=("Segoe UI", 12)).pack(pady=(0,10))
        

        self.scroll = ctk.CTkScrollableFrame(self, fg_color="transparent")
        self.scroll.pack(fill="both", expand=True, padx=20, pady=10)
        

        self.entries = {}
        self.header_rows = {}
        
        for i, (sig, data) in enumerate(buckets_info.items()):
            frame = ctk.CTkFrame(self.scroll, fg_color="white")
            frame.pack(fill="x", pady=5)
        


            file_count = len(set(r[-1] for r in data['rows'])) 
            col_preview = ", ".join(data['cols'][:5])
            if len(data['cols']) > 5: col_preview += "..."
            
            info_text = f"Group {i+1}: {file_count} File(s) | Cols: {len(data['cols'])}\nPreview: [{col_preview}]"
            
            ctk.CTkLabel(frame, text=info_text, anchor="w", justify="left", text_color="gray", font=("Consolas", 11)).pack(fill="x", padx=10, pady=5)
            
            ent = ctk.CTkEntry(frame, placeholder_text=f"Enter Name for Table {i+1}...")
            ent.pack(fill="x", padx=10, pady=(0, 10))
            self.entries[sig] = ent


            hdr_var = ctk.StringVar(value="1")
            self.header_rows[sig] = hdr_var

            hdr_frame = ctk.CTkFrame(frame, fg_color="transparent")
            hdr_frame.pack(fill="x", padx=10, pady=(0, 10))

            ctk.CTkLabel(
            hdr_frame,
                text="Header rows to combine:",
                font=("Segoe UI", 11),
                text_color="#636e72"
            ).pack(side="left")

            hdr_menu = ctk.CTkOptionMenu(
                hdr_frame,
                values=["1", "2", "3", "4", "5"],
                variable=hdr_var,
                width=70
            )
            hdr_menu.pack(side="left", padx=(10, 0))
            
        btn_frame = ctk.CTkFrame(self, fg_color="transparent")
        btn_frame.pack(fill="x", pady=15, padx=20)
        
        ctk.CTkButton(btn_frame, text="Cancel", fg_color="#b2bec3", text_color="#2d3436",
              command=self.on_cancel).pack(side="left", expand=True, fill="x", padx=(0, 8))

        ctk.CTkButton(btn_frame, text="Generate Dashboard", fg_color=C_MAIN,
              command=self.on_submit).pack(side="left", expand=True, fill="x")
        

    def on_cancel(self):
        self.result = None
        try:
            self.callback(None)
        finally:
            self.destroy()

    def on_submit(self):
        final_map = {}
        used_names = set()
        


        for sig, ent in self.entries.items():
            name = get_stripped_text(ent)
            if not name:
                messagebox.showwarning("Missing Name", "Please provide a name for all tables.")
                return
            if name in used_names:
                messagebox.showwarning("Duplicate Name", f"Table name '{name}' is used twice.")
                return
            used_names.add(name)
            final_map[name] = {
                "cols": self.buckets_info[sig]["cols"],
                "rows": self.buckets_info[sig]["rows"],
                "files": self.buckets_info[sig].get("files", []),
                "header_rows": int(self.header_rows[sig].get())
            }
        
        self.result = final_map
        self.destroy()
        self.callback(self.result)

# MAIN APPLICATION

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Data Manager v8.0") 
        self.root.geometry("1150x800")
        
        # Determine paths
        if getattr(sys, 'frozen', False):
            self.app_dir = os.path.dirname(sys.executable)
        else:
            self.app_dir = os.path.dirname(os.path.abspath(__file__))
        

        self.data_dir = os.path.join(self.app_dir, 'Data')

        self.out_file = os.path.join(self.app_dir, 'Tool_Result.html')
        
        self.master_cols = []
        self.master_rows = []
        
        # State variables
        self.mode = ctk.StringVar(value="standard")
        self.skip_rows = ctk.StringVar(value="1")
        self.file_count = 0
        self.file_states = {}



        self.settings_visible = False
        self.var_fuzzy = ctk.StringVar(value="off")
        self.var_fuzzy_threshold = ctk.DoubleVar(value=0.85) 
        self.var_fuzzy_col = ctk.StringVar(value="Company")  
        self.var_export_fmt = ctk.StringVar(value="xlsx")
        self.weight_inputs = {}

        self.default_weights = dict(CONFIG["ROLE_WEIGHTS"])



        self.vcmd = (self.root.register(self.check_num), '%P')
        self.init_view()
        self.refresh_file_list()


    def init_view(self):
      

        self.nav = ctk.CTkFrame(self.root, width=260, corner_radius=0, fg_color="#006039")
        self.nav.pack(side="left", fill="y")
        self.nav.pack_propagate(False) 
        
        self.lbl_logo = ctk.CTkLabel(self.nav, text="Data Manager", 
                                     font=ctk.CTkFont(family="Segoe UI", size=26, weight="bold"), 
                                     text_color="white")
        self.lbl_logo.pack(pady=(50, 20))
        

        self.runBtn = ctk.CTkButton(
            self.nav, 
            text="PROCESS DATA", 
            font=ctk.CTkFont(family="Segoe UI", size=13, weight="bold"),
            fg_color=C_MAIN, 
            text_color="white",
            hover_color="#2eb84a",
            height=45,
            corner_radius=8,
            command=self.on_run_click
        )
        self.runBtn.pack(pady=(15, 5), padx=25, fill="x")
        self.pbar = ctk.CTkProgressBar(self.nav, height=10, corner_radius=5, progress_color=C_MAIN, fg_color="#004d2e")
        self.pbar.set(0)


        self.pbar.pack_forget()
        self.btn_settings = ctk.CTkButton(
            self.nav, 
            text="SETTINGS", 
            font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"),
            fg_color="transparent",
            border_width=1,
            border_color="#aeb6bf",
            text_color="#aeb6bf",
            hover_color="#004d2e",
            height=30,
            corner_radius=8,
            command=self.toggle_settings
        )
        self.btn_settings.pack(pady=5, padx=25, fill="x")
        self.settings_frame = ctk.CTkScrollableFrame(self.nav, fg_color="#004d2e", height=320, label_text="Configuration", label_text_color="white")
        

        # Op Dashboard
        self.openBtn = ctk.CTkButton(
            self.nav, 
            text="VIEW DASHBOARD", 
            font=ctk.CTkFont(family="Segoe UI", size=13, weight="bold"),
            fg_color="transparent",
            border_width=2,
            border_color="#aeb6bf",
            text_color="#aeb6bf",
            hover_color="#004d2e",
            height=45,
            corner_radius=8,
            state="disabled", 
            command=self.launch_report
        )
        self.openBtn.pack(side="bottom", pady=20, padx=25, fill="x")
        ctk.CTkLabel(self.nav, text="v8.0 | Internal Use", text_color="#8FBC8F", font=("Segoe UI", 11)).pack(side="bottom", pady=5)






        # Main
        self.main_area = ctk.CTkFrame(self.root, fg_color="#F5F7FA", corner_radius=0)
        self.main_area.pack(side="right", fill="both", expand=True)       
        self.container = ctk.CTkFrame(self.main_area, fg_color="transparent")
        self.container.pack(fill="both", expand=True, padx=30, pady=30)
        self.top_bar = ctk.CTkFrame(self.container, fg_color="transparent")
        self.top_bar.pack(fill="x", pady=(0, 15))
        self.cv_led = tk.Canvas(self.top_bar, width=15, height=15, bg="#f0f2f5", highlightthickness=0)
        self.led = self.cv_led.create_oval(2, 2, 13, 13, fill="gray", outline="")
        self.cv_led.pack(side="left", padx=(0, 10))
        self.lbl_path = ctk.CTkLabel(self.top_bar, text=f"SOURCE: {self.data_dir}", 
                                     font=("Consolas", 12, "bold"), text_color="#636e72")
        self.lbl_path.pack(side="left")
       
        self.mainGrid = ctk.CTkFrame(self.container, fg_color="transparent")
        self.mainGrid.pack(fill="x", expand=False, pady=(0, 20))
        self.mainGrid.grid_columnconfigure(0, weight=3) 
        self.mainGrid.grid_columnconfigure(1, weight=2) 



        # Left 
        self.configCard = ctk.CTkFrame(self.mainGrid, fg_color="white", corner_radius=15)
        self.configCard.grid(row=0, column=0, sticky="nsew", padx=(0, 15))
        
        ctk.CTkLabel(self.configCard, text="Import Settings", font=("Segoe UI", 16, "bold"), text_color="#2d3436").pack(anchor="w", padx=20, pady=15)
        

        self.rb_std = ctk.CTkRadioButton(self.configCard, text="Standard Mode (Committee List Dataset)",  
                                         variable=self.mode, value="standard", font=("Segoe UI", 13),
                                         fg_color=C_MAIN, command=self.update_inputs)
        self.rb_std.pack(anchor="w", padx=20, pady=(5, 10))
        tooltip(self.rb_std, "Standard mode. Choose this for committee list files\nwith single-row headers. This mode enables advanced analytics.")
        self.rb_gen = ctk.CTkRadioButton(self.configCard, text="Normal Mode", 
                                         variable=self.mode, value="normal", font=("Segoe UI", 13),
                                         fg_color=C_MAIN, command=self.update_inputs)
        self.rb_gen.pack(anchor="w", padx=20, pady=(5, 10))
        
        tooltip(self.rb_gen, "Generic mode. Choose this for standard 1-row files\nOR multi-row headers. Choosing this option will not give access to advanced analytics.")

        self.rb_comm = ctk.CTkRadioButton(self.configCard, text="Community Mode (SESA Multi-Sheet Template)", 
                                         variable=self.mode, value="community", font=("Segoe UI", 13),
                                         fg_color=C_MAIN, command=self.update_inputs)
        self.rb_comm.pack(anchor="w", padx=20, pady=(5, 10))
        tooltip(self.rb_comm, "Community mode. Use this for the Community Excel template\nwith multiple sheets (tabs) and the unique ID 'SESA'.")

        self.spin_box = ctk.CTkFrame(self.configCard, fg_color="transparent")
        self.spin_box.pack(anchor="w", padx=45, pady=(0, 20))
        
        ctk.CTkLabel(self.spin_box, text="Default Header Rows for preview:", font=("Segoe UI", 12), text_color="#636e72").pack(side="left", padx=(0, 10))
        
        self.ent_rows = ctk.CTkEntry(self.spin_box, width=50, textvariable=self.skip_rows,
                                     justify="center", font=("Segoe UI", 13), border_color="#dfe6e9",
                                     validate='key', validatecommand=self.vcmd)
        self.ent_rows.pack(side="left")
        self.ent_rows.configure(state="disabled") 


        # Right 
        self.fileCard = ctk.CTkFrame(self.mainGrid, fg_color="white", corner_radius=15)
        self.fileCard.grid(row=0, column=1, sticky="nsew")
        
        self.file_header_frame = ctk.CTkFrame(self.fileCard, fg_color="transparent")
        self.file_header_frame.pack(fill="x", padx=15, pady=10)

        self.lbl_files = ctk.CTkLabel(self.file_header_frame, text="Pending Files (0)", font=("Segoe UI", 14, "bold"), text_color="#2d3436")
        self.lbl_files.pack(side="left")

        self.btn_refresh = ctk.CTkButton(self.file_header_frame, text="↻", width=30, height=24,
                                         fg_color="#dfe6e9", text_color="#2d3436", hover_color="#b2bec3",
                                         font=("Segoe UI", 14), command=self.refresh_file_list)
        self.btn_refresh.pack(side="right")
        tooltip(self.btn_refresh, "Refresh file list")
        
        self.file_scroll = ctk.CTkScrollableFrame(self.fileCard, fg_color="transparent", height=120)
        self.file_scroll.pack(fill="both", expand=True, padx=5, pady=(0, 10))


        # Logs CMD lke
        ctk.CTkLabel(self.container, text="System Logs", font=("Segoe UI", 14, "bold"), text_color="#2d3436").pack(anchor="w", pady=(0, 5))
        
        self.log_frame = ctk.CTkFrame(self.container, fg_color="white", corner_radius=15)
        self.log_frame.pack(fill="both", expand=True)
        
        self.txt_log = scrolledtext.ScrolledText(self.log_frame, height=6, bg="#1e272e", fg="#d2dae2", 
                                                 font=("Consolas", 11), insertbackground="white", relief="flat", padx=15, pady=15, bd=0)
        self.txt_log.pack(fill="both", expand=True, padx=5, pady=5)



    def toggle_settings(self):
        if not self.settings_visible:
            self.btn_settings.configure(fg_color="#004d2e", text="⚙️  CLOSE SETTINGS")
            self.settings_frame.pack(after=self.btn_settings, pady=10, padx=20, fill="x")
            self.render_settings_content()
            self.settings_visible = True
        else:
            self.settings_frame.pack_forget()
            self.btn_settings.configure(fg_color="transparent", text="⚙️  SETTINGS")
            self.settings_visible = False

    def render_settings_content(self):
        for w in self.settings_frame.winfo_children():
            w.destroy()

        # Weights
        ctk.CTkLabel(self.settings_frame, text="Power Weights", font=("Segoe UI", 11, "bold"), text_color="#bdc3c7").pack(anchor="w", pady=(5,5))
        self.weight_inputs = {}
        for role, val in self.default_weights.items():
            r_frame = ctk.CTkFrame(self.settings_frame, fg_color="transparent", height=25)
            r_frame.pack(fill="x", pady=2)
            
            lbl = ctk.CTkLabel(r_frame, text=role, font=("Segoe UI", 11), text_color="white", width=120, anchor="w")
            lbl.pack(side="left")
            
            entry = ctk.CTkEntry(r_frame, width=50, height=22, font=("Consolas", 11), justify="center")
            entry.insert(0, val)
            entry.configure(state="disabled", text_color="gray") 
            entry.pack(side="right")
            self.weight_inputs[role] = entry

        ctk.CTkFrame(self.settings_frame, height=2, fg_color="#bdc3c7").pack(fill="x", pady=10)





        # Fuzzy
        ctk.CTkLabel(self.settings_frame, text="Advanced Logic (RapidFuzz)", font=("Segoe UI", 11, "bold"), text_color="#bdc3c7").pack(anchor="w", pady=(0,5))
    
        fuzzy_switch = ctk.CTkSwitch(self.settings_frame, text="Enable Smart Grouping", variable=self.var_fuzzy, 
                                     onvalue="on", offvalue="off", command=self.toggle_fuzzy_slider,
                                     font=("Segoe UI", 11), text_color="white", progress_color=C_MAIN)
        fuzzy_switch.pack(anchor="w", pady=5)
        
        self.ent_fuzzy_col = ctk.CTkEntry(self.settings_frame, textvariable=self.var_fuzzy_col, height=24, font=("Segoe UI", 11))
        self.ent_fuzzy_col.pack(fill="x", padx=25, pady=(0, 10))
        


        # Slider
        self.fuzzy_container = ctk.CTkFrame(self.settings_frame, fg_color="transparent")
        self.fuzzy_container.pack(fill="x", pady=(0, 10))
        
        self.lbl_fuzzy_val = ctk.CTkLabel(self.fuzzy_container, text=f"Match Sensitivity: {int(self.var_fuzzy_threshold.get()*100)}%", 
                                          font=("Segoe UI", 10), text_color="#dfe6e9")
        self.lbl_fuzzy_val.pack(anchor="w", padx=25)
        
        self.slider_fuzzy = ctk.CTkSlider(self.fuzzy_container, from_=0.5, to=1.0, variable=self.var_fuzzy_threshold,
                                          command=self.update_fuzzy_lbl, height=16, progress_color=C_MAIN)
        self.slider_fuzzy.pack(fill="x", padx=20)
        self.toggle_fuzzy_slider()

        # Export Format
        ctk.CTkLabel(self.settings_frame, text="Export Format", font=("Segoe UI", 11, "bold"), text_color="#bdc3c7").pack(anchor="w", pady=(5,5))
        ctk.CTkOptionMenu(self.settings_frame, values=["xlsx", "csv"], variable=self.var_export_fmt,
                          fg_color="#1e272e", button_color="#1e272e", width=100).pack(anchor="w", pady=5)






    def toggle_fuzzy_slider(self):
        if self.var_fuzzy.get() == "on":
            self.slider_fuzzy.configure(state="normal", progress_color=C_MAIN)
            self.ent_fuzzy_col.configure(state="normal", border_color="#555")
            self.lbl_fuzzy_val.configure(text_color="#dfe6e9")
        else:
            self.slider_fuzzy.configure(state="disabled", progress_color="gray")
            self.ent_fuzzy_col.configure(state="disabled", border_color="#333")
            self.lbl_fuzzy_val.configure(text_color="gray")

    def update_fuzzy_lbl(self, val):
        self.lbl_fuzzy_val.configure(text=f"Strictness: {int(val*100)}%")

    def get_current_weights(self):
        if not self.weight_inputs: return self.default_weights.copy()
        current = {}
        for role, entry_widget in self.weight_inputs.items():
            try:
                val = get_stripped_text(entry_widget)
                if val.isdigit(): current[role] = val
                else: current[role] = self.default_weights[role]
            except:
                current[role] = self.default_weights[role]
        return current

    def check_num(self, val):
        return val.isdigit() or val == ""

    def update_inputs(self):
        if self.mode.get() == "normal":
            self.ent_rows.configure(state="normal", border_color=C_MAIN)
            self.ent_rows.focus_set()
        else:
            self.ent_rows.configure(state="disabled", border_color="#dfe6e9")

    def toggle_led(self, color):
        self.cv_led.itemconfig(self.led, fill=color)

    def log(self, txt):
        self.txt_log.insert(tk.END, f"> {txt}\n")
        self.txt_log.see(tk.END)

    def refresh_file_list(self):
        self.log("scanning folder...") 
        self.file_states.clear()
        
        if not os.path.exists(self.data_dir):
            os.makedirs(self.data_dir)
            self.log(f"Created input folder: {self.data_dir}")
            self.toggle_led("gray")
            self.lbl_files.configure(text="Pending Files (0)")
        else:
            files = glob.glob(os.path.join(self.data_dir, '*.xlsx'))
            self.file_count = len(files)
            
            for w in self.file_scroll.winfo_children(): w.destroy()

            if self.file_count > 0:
                self.toggle_led(C_MAIN)
                self.lbl_files.configure(text=f"Pending Files ({self.file_count})")
                
                self.var_select_all = ctk.StringVar(value="on")
                
             
                ctrl_frame = ctk.CTkFrame(self.file_scroll, fg_color="transparent")
                ctrl_frame.pack(fill="x", pady=(0, 5))
                
                switch = ctk.CTkSwitch(ctrl_frame, text="Select All", variable=self.var_select_all, 
                                       onvalue="on", offvalue="off", command=self.toggle_select_all,
                                       font=("Segoe UI", 11, "bold"), height=20, progress_color=C_MAIN)
                switch.pack(side="left", padx=5)
                
                files.sort()
                for f in files:
                    fname = os.path.basename(f)
                    kb = os.path.getsize(f) / 1024
                    
                    f_var = ctk.StringVar(value="on")
                    self.file_states[f] = f_var
                    
                    r = ctk.CTkFrame(self.file_scroll, fg_color="transparent")
                    r.pack(fill="x", pady=2)
                    
                    chk = ctk.CTkCheckBox(r, text=fname, variable=f_var, onvalue="on", offvalue="off",
                                          font=("Segoe UI", 12), checkbox_height=18, checkbox_width=18,
                                          hover_color=C_MAIN, fg_color=C_MAIN)
                    chk.pack(side="left", padx=5)
                    
                    ctk.CTkLabel(r, text=f"{kb:.1f} KB", font=("Segoe UI", 11), text_color="gray").pack(side="right", padx=5)

                self.log(f"Ready. Found {self.file_count} files.")
            else:
                self.toggle_led("gray")
                self.lbl_files.configure(text="Pending Files (0)")
                self.log("Folder is empty.")
                ctk.CTkLabel(self.file_scroll, text="No .xlsx files found", text_color="gray").pack(pady=10)

    def toggle_select_all(self):
        val = self.var_select_all.get()
        for var in self.file_states.values(): var.set(val)

    def set_ui_idle(self, success=False):
        self.runBtn.configure(state="normal", text="▶  PROCESS DATA")
        self.pbar.pack_forget() 
        
        if success:
            self.toggle_led(C_MAIN)
            self.anim_btn() 
        else:
            self.toggle_led(C_ERR)
            self.openBtn.configure(state="disabled", border_color="#aeb6bf", text_color="#aeb6bf", text="🌐  OPEN DASHBOARD")

    def anim_btn(self, step=0):
        self.openBtn.configure(state="normal", border_color=C_MAIN, text_color=C_MAIN)
        cols = ["#e8f5e9", "transparent"] 
        if step < 6: 
            c = cols[step % 2]
            self.openBtn.configure(fg_color=c)
            self.root.after(300, lambda: self.anim_btn(step + 1))
        else:
            self.openBtn.configure(fg_color=C_MAIN, text_color="white", text="✔  OPEN DASHBOARD")

    def on_run_click(self):
        self.txt_log.delete(1.0, tk.END)
        selected_files = [f for f, var in self.file_states.items() if var.get() == "on"]
        
        if not selected_files:
            messagebox.showwarning("Selection Error", "Please select at least one file to process.")
            return


        weights_config = self.get_current_weights()
        fuzzy_enabled = (self.var_fuzzy.get() == "on")
        fuzzy_thresh = self.var_fuzzy_threshold.get()
        export_fmt = self.var_export_fmt.get()
        target_col_name = get_stripped_text(self.var_fuzzy_col) or "Company"

        self.runBtn.configure(state="disabled", text="PROCESSING...")
        self.openBtn.configure(state="disabled", fg_color="transparent", text_color="#aeb6bf", text="🌐  OPEN DASHBOARD", border_color="#aeb6bf")


        self.pbar.pack(pady=(0, 20), padx=25, fill="x", after=self.runBtn)
        self.pbar.set(0)

        self.toggle_led("#f1c40f") 
        
        t = threading.Thread(
            target=self.worker_process_files, 
            args=(selected_files, weights_config, fuzzy_enabled, fuzzy_thresh, export_fmt, target_col_name), 
            daemon=True
        )

        t.start()


    # FILE INGESTION PIPELINE

    # Community mode helpers
    def is_community_workbook(self, wb):
        required = ["Personal Information", "Standards Group", "Internal Groups", "Association", "Expertise"]
        names = [s.strip() for s in wb.sheetnames]
        if not all(r in names for r in required):
            return False
        try:
            for s in required:
                ws = wb[s]
                hdr = [str(v).strip().lower() for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)) if v is not None]
                if "sesa" not in hdr:
                    return False
            return True
        except Exception:
            return False

    def _read_sheet_table(self, ws, fname, sheet_name):
        cols = extract_headers(ws, header_rows=1)
        
        last = 0
        for i, v in enumerate(cols):
            if v and not v.startswith("Column_"):
                last = i + 1
        cols = cols[:last] if last > 0 else cols

        sesa_idx = -1
        for i, c in enumerate(cols):
            if c.strip().lower() == "sesa":
                sesa_idx = i
                break

        rows = []
        seen = set()
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not any(r):
                continue

            row_data = []
            for idx in range(len(cols)):
                val = r[idx] if idx < len(r) else None
                row_data.append(clean_cell_value(val))

            if sesa_idx != -1:
                if not row_data[sesa_idx] or str(row_data[sesa_idx]).strip() == "":
                    continue

           
            row_data.append(fname)
            row_data.append(sheet_name)

            sig = tuple(row_data)
            if sig in seen:
                continue
            seen.add(sig)
            rows.append(row_data)

        
        cols2 = cols + ["📄 Source File", "🧾 Source Sheet"]
        return cols2, rows

    def build_master_table(self, tables_by_sheet):
        def find_sesa_idx(cols):
            for i, c in enumerate(cols):
                if str(c).strip().lower() == "sesa":
                    return i
            return -1



        all_sesa = set()
        for t in tables_by_sheet.values():
            si = find_sesa_idx(t["cols"])
            if si == -1:
                continue
            for r in t["rows"]:
                all_sesa.add(str(r[si]).strip())
        all_sesa = sorted([s for s in all_sesa if s])

        
        
        base = tables_by_sheet.get("Personal Information")
        base_cols = base["cols"] if base else ["SESA"]
        base_sesa_idx = find_sesa_idx(base_cols)
        base_map = {}
        if base and base_sesa_idx != -1:
            for r in base["rows"]:
                base_map[str(r[base_sesa_idx]).strip()] = r

        agg_cols = []
        agg_maps = {}
        for sheet, t in tables_by_sheet.items():
            if sheet == "Personal Information":
                continue
            cols = t["cols"]
            si = find_sesa_idx(cols)
            if si == -1:
                continue
            biz_cols = [c for c in cols if c not in ["📄 Source File", "🧾 Source Sheet"]]
            out_col = f"{sheet} - Entries"
            agg_cols.append(out_col)
            m = {}
            for r in t["rows"]:
                sesa = str(r[si]).strip()
                parts = []
                for j, c in enumerate(biz_cols):
                    if j == si:
                        continue
                    v = (r[j] if j < len(r) else "")
                    if v and str(v).strip():
                        parts.append(str(v).strip())
                entry = " | ".join(parts)
                if not entry:
                    continue
                m.setdefault(sesa, set()).add(entry)

            agg_maps[out_col] = {k: "; ".join(sorted(v)) for k, v in m.items()}


        final_cols = list(base_cols)
        if find_sesa_idx(final_cols) == -1:
            final_cols.insert(0, "SESA")
        for c in agg_cols:
            if c not in final_cols:
                final_cols.append(c)


        out_rows = []
        out_sesa_idx = find_sesa_idx(final_cols)

        for sesa in all_sesa:
            if base and sesa in base_map:
                row = list(base_map[sesa])

                if len(row) < len(base_cols):
                    row += [""] * (len(base_cols) - len(row))
            else:
                row = [""] * len(base_cols)
                if out_sesa_idx != -1:
                    if out_sesa_idx < len(row):
                        row[out_sesa_idx] = sesa
                    else:
                        row = [sesa] + row

            for c in agg_cols:
                row.append(agg_maps.get(c, {}).get(sesa, ""))
            out_rows.append(row)

        return {"cols": final_cols, "rows": out_rows}

    def export_dashboard_direct(self, named_buckets_simple, is_std, weights, export_fmt):
        try:
            self.log("Generating HTML Dashboard...")
            html_out = render_dashboard_html(named_buckets_simple, is_std, weights, export_fmt)
            with open(self.out_file, 'w', encoding='utf-8') as f:
                f.write(html_out)
            self.log("Done.")
            self.set_ui_idle(True)
            messagebox.showinfo("Success", "Processing complete!")
        except DataManagerError as e:
            self.log(f"{e.__class__.__name__}: {e.user_message}")
            if e.detail:
                self.log(f"Details: {e.detail}")
            messagebox.showerror("Export error", e.user_message)
            self.set_ui_idle(False)
        except Exception as e:
            self.log(f"Unexpected export error: {e}")
            if CONFIG["PRINT_TRACEBACKS"]:
                import traceback; traceback.print_exc()
            messagebox.showerror("Export error", "Unexpected error while exporting. Check the logs for details.")
            self.set_ui_idle(False)

    def run_fuzzy_logic(self, rows, headers, threshold, target_col_name):
        col_idx = -1
        target = target_col_name.lower()
        
        lower_headers = [h.lower() for h in headers]
        for i, h in enumerate(lower_headers):
            if target in h:
                col_idx = i
                break
        
        if col_idx == -1:
            self.log(f"WARN: Grouping skipped, col '{target_col_name}' not found.")
            return

        rf_threshold = threshold * 100
    
        
        unique_vals = list(set([str(r[col_idx]).strip() for r in rows if r[col_idx] and str(r[col_idx]).strip() != ""]))
        unique_vals.sort(key=len)
        
        mapping = {}
        self.log(f"analyzing {len(unique_vals)} unique values for grouping...")
        
        for i, short_val in enumerate(unique_vals):
            s_lower = short_val.lower()
         
            # Common internal shorthand: "SE", "S.E.", "Schneider-elec" etc.
            # Kept explicit because fuzzy alone can be inconsistent with very short tokens
            if "schneider" in s_lower or s_lower in ["se", "s.e."]:
                mapping[short_val] = "Schneider Electric"
                continue
            
            if short_val in mapping: continue 
            if len(short_val) < 2:
                mapping[short_val] = short_val
                continue

            mapping[short_val] = short_val
            
            for long_val in unique_vals[i+1:]:
                if long_val in mapping: continue

                l_lower = long_val.lower()
                if "schneider" in l_lower or l_lower in ["se", "s.e."]:
                    mapping[long_val] = "Schneider Electric"
                    continue

                if l_lower.startswith(s_lower):
                    mapping[long_val] = short_val
                    continue

                len_ratio = len(short_val) / len(long_val)
                if len_ratio < 0.6: continue 

                score = fuzz.token_sort_ratio(s_lower, l_lower)
                if score >= rf_threshold:
                    mapping[long_val] = short_val
       
        count = 0
        for r in rows:
            orig = str(r[col_idx]).strip()
            if orig in mapping and mapping[orig] != orig:
                r[col_idx] = mapping[orig]
                count += 1
        
        self.log(f"Grouping done. Merged {count} records.")

    # Main worker
    def worker_process_files(self, target_files, weights, is_fuzzy, fuzzy_thresh, export_fmt, target_col_name):
        try:
            t0 = time.time()  # Track processing time
            curr_mode = self.mode.get()
            is_std = (curr_mode == "standard")

            #  COMMUNITY MODE PIPELINE 
            if curr_mode == "community":
                self.log("Community mode: reading templates...")
                tables_by_sheet = {}
                required = ["Personal Information", "Standards Group", "Internal Groups", "Association", "Expertise"]

                for i, fpath in enumerate(target_files):
                    prog = (i + 1) / len(target_files)
                    self.root.after(0, lambda p=prog: self.pbar.set(p))
                    fname = os.path.basename(fpath)
                    self.log(f"Reading {fname} ({i+1}/{len(target_files)})...")

                    wb = None
                    try:
                        wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)

                        if not self.is_community_workbook(wb):
                            raise InvalidTemplateError(
                            user_message=f"'{fname}' is not a valid Community template.",
                            detail="Missing required sheets or missing a 'SESA' header in at least one required sheet."
                        )

                        for sheet_name in required:
                            ws = wb[sheet_name]
                            cols, rows = self._read_sheet_table(ws, fname, sheet_name)

                            if sheet_name not in tables_by_sheet:
                                tables_by_sheet[sheet_name] = {"cols": cols, "rows": []}
                            tables_by_sheet[sheet_name]["rows"].extend(rows)

                    except Exception as e:
                        self.log(f"Error reading {fname}: {e}")
                        raise
                    finally:
                        if wb: wb.close()

                if not tables_by_sheet:
                    self.log("No valid data found.")
                    self.root.after(0, lambda: self.set_ui_idle(False))
                    return

                for sheet_name, t in tables_by_sheet.items():
                    seen = set()
                    unique_rows = []
                    for r in t["rows"]:
                        sig = tuple(r)
                        if sig in seen: 
                            continue
                        seen.add(sig)
                        unique_rows.append(r)
                    t["rows"] = unique_rows


                master = self.build_master_table(tables_by_sheet)
                tables_by_sheet["Community Master"] = master

                named_buckets_simple = {k: {"cols": v["cols"], "rows": v["rows"]} for k, v in tables_by_sheet.items()}
                self.root.after(0, lambda: self.export_dashboard_direct(named_buckets_simple, False, weights, export_fmt))
                return

            h_count = 1
            if not is_std:
                try:
                    h_val = safe_int(self.skip_rows.get(), 1)
                    if h_val > 0: h_count = h_val
                except: pass

            self.log(f"{'Standard' if is_std else 'Normal'} mode: detecting file structures...")
            target_files.sort()
            
            
            buckets = {}
            first_sig_standard = None

            for i, fpath in enumerate(target_files):
                prog = (i + 1) / len(target_files)
                self.root.after(0, lambda p=prog: self.pbar.set(p))
                fname = os.path.basename(fpath)
                self.log(f"Reading {fname} ({i+1}/{len(target_files)})...")
                
                wb = None
                try:
                    wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
                    ws = wb.active
                    
                    if ws.max_row is not None and ws.max_row < h_count: 
                        self.log("Skipping empty sheet")
                        continue

                    # Extract Headers
                    local_cols = extract_headers(ws, h_count)
                    if not local_cols: continue
                    
                    sig = tuple(local_cols)
                    
                    if is_std:
                        if first_sig_standard is None: first_sig_standard = sig
                        elif sig != first_sig_standard:
                            raise ColumnMismatchError(
                            user_message=f"'{fname}' has different columns. Standard mode requires identical headers across all files.",
                            detail="Ensure the header row is identical in every file, or switch to Normal Mode."
                        )
                    
                    if sig not in buckets:
                        buckets[sig] = { 'cols': list(local_cols), 'rows': [], 'files': [] }

                    buckets[sig]['files'].append(fpath)
                    
                    file_rows = []
                    # FIXME: dedupe should be consistent between initial read and finalize_export
                    seen_rows = set()  

                    for r in ws.iter_rows(min_row=h_count+1, values_only=True):
                        if not any(r):
                            continue

                        row_data = []
                        for idx in range(len(local_cols)):
                            val = r[idx] if idx < len(r) else None
                            row_data.append(clean_cell_value(val))


                        row_data.append(fname)

                        sig_row = tuple(row_data) 
                        if sig_row in seen_rows:
                            continue
                        seen_rows.add(sig_row)

                        file_rows.append(row_data)

                    buckets[sig]['rows'].extend(file_rows)

                except Exception as e:
                    self.log(f"Error reading {fname}: {e}")
                    if is_std: raise e 
                finally:
                    if wb: wb.close()

            if not buckets:
                self.log("No valid data found."); self.root.after(0, lambda: self.set_ui_idle(False)); return

            self.log("Processing buckets...")
            
            for sig in buckets:
                b_cols = buckets[sig]['cols']
                b_rows = buckets[sig]['rows']
                if is_fuzzy: self.run_fuzzy_logic(b_rows, b_cols, fuzzy_thresh, target_col_name)
            
            self.bucket_data_payload = (buckets, weights, is_std, export_fmt, is_fuzzy, fuzzy_thresh, target_col_name)
            self.root.after(0, self.prompt_table_names)
            
        except DataManagerError as e:
            self.log(f"{e.__class__.__name__}: {e.user_message}")
            if e.detail:
                self.log(f"Details: {e.detail}")
            messagebox.showerror("Processing error", e.user_message)
            self.root.after(0, lambda: self.set_ui_idle(False))
        except Exception as e:
            self.log(f"Unexpected error: {e}")
            if CONFIG["PRINT_TRACEBACKS"]:
                import traceback; traceback.print_exc()
            messagebox.showerror("Processing error", "Unexpected error while processing. Check the logs for details.")
            self.root.after(0, lambda: self.set_ui_idle(False))

    def prompt_table_names(self):
        buckets, weights, is_std, export_fmt, is_fuzzy, fuzzy_thresh, target_col_name = self.bucket_data_payload
        
        if len(buckets) > 15:
            messagebox.showerror("Limit Exceeded", f"Found {len(buckets)} different file structures. The limit is 15. Please cleanup your folder.")
            self.set_ui_idle(False)
            return

        if is_std or len(buckets) == 1:
            named_buckets = {}
            for sig, data in buckets.items():
                named_buckets["Main Table" if len(buckets)==1 else "Data"] = data
            self.finalize_export(named_buckets)
        else:
            TableStructureNamerDialog(self.root, buckets, self.finalize_export)

    def finalize_export(self, named_buckets):
        if not named_buckets:
            self.log("Operation cancelled."); self.set_ui_idle(False); return
            
        try:
            t0 = time.time()
            # TODO: This rebuild re-reads all files; acceptable for now but slow on 50+ files
            _, weights, is_std, export_fmt, is_fuzzy, fuzzy_thresh, target_col_name = self.bucket_data_payload
            self.log("Generating HTML Dashboard...")
            
            rebuilt = {}
            for tab_name, data in named_buckets.items():
                h_count = safe_int(data.get("header_rows"), 1)

                new_cols = None
                new_rows = []

                for fpath in data.get("files", []):
                    wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
                    ws = wb.active
                    fname = os.path.basename(fpath)


                    local_cols = extract_headers(ws, h_count)
                    if not local_cols:
                        wb.close()
                        continue


                    if new_cols is None:
                        new_cols = local_cols

                    for r in ws.iter_rows(min_row=h_count + 1, values_only=True):
                        if not any(r):
                            continue
                    
                        row_data = []
                        for idx in range(len(local_cols)):
                            val = r[idx] if idx < len(r) else None
                            row_data.append(clean_cell_value(val))

                        row_data.append(fname)
                        new_rows.append(row_data)

                    wb.close()


                if is_fuzzy:
                    self.run_fuzzy_logic(new_rows, new_cols or [], fuzzy_thresh, target_col_name)
                rebuilt[tab_name] = {"cols": new_cols or [], "rows": new_rows}

            named_buckets = rebuilt

            html_out = render_dashboard_html(named_buckets, is_std, weights, export_fmt)
            
            with open(self.out_file, 'w', encoding='utf-8') as f:
                f.write(html_out)
            

            elapsed = time.time() - t0
            self.log(f"Done — dashboard updated ({elapsed:.1f}s)"); self.set_ui_idle(True)
            messagebox.showinfo("Success", "Processing complete!")
        except DataManagerError as e:
            self.log(f"{e.__class__.__name__}: {e.user_message}")
            if e.detail:
                self.log(f"Details: {e.detail}")
            messagebox.showerror("Export error", e.user_message)
            self.set_ui_idle(False)
        except Exception as e:
            self.log(f"Unexpected export error: {e}")
            if CONFIG["PRINT_TRACEBACKS"]:
                import traceback; traceback.print_exc()
            messagebox.showerror("Export error", "Unexpected error while exporting. Check the logs for details.")
            self.set_ui_idle(False)

    def launch_report(self):
        if os.path.exists(self.out_file):
            webbrowser.open('file://' + os.path.abspath(self.out_file))
            self.log("Opening Dashboard...")
        else:
            messagebox.showerror("Error", "File not found.")



# DASHBOARD (HTML/CSS/JS)

HTML_HEAD = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Schneider Data Manager</title>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet">
    <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css">
    <link rel="stylesheet" href="https://cdn.datatables.net/1.13.6/css/jquery.dataTables.min.css">
"""

HTML_SCRIPTS = """
    <script src="https://code.jquery.com/jquery-3.7.0.min.js"></script>
    <script src="https://cdn.datatables.net/1.13.6/js/jquery.dataTables.min.js"></script>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/xlsx/0.18.5/xlsx.full.min.js"></script>
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <script src="https://cdn.jsdelivr.net/npm/echarts@5.4.3/dist/echarts.min.js"></script>
"""

# Dashboard rendering

def render_dashboard_html(named_buckets, is_std, weights_dict=None, export_fmt='xlsx'):
    tables_js = {}
    
    for t_name, data in named_buckets.items():
        js_cols = [{"title": c} for c in data['cols']]
        
        if "📄 Source File" not in data['cols']:
             js_cols.append({"title": "📄 Source File"})
        f_counts = {}
        for r in data['rows']:
            fn = r[-1] 
            f_counts[fn] = f_counts.get(fn, 0) + 1
        
        tables_js[t_name] = {
            "columns": js_cols,
            "data": data['rows'],
            "file_labels": list(f_counts.keys()),
            "file_counts": list(f_counts.values())
        }

    # Get CSS and JS 
    css_block = _get_dashboard_css()
    js_block = _get_dashboard_js(tables_js, is_std, export_fmt, weights_dict)
    
    logo_url = "https://upload.wikimedia.org/wikipedia/commons/thumb/9/95/Schneider_Electric_2007.svg/320px-Schneider_Electric_2007.svg.png"
    
    # search options
    first_table_name = list(tables_js.keys())[0]
    first_cols = tables_js[first_table_name]['columns']
    col_opts = '<option value="all">Global Search...</option>'
    for i, c in enumerate(first_cols):
        col_opts += f'<option value="{i}">{c["title"]}</option>'

    #html anal
    if is_std:
        analytics_html = """
            <div class="analytics-grid">
                <div class="action-card" id="powerBtn">
                    <div class="icon-box bg-purple"><i class="fas fa-bolt"></i></div>
                    <div class="action-text"><h4>Company Rankings</h4><p>Power Analysis</p></div>
                </div>
                <div class="action-card" id="radarBtn">
                    <div class="icon-box bg-red"><i class="fas fa-crosshairs"></i></div>
                    <div class="action-text"><h4>Competitor Comparison</h4><p>Domain Analysis</p></div>
                </div>
                <div class="action-card" id="networkBtn">
                    <div class="icon-box bg-teal"><i class="fas fa-project-diagram"></i></div>
                    <div class="action-text"><h4>Network Map</h4><p>Connection Analysis</p></div>
                </div>
                <div class="action-card" id="landscapeBtn">
                    <div class="icon-box bg-blue"><i class="fas fa-balance-scale"></i></div>
                    <div class="action-text"><h4>Gaps & Leads</h4><p>Competitors Analysis</p></div>
                </div>
                <div class="action-card" id="geoBtn">
                    <div class="icon-box bg-green-dark" style="background:#2d3436;"><i class="fas fa-globe-americas"></i></div>
                    <div class="action-text"><h4>Geographic Heatmap</h4><p>Regional Power Analysis</p></div>
                </div>
                <div class="action-card" id="benchBtn">
                    <div class="icon-box bg-orange"><i class="fas fa-user-shield"></i></div>
                    <div class="action-text"><h4>Expert Workload S.E.</h4><p>Role Strength Analysis</p></div>
                </div>
            </div>
        """
    else:
        analytics_html = """
            <div class="empty-state">
                <i class="fas fa-info-circle"></i>
                <span>Advanced Analytics are disabled in Normal Mode (Multiple file structures detected).</span>
            </div>
        """

    # Weights Legend 
    if weights_dict is None: weights_dict = {}
    weight_visual_html = """
    <div class="card" style="margin-top: 20px; padding: 15px;">
        <div class="sidebar-title">Position Weight</div>
        <table style="width:100%; font-size:12px; color: var(--text-muted);">
    """
    sorted_weights = sorted(weights_dict.items(), key=lambda x: int(x[1]), reverse=True)
    for role, score in sorted_weights:
        weight_visual_html += f"""
        <tr>
            <td style="padding: 4px 0;">{role}</td>
            <td style="text-align:right; font-weight:bold; color:var(--text-main);">{score}</td>
        </tr>
        """
    weight_visual_html += "</table></div>"

    # Tab navigation
    tabs_nav = '<div class="tabs-nav">'
    for i, t_name in enumerate(tables_js.keys()):
        active = "active" if i == 0 else ""
        tabs_nav += f'<button class="tab-btn {active}" onclick="switchTab(\'{t_name}\')">{t_name}</button>'
    tabs_nav += '</div>'

    # Assemble the page
    html = HTML_HEAD + f"""
    <style>{css_block}</style>
</head>
<body>
    <nav class="navbar">
        <div class="logo-area">
            <img src="{logo_url}" alt="SE">
            <div><h1>Data Manager</h1></div>
        </div>
        <div style="font-size: 13px; color: var(--text-muted); font-weight: 500;">
            {datetime.now().strftime('%d %b %Y, %H:%M')}
        </div>
    </nav>
    <div class="container">
        <aside>
            <div class="card">
                <div class="sidebar-title">Dataset Overview</div>
                <div class="stat-item"><span class="stat-label">Total Records</span><span class="stat-val" id="statTotal">0</span></div>
                <div class="stat-item"><span class="stat-label">Visible</span><span class="stat-val" style="color: var(--se-green);" id="statFiltered">0</span></div>
                <div class="stat-item"><span class="stat-label">Active Files</span><span class="stat-val" id="statFiles">0</span></div>
                
                <div class="sidebar-title" style="margin-top: 20px; margin-bottom: 5px;">Active Files</div>
                <div id="fileListContainer" class="file-list-container"></div>
            </div>
            {weight_visual_html}
        </aside>
        <main>
            <div class="card">
                <div class="search-wrapper">
                    <div class="input-group">
                        <select id="columnSelect">{col_opts}</select>
                        <input type="text" id="customSearch" placeholder="Type keywords to filter...">
                    </div>
                    <button class="btn btn-primary" id="addFilter">Apply</button>
                    <button class="btn btn-secondary" id="excludeBtn" title="Exclude"><i class="fas fa-ban"></i></button>
                    <button class="btn btn-secondary" id="clearAll" title="Reset"><i class="fas fa-undo"></i></button>
                </div>
                <div class="filter-logic-hint">
                    <i class="fas fa-lightbulb"></i> <strong>Tip:</strong> Filters in the SAME column are combined with <b>OR</b>. Filters in DIFFERENT columns are combined with <b>AND</b>.
                </div>
                <div class="filters-area" id="filtersContainer"><span style="color: var(--text-muted); font-size: 13px; padding-left: 5px;">No active filters</span></div>
                {analytics_html}
            </div>
            
            {tabs_nav}
            
            <div class="card" style="padding: 0; overflow: hidden; border-top-left-radius: 0;">
                 <div style="padding: 15px 25px; border-bottom: 1px solid var(--border-subtle); display: flex; justify-content: space-between; align-items: center; background: #fff;">
                    <div id="resultsInfo" style="font-weight: 600; font-size: 14px; color: var(--text-main);">Loading...</div>
                    <div>
                        <button class="btn btn-secondary" id="topExportBtn" style="padding: 6px 15px; font-size: 12px;">
                             <i class="fas fa-file-download"></i> Export Table
                        </button>
                    </div>
                 </div>
                <div style="padding: 0 25px 25px 25px;">
                    <table id="dataTable" class="display" style="width:100%"></table>
                </div>
            </div>
        </main>
    </div>
    
    <div id="detailsModal" class="modal"><div class="modal-content"><div class="modal-header"><h2>Record Details</h2><span class="close-btn" onclick="closeModal('detailsModal')">&times;</span></div><div class="modal-body" id="modalBody"></div></div></div>
    
    <div id="radarModal" class="modal">
        <div class="modal-content" style="max-width:1100px; height:85vh;">
            <div class="modal-header">
                <h2>Competitor Radar Comparison</h2>
                <div>
                    <button class="btn btn-secondary" style="margin-right: 15px; padding: 6px 15px; font-size: 12px;" onclick="exportInteractiveChart('echarts', radarChartInst, 'Interactive_Radar_Chart')"><i class="fas fa-file-download"></i> Export Interactive Chart</button>
                    <span class="close-btn" onclick="closeModal('radarModal')" style="cursor:pointer; font-size:24px;">&times;</span>
                </div>
            </div>
            <div class="modal-body" style="display: flex; flex-direction: column;">
                <div style="background: #f1f2f6; padding: 15px; border-radius: 8px; margin-bottom: 15px;">
                    <div style="font-size: 12px; font-weight: bold; color: #636e72; margin-bottom: 8px;">SELECT COMPANIES TO COMPARE (MAX 5):</div>
                    <div id="radarCompanyList" class="checkbox-list"></div>
                    <div style="margin-top: 10px; font-size: 11px; color: #636e72; font-style: italic;">
                        * The chart automatically selects the Top 8 Domains where the selected companies are most active combined.
                    </div>
                </div>
                <div id="radarChartContainer" style="flex-grow: 1; width: 100%; min-height: 400px;"></div>
            </div>
        </div>
    </div>

    <div id="networkModal" class="modal">
        <div class="modal-content">
            <div class="modal-header">
                <h2><span class="badge badge-green">Top 10</span> Map of Connections</h2>
                <span class="close-btn" onclick="closeModal('networkModal')" style="cursor:pointer; font-size:24px;">&times;</span>
            </div>
            <div class="modal-body" style="padding:0; overflow:hidden; display:flex; flex-direction:column;">
                <div style="padding: 15px; background: #fff; border-bottom: 1px solid #eee; display: flex; gap: 10px; align-items: center;">
                    <input list="networkDatalist" id="networkSearchInput" placeholder="Type to search Company or Group..." style="flex-grow: 1; padding: 10px; border: 1px solid #ddd; border-radius: 8px;">
                    <datalist id="networkDatalist"></datalist>
                    <button class="btn btn-primary" onclick="searchNetworkNode()">Search</button>
                    <button class="btn btn-secondary" onclick="resetNetworkMap()">Reset</button>
                    <button class="btn btn-secondary" onclick="exportInteractiveChart('echarts', networkChartInst, 'Interactive_Network_Map')"><i class="fas fa-file-download"></i> Export Viz</button>
                </div>
                <div id="networkMessage" style="position: absolute; top: 50%; left: 50%; transform: translate(-50%, -50%); text-align: center; color: #636E72; z-index: 10;">
                    <i class="fas fa-search" style="font-size: 3rem; margin-bottom: 10px; color: #dfe6e9;"></i>
                    <h3 style="margin: 0;">Map is empty</h3>
                    <p>Search for a Company or Group to visualize connections.</p>
                </div>
                <div id="networkContainer" style="width: 100%; height: 70vh;"></div>
            </div>
        </div>
    </div>
    
    <div id="geoModal" class="modal">
        <div class="modal-content" style="max-width:1300px; height:90vh;">
            <div class="modal-header">
                <h2>Geographic Heatmap</h2>
                <div>
                    <button class="btn btn-secondary" style="margin-right: 15px; padding: 6px 15px; font-size: 12px;" onclick="exportInteractiveChart('echarts', geoChartInst, 'Interactive_Geo_Heatmap')"><i class="fas fa-file-download"></i> Export Interactive Chart</button>
                    <span class="close-btn" onclick="closeModal('geoModal')" style="cursor:pointer; font-size:24px;">&times;</span>
                </div>
            </div>
            <div class="modal-body" style="display:flex; flex-direction:column; padding:15px;">
                <div id="geoLoader" class="loader"></div>
                <div id="geoContainer" style="width: 100%; flex-grow:1; min-height: 500px; border-radius: 12px; overflow: hidden;"></div>
            </div>
        </div>
    </div>
    
    <div id="powerModal" class="modal">
        <div class="modal-content" style="max-width:1000px; height:80vh;">
            <div class="modal-header">
                <h2>Company Power Rankings</h2>
                <div>
                    <button class="btn btn-secondary" style="margin-right: 15px; padding: 6px 15px; font-size: 12px;" onclick="exportInteractiveChart('chartjs', powerChartInst, 'Interactive_Power_Rankings')"><i class="fas fa-file-download"></i> Export Interactive Chart</button>
                    <span class="close-btn" onclick="closeModal('powerModal')" style="cursor:pointer; font-size:24px;">&times;</span>
                </div>
            </div>
            <div class="modal-body">
                <div style="height:300px; margin-bottom:20px;"><canvas id="powerChart"></canvas></div>
                <div id="powerTableContainer"></div>
            </div>
        </div>
    </div>

    <div id="landscapeModal" class="modal">
        <div class="modal-content" style="max-width:1100px; height:80vh;">
            <div class="modal-header">
                <h2>Gaps & Leads</h2>
                <span class="close-btn" onclick="closeModal('landscapeModal')" style="cursor:pointer; font-size:24px;">&times;</span>
            </div>
            <div class="modal-body">
                <div class="alert" style="background:#f1f2f6; padding:10px; border-radius:8px; margin-bottom:15px; font-size:12px; color:#636e72;">
                    <strong>Logic:</strong> Shows groups where the difference between SE Score and the Top Competitor Score is significant (Gap > SIGNIFICANT_GAP_THRESHOLD or Lead > SIGNIFICANT_GAP_THRESHOLD).
                </div>
                <div id="landscapeTableContainer"></div>
            </div>
        </div>
    </div>
    
    <div id="benchModal" class="modal">
        <div class="modal-content" style="max-width:1000px; height:80vh;">
            <div class="modal-header">
                <h2>Expert Workload Analysis</h2>
                <div>
                    <button class="btn btn-secondary" style="margin-right: 15px; padding: 6px 15px; font-size: 12px;" onclick="exportInteractiveChart('echarts', benchGeoChartInst, 'Interactive_Workload_Map')"><i class="fas fa-file-download"></i> Export Interactive Chart</button>
                    <span class="close-btn" onclick="closeModal('benchModal')" style="cursor:pointer; font-size:24px;">&times;</span>
                </div>
            </div>
            <div class="modal-body">
                <div style="margin-bottom: 20px; font-size:12px; color:#636e72; background:#f1f2f6; padding:10px; border-radius:8px;">
                      <strong>Metrics:</strong> "Strength" is the total weighted power score derived from all roles held by the expert.
                </div>
                <div id="benchGeoWrap" style="height:520px; margin-bottom:20px; border-radius:12px; overflow:hidden;">
                    <div id="benchGeoContainer" style="width:100%; height:100%;"></div>
                </div>
                <div id="benchTableContainer"></div>
            </div>
        </div>
    </div>
"""
    html += HTML_SCRIPTS + js_block + """
</body>
</html>
    """
    return html

def _get_dashboard_css():
    return """
        :root {
            --se-green: #3DCD58;
            --se-dark: #006039;
            --bg-body: #F5F7FA;
            --bg-card: #FFFFFF;
            --text-main: #2D3436;
            --text-muted: #636E72;
            --border-subtle: #EDF2F7;
            --shadow-sm: 0 2px 4px rgba(0,0,0,0.02);
            --shadow-md: 0 4px 12px rgba(0,0,0,0.05);
            --radius: 12px;
        }
        * { box-sizing: border-box; outline: none; }
        body { font-family: 'Inter', sans-serif; background: var(--bg-body); margin: 0; padding: 0; color: var(--text-main); -webkit-font-smoothing: antialiased; }
        .navbar { background: var(--bg-card); height: 70px; padding: 0 40px; display: flex; align-items: center; justify-content: space-between; box-shadow: var(--shadow-sm); position: sticky; top: 0; z-index: 100; }
        .logo-area { display: flex; align-items: center; gap: 15px; }
        .logo-area img { height: 35px; }
        .logo-area h1 { font-size: 18px; font-weight: 700; color: var(--se-dark); margin: 0; letter-spacing: -0.5px; }
        .container { max-width: 1600px; margin: 30px auto; padding: 0 30px; display: grid; grid-template-columns: 300px 1fr; gap: 30px; }
        .card { background: var(--bg-card); border-radius: var(--radius); box-shadow: var(--shadow-md); padding: 25px; margin-bottom: 25px; border: 1px solid var(--border-subtle); }
        .sidebar-title { font-size: 12px; font-weight: 700; color: var(--text-muted); text-transform: uppercase; margin-bottom: 15px; letter-spacing: 1px; }
        .stat-item { display: flex; justify-content: space-between; align-items: center; margin-bottom: 12px; }
        .stat-label { color: var(--text-muted); font-size: 14px; }
        .stat-val { font-weight: 700; font-size: 16px; color: var(--se-dark); }
        
        .file-list-container { max-height: 300px; overflow-y: auto; margin: 15px -10px 0 -10px; padding: 0 10px; }
        .file-item { 
            display: flex; justify-content: space-between; align-items: center; 
            padding: 8px 10px; margin-bottom: 4px; border-radius: 6px; 
            cursor: pointer; transition: background 0.2s; font-size: 13px; color: var(--text-main);
        }
        .file-item:hover { background-color: #f1f2f6; }
        .file-item.active { background-color: #e8f5e9; color: var(--se-dark); font-weight: 600; }
        .file-count { font-size: 11px; color: var(--text-muted); background: #dfe6e9; padding: 2px 6px; border-radius: 10px; }
        .file-item:hover .file-count { background: #fff; }

        
        .tabs-nav { display: flex; gap: 5px; margin-bottom: 0; padding-left: 10px; }
        .tab-btn { 
            padding: 10px 25px; background: #e0e0e0; border: none; 
            border-radius: 8px 8px 0 0; cursor: pointer; font-weight: 600; 
            color: #666; transition: 0.2s; font-size: 13px;
        }
        .tab-btn.active { 
            background: white; color: var(--se-dark); 
            border-top: 3px solid var(--se-green); 
            box-shadow: 0 -2px 5px rgba(0,0,0,0.05);
        }
        .tab-btn:hover:not(.active) { background: #d0d0d0; }

        .search-wrapper { display: flex; gap: 10px; margin-bottom: 10px; }
        .input-group { flex-grow: 1; display: flex; gap: 10px; }
        select, input { padding: 12px 16px; border: 1px solid #E2E8F0; border-radius: 8px; font-family: inherit; font-size: 14px; background: #FAFAFA; transition: 0.2s; }
        select:focus, input:focus { border-color: var(--se-green); background: #fff; box-shadow: 0 0 0 3px rgba(61, 205, 88, 0.1); }
        #columnSelect { width: 180px; }
        #customSearch { flex-grow: 1; }
        .analytics-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 15px; margin-top: 20px; }
        .action-card { background: #fff; border: 1px solid var(--border-subtle); border-radius: 10px; padding: 15px; display: flex; align-items: center; gap: 15px; cursor: pointer; transition: all 0.2s; }
        .action-card:hover { transform: translateY(-3px); box-shadow: var(--shadow-md); border-color: var(--se-green); }
        .icon-box { width: 40px; height: 40px; border-radius: 8px; display: flex; align-items: center; justify-content: center; font-size: 18px; color: white; }
        .bg-purple { background: #6c5ce7; }
        .bg-red { background: #e74c3c; }
        .bg-yellow { background: #f1c40f; }
        .bg-orange { background: #e67e22; }
        .bg-teal { background: #00b894; }
        .bg-blue { background: #0984e3; }
        .bg-green-dark { background: #2d3436; }
        .action-text h4 { margin: 0; font-size: 14px; font-weight: 600; color: var(--text-main); }
        .action-text p { margin: 2px 0 0 0; font-size: 11px; color: var(--text-muted); }
        .filters-area { margin-bottom: 20px; display: flex; flex-wrap: wrap; gap: 8px; min-height: 24px; }
        .chip { background: #E8F5E9; color: var(--se-dark); padding: 5px 12px; border-radius: 20px; font-size: 12px; font-weight: 600; display: flex; align-items: center; gap: 8px; }
        .chip i { cursor: pointer; opacity: 0.5; }
        .chip i:hover { opacity: 1; }
        table.dataTable { border-collapse: collapse !important; width: 100% !important; }
        table.dataTable thead th { background: white !important; color: var(--text-muted) !important; font-weight: 600 !important; font-size: 12px !important; text-transform: uppercase; border-bottom: 2px solid var(--border-subtle) !important; padding: 15px 10px !important; }
        table.dataTable tbody td { padding: 12px 10px !important; font-size: 13px; color: var(--text-main); border-bottom: 1px solid var(--border-subtle); }
        table.dataTable tbody tr:hover { background-color: #FAFAFA !important; }
        .dataTables_info, .dataTables_paginate { font-size: 12px; margin-top: 15px; }
        .paginate_button.current { background: var(--se-green) !important; color: white !important; border: none !important; border-radius: 4px; }
        .btn { border: none; padding: 10px 20px; border-radius: 8px; font-weight: 600; font-size: 13px; cursor: pointer; transition: 0.2s; display: inline-flex; align-items: center; justify-content: center; gap: 8px; }
        .btn:hover { opacity: 0.9; transform: translateY(-1px); }
        .btn-primary { background: var(--se-green); color: white; }
        .btn-secondary { background: white; border: 1px solid #CBD5E0; color: var(--text-main); }
        .btn-secondary:hover { background: #F7FAFC; }
        .btn-link { background: transparent; color: var(--se-green); padding: 0; }
        .full-width { width: 100%; }
        .mb-3 { margin-bottom: 15px; }
        .modal { display: none; position: fixed; z-index: 1000; left: 0; top: 0; width: 100%; height: 100%; background: rgba(0,0,0,0.4); backdrop-filter: blur(4px); }
        .modal-content { background: white; margin: 3vh auto; width: 95%; max-width: 1400px; max-height: 85vh; border-radius: 16px; overflow: hidden; display: flex; flex-direction: column; }
        .modal-header { padding: 20px 30px; border-bottom: 1px solid var(--border-subtle); display: flex; justify-content: space-between; align-items: center; }
        .modal-header h2 { margin: 0; font-size: 18px; color: var(--text-main); display: flex; align-items: center; gap: 10px; }
        .close-btn { font-size: 24px; color: var(--text-muted); cursor: pointer; }
        .modal-body { padding: 20px; overflow-y: auto; flex: 1; min-height: 0; background: #FBFCFD; position: relative; }
        .badge { padding: 4px 8px; border-radius: 6px; font-size: 11px; font-weight: 700; }
        .badge-green { background: #E8F5E9; color: var(--se-dark); }
        .badge-red { background: #FEE2E2; color: #991B1B; }
        .badge-gray { background: #dfe6e9; color: #636e72; }
        .empty-state { text-align: center; color: var(--text-muted); padding: 20px; background: #F7FAFC; border-radius: 8px; border: 1px dashed #CBD5E0; font-size: 13px; }
        .filter-logic-hint { font-size: 11px; color: #636e72; margin-bottom: 10px; font-style: italic; background: #f1f2f6; padding: 5px 10px; border-radius: 4px; display: inline-block; }
        .loader { display: none; border: 4px solid #f3f3f3; border-top: 4px solid #3DCD58; border-radius: 50%; width: 30px; height: 30px; animation: spin 1s linear infinite; margin: 20px auto; }
        
        .checkbox-list { display: grid; grid-template-columns: repeat(auto-fill, minmax(180px, 1fr)); gap: 8px; max-height: 150px; overflow-y: auto; padding: 10px; border: 1px solid #eee; border-radius: 8px; background: #fff; }
        .checkbox-item { display: flex; align-items: center; gap: 8px; font-size: 13px; padding: 4px; border-radius: 4px; }
        .checkbox-item:hover { background: #f1f2f6; }
        .checkbox-item input { margin: 0; padding: 0; width: 16px; height: 16px; }
        
        @keyframes spin { 0% { transform: rotate(0deg); } 100% { transform: rotate(360deg); } }
    """

def _get_dashboard_js(tables_js, is_std, export_fmt, weights_dict):
    
    weight_js = ""
    mapping = {
        "Secretary": "secretary", "Assistant Sec": "assistant sec", "Chair": "chair",
        "Vice-Chair": "vice-chair", "Liaison": "liaison", "Convenor": "convenor",
        "Nat. Part": "national participant", "Member": "member"
    }
    sorted_keys = sorted(mapping.keys(), key=lambda k: len(mapping[k]), reverse=True)
    
    for k in sorted_keys:
        search_str = mapping[k]
        val = weights_dict.get(k, "0")
        weight_js += f"if (func.includes('{search_str}')) return {val};\n"
    
    mem_val = weights_dict.get("Member", "5")
    weight_js += f"if (func.includes('nc official') || func.includes('observer')) return {mem_val};\n"

    return f"""
    <script>
    const tablesData = {json.dumps(tables_js)};
    const isStdMode = {str(is_std).lower()};
    const exportFormat = "{export_fmt}";
    
    const MAX_COMPANIES_COMPARE = {MAX_COMPANIES_COMPARE};
    const TOP_GROUPS = {CONFIG["TOP_GROUPS"]};
    const SIGNIFICANT_GAP_THRESHOLD = {CONFIG["SIGNIFICANT_GAP_THRESHOLD"]};
    
    let currentTableName = Object.keys(tablesData)[0];

    let fullData = [];
    let fullColumns = [];
    let sourceFileIndex = -1;
    
    let activeFilters = [];
    let filteredByTable = {{}};
    let filteredIndices = [];
    let table;
    
    let powerChartInst, benchChartInst, networkChartInst, geoChartInst, radarChartInst, benchGeoChartInst;
    let mapDataLoaded = false;
    let radarCompanies = [];
    let geoHeatmapMax = 0;
    let powerExportData = [], landscapeExportData = [], benchExportData = [];
    






    // FROM IEC WEBSITE so if there is changes in the participant visit IEC WEB to get the ISOCODE 
    const isoMap = {{
        "FR": "France", "US": "United States of America", "DE": "Germany", "CN": "China", "GB": "United Kingdom", "UK": "United Kingdom",
        "IN": "India", "JP": "Japan", "IT": "Italy", "ES": "Spain", "CA": "Canada", "AU": "Australia", "BR": "Brazil",
        "RU": "Russia", "KR": "South Korea", "MX": "Mexico", "ID": "Indonesia", "TR": "Turkey", "SA": "Saudi Arabia",
        "SE": "Sweden", "CH": "Switzerland", "NL": "Netherlands", "PL": "Poland", "BE": "Belgium", "AT": "Austria",
        "NO": "Norway", "DK": "Denmark", "FI": "Finland", "IE": "Ireland", "NZ": "New Zealand", "SG": "Singapore",
        "ZA": "South Africa", "EG": "Egypt", "MY": "Malaysia", "TH": "Thailand", "VN": "Vietnam", "PH": "Philippines",
        "AR": "Argentina", "CL": "Chile", "CO": "Colombia", "PE": "Peru", "UA": "Ukraine", "CZ": "Czech Republic",
        "RO": "Romania", "HU": "Hungary", "GR": "Greece", "PT": "Portugal", "IL": "Israel"
    }};




    Chart.defaults.font.family = "'Inter', sans-serif";
    Chart.defaults.color = '#636E72';

    function buildNameToIso() {{
        const m = {{}};
        Object.keys(isoMap).forEach(code => {{
            m[isoMap[code].toLowerCase()] = code;
        }});
        return m;
    }}
    const nameToIso = buildNameToIso();
    
    function initTable(tableName) {{
        currentTableName = tableName;
        const tData = tablesData[tableName];
        
        fullData = tData.data;
        fullColumns = tData.columns;
        sourceFileIndex = fullColumns.length - 1; 
        
        if (!filteredByTable[tableName]) {{
            filteredByTable[tableName] = fullData.map((_, idx) => idx);
        }}
        filteredIndices = filteredByTable[tableName];
        
        const sel = $('#columnSelect');
        sel.empty();
        sel.append('<option value="all">Global Search...</option>');
        fullColumns.forEach((c, i) => {{
            sel.append(`<option value="${{i}}">${{c.title}}</option>`);
        }});
        
        $('#statTotal').text(fullData.length.toLocaleString());
        
        updateTableRender();
        updateFileList();
    }}



    
    function switchTab(tableName) {{
        
        $('.tab-btn').removeClass('active');
       
        $(`.tab-btn`).filter(function() {{ return $(this).text() === tableName; }}).addClass('active');
        
        initTable(tableName);
    }}

    function updateTableRender() {{
        let maxCols = 5;

        
        let displayCols = fullColumns.slice(0, maxCols).map(c => ({{ title: c.title }}));
        displayCols.push({{ title: "Details" }});

        
        let tableData = filteredIndices.map(i => {{
            let row = fullData[i];
            let newRow = row.slice(0, maxCols);
            newRow.push(i); 
            return newRow;
        }});

        


if ($.fn.DataTable.isDataTable('#dataTable')) {{
    table.destroy();
    $('#dataTable').empty();
}}

table = $('#dataTable').DataTable({{
    data: tableData,
    columns: displayCols,
    pageLength: 20,
    searching: false,
    lengthChange: false,
    language: {{ emptyTable: "No data available" }},
    columnDefs: [{{
        targets: -1,
        render: (data) => `<button class="btn btn-link" onclick="showDetails(${{data}})">View</button>`
    }}]
}});

        $('#statFiltered').text(filteredIndices.length.toLocaleString());
        $('#resultsInfo').html(activeFilters.length === 0 ? `All Records` : `Found ${{filteredIndices.length}} matches`);
        renderFilters();
    }}

    function updateFileList() {{
        const container = $('#fileListContainer');
        container.empty();
        
    
        
        const currentCounts = {{}};
        
        

        filteredIndices.forEach(idx => {{ 
            const file = fullData[idx][sourceFileIndex]; 
            currentCounts[file] = (currentCounts[file] || 0) + 1;
        }});

        const tData = tablesData[currentTableName];
        const allFiles = tData.file_labels; 





        
        let activeFileCount = 0;

        allFiles.forEach(file => {{
            const count = currentCounts[file] || 0;
            if (count > 0) activeFileCount++;
            
            const isActive = activeFilters.some(f => f.colIndex === sourceFileIndex && f.value === file);
            
            const itemHtml = `
                <div class="file-item ${{isActive ? 'active' : ''}}" onclick="addFilter('${{file}}', ${{sourceFileIndex}}, 'Source File')">
                    <span style="white-space:nowrap; overflow:hidden; text-overflow:ellipsis; margin-right:5px;" title="${{file}}"><i class="far fa-file-excel"></i> ${{file}}</span>
                    <span class="file-count">${{count}}</span>
                </div>`;
            container.append(itemHtml);
        }});
        
        $('#statFiles').text(activeFileCount.toLocaleString());
    }}

    


    function applyFilters() {{
        const globalIncludeFilters = activeFilters.filter(f => f.type === 'all');        
        const globalExcludeFilters = activeFilters.filter(f => f.type === 'excludeAll');

        const tableIncludeFilters = activeFilters.filter(f => f.type === 'col' || f.type === 'nonEmpty'); 
        const tableExcludeFilters = activeFilters.filter(f => f.type === 'excludeCol');                  

        
    Object.keys(tablesData).forEach(tName => {{
        const tData = tablesData[tName];
        const data = tData.data;

        const out = [];
        data.forEach((row, rowIdx) => {{
            const rowStr = row.join(' ').toLowerCase();

            if (globalIncludeFilters.length > 0) {{
                const matchesInclude = globalIncludeFilters.some(f =>
                    rowStr.includes((f.value || "").toLowerCase())
                );
                if (!matchesInclude) return;
            }}

            if (globalExcludeFilters.length > 0) {{
                const matchesExclude = globalExcludeFilters.some(f =>
                    rowStr.includes((f.value || "").toLowerCase())
                );
                if (matchesExclude) return;
            }}

            out.push(rowIdx);
        }});

        filteredByTable[tName] = out;
    }});
    const tName = currentTableName;
    const data = tablesData[tName].data;

    let base = filteredByTable[tName] || data.map((_, idx) => idx);

    

    if (tableIncludeFilters.length > 0) {{
        const filtersByCol = {{}};

        tableIncludeFilters.forEach(f => {{
        const key = f.colIndex;
        if (!filtersByCol[key]) filtersByCol[key] = [];
        filtersByCol[key].push(f);
        }});

        const colKeys = Object.keys(filtersByCol);
        const out = [];

        base.forEach(rowIdx => {{
        const row = data[rowIdx];

        const matchesAllCols = colKeys.every(key => {{
            const group = filtersByCol[key];

            


            return group.some(f => {{
            const cellVal = (row[f.colIndex] || "").toString();
            if (f.type === 'nonEmpty') return cellVal.trim() !== "";
            return cellVal.toLowerCase().includes((f.value || "").toLowerCase());
        }});
      }});

      if (matchesAllCols) out.push(rowIdx);
    }});

    base = out; 
  }}

  if (tableExcludeFilters.length > 0) {{
    const excludeByCol = {{}};

    tableExcludeFilters.forEach(f => {{
      const key = f.colIndex;
      if (!excludeByCol[key]) excludeByCol[key] = [];
      excludeByCol[key].push(f);
    }});

    const colKeys = Object.keys(excludeByCol);
    const out = [];

    base.forEach(rowIdx => {{
      const row = data[rowIdx];

      const hitExclude = colKeys.some(key => {{
        const group = excludeByCol[key];
        return group.some(f => {{
          const cellVal = (row[f.colIndex] || "").toString();
          return cellVal.toLowerCase().includes((f.value || "").toLowerCase());
        }});
      }});

      if (!hitExclude) out.push(rowIdx);
    }});

    base = out;
  }}

  filteredByTable[tName] = base;
  filteredIndices = filteredByTable[currentTableName] || fullData.map((_, idx) => idx);

  updateTableRender();
  updateFileList();
}}


    function addFilter(val, colIndex = 'all', colName = 'All', isNonEmpty = false, isExclude = false) {{
  if(!isNonEmpty && (!val || val.trim() === '')) return;
  if(!isNonEmpty) val = val.trim();

  let type;
  if (isNonEmpty) type = 'nonEmpty';
  else if (isExclude && colIndex === 'all') type = 'excludeAll';
  else if (isExclude && colIndex !== 'all') type = 'excludeCol';
  else type = (colIndex === 'all' ? 'all' : 'col');

  const exists = activeFilters.some(f =>
    f.colIndex === colIndex &&
    f.type === type &&
    (type === 'nonEmpty' ? true : f.value === val)
  );
  if(exists) return;

  activeFilters.push({{
    type: type,
    value: isNonEmpty ? 'Not Blank' : val,
    colIndex: colIndex,
    colName: colName
  }});

  renderFilters();
  applyFilters();
  $('#customSearch').val('');
}}

    function removeFilter(index) {{ activeFilters.splice(index, 1); renderFilters(); applyFilters(); }}
    
    function renderFilters() {{
  const c = $('#filtersContainer').empty();
  if(activeFilters.length === 0) {{
    c.html('<span style="color: var(--text-muted); font-size: 13px; padding-left: 5px;">No active filters</span>');
    return;
  }}

  activeFilters.forEach((f, idx) => {{
    const isEx = (f.type === 'excludeAll' || f.type === 'excludeCol');
    let label;

    if (f.colIndex === 'all') {{
      label = isEx ? `Global: NOT "${{f.value}}"` : `Global: "${{f.value}}"`;
    }} else {{
      label = isEx
        ? `<strong>${{f.colName}}:</strong> NOT ${{f.value}}`
        : `<strong>${{f.colName}}:</strong> ${{f.value}}`;
    }}

    c.append(`<div class="chip">${{label}} <i class="fas fa-times-circle" onclick="removeFilter(${{idx}})"></i></div>`);
  }});
}}

    window.exportDataToExcel = function(headers, data, filename) {{
        if(!data || data.length === 0) {{ alert('No data to export'); return; }}
        if (exportFormat === 'csv') {{
            let csvContent = "data:text/csv;charset=utf-8,";
            csvContent += headers.join(",") + "\\r\\n";
            data.forEach(function(rowArray) {{
                let row = rowArray.join(",");
                csvContent += row + "\\r\\n";
            }});
            var encodedUri = encodeURI(csvContent);
            var link = document.createElement("a");
            link.setAttribute("href", encodedUri);
            link.setAttribute("download", filename + ".csv");
            document.body.appendChild(link);
            link.click();
        }} else {{
            const ws = XLSX.utils.aoa_to_sheet([headers].concat(data));
            const wb = XLSX.utils.book_new();
            XLSX.utils.book_append_sheet(wb, ws, "Data");
            XLSX.writeFile(wb, filename + ".xlsx");
        }}
    }};

    window.showDetails = function(rowIndex) {{
        let html = '<div style="display:grid; grid-template-columns: 1fr 1fr; gap:20px;">';
        fullColumns.forEach((col, i) => html += `<div style="border-bottom:1px solid #eee; padding-bottom:5px;"><div style="font-size:11px; color:#999; text-transform:uppercase;">${{col.title}}</div><div style="font-weight:500;">${{fullData[rowIndex][i]}}</div></div>`);
        html += '</div>';
        $('#modalBody').html(html); $('#detailsModal').show();
    }};

    

    
    function getWeight(func) {{
        func = (func || "").toLowerCase();
        {weight_js}
        return 0;
    }}

    function getIndices() {{
        const h = fullColumns.map(c => c.title.toLowerCase());
        return {{
            comp: h.findIndex(x => x.includes('company') || x.includes('organization') || x.includes('organi')),
            func: h.findIndex(x => x.includes('function') || x.includes('role')),
            grp: h.findIndex(x => x.includes('group') || x.includes('reference') || x.includes('committee') || x.includes('ref')),
            nc: h.findIndex(x => x === 'nc' || x.includes('national committee')),
            desc: h.findIndex(x => x === 'group name' || x === 'groupname' || x === 'title'),
            lname: h.findIndex(x => x.includes('last_name') || x.includes('lastname') || x === 'last name' || x.includes('surname') || x.includes('nom')),
            fname: h.findIndex(x => x.includes('first_name') || x.includes('firstname') || x === 'first name' || x.includes('prénom'))
        }};
    }}

    function getFullName(row, idxs) {{
        if (idxs.lname === -1 && idxs.fname === -1) return "Unknown Name";
        const l = (idxs.lname !== -1 ? (row[idxs.lname] || "") : "").trim();
        const f = (idxs.fname !== -1 ? (row[idxs.fname] || "") : "").trim();
        return (l + " " + f).trim() || "Unknown Name";
    }}
    
    function makeModalTable(id) {{
        if ($.fn.DataTable.isDataTable('#' + id)) $('#' + id).DataTable().destroy();
        $('#' + id).DataTable({{ pageLength: 10, searching: true, lengthChange: false, language: {{ search: "", searchPlaceholder: "Search table..." }} }});
    }}
    


    function initRadar() {{
        if(!isStdMode) return;
        const idxs = getIndices();
        if (idxs.comp === -1 || idxs.grp === -1) {{ alert("Missing Company or Group columns."); return; }}
        const compStats = {{}};
        filteredIndices.forEach(i => {{
            const row = fullData[i]; const comp = (row[idxs.comp]||"").trim(); const w = getWeight(row[idxs.func]);
            if(comp && comp.toLowerCase() !== 'unknown' && w > 0) {{ if(!compStats[comp]) compStats[comp] = 0; compStats[comp] += w; }}
        }});
        const sortedComps = Object.entries(compStats).sort((a,b) => b[1] - a[1]).map(x => x[0]);
        const listContainer = $('#radarCompanyList'); listContainer.empty();
        let schneiderSelected = false; 
        sortedComps.forEach(c => {{
            let isChecked = ''; const isSchneider = c.toLowerCase().includes('schneider') || c.toLowerCase() === 'se';
            if (isSchneider && !schneiderSelected) {{ isChecked = 'checked'; schneiderSelected = true; }}
            listContainer.append(`<label class="checkbox-item"><input type="checkbox" value="${{c}}" ${{isChecked}} onchange="updateRadarChart()"> ${{c}}</label>`);
        }});
        $('#radarModal').show();
        if(radarChartInst) radarChartInst.dispose();
        radarChartInst = echarts.init(document.getElementById('radarChartContainer'));
        window.addEventListener('resize', () => radarChartInst.resize());
        updateRadarChart();
    }}



    function updateRadarChart() {{
        const idxs = getIndices();
        const selected = []; $('#radarCompanyList input:checked').each(function() {{ selected.push($(this).val()); }});
        if(selected.length > MAX_COMPANIES_COMPARE) {{ alert("Please select maximum MAX_COMPANIES_COMPARE companies."); return; }}
        if(selected.length === 0) {{ radarChartInst.clear(); return; }}
        const groupData = {{}};
        filteredIndices.forEach(i => {{
            const row = fullData[i]; const comp = (row[idxs.comp]||"").trim(); const grp = (row[idxs.grp]||"").trim(); const w = getWeight(row[idxs.func]);
            if(selected.includes(comp) && grp) {{
                if(!groupData[grp]) groupData[grp] = {{ total: 0 }};
                if(!groupData[grp][comp]) groupData[grp][comp] = 0;
                groupData[grp][comp] += w; groupData[grp].total += w;
            }}
        }});



        const topGroups = Object.entries(groupData).sort((a,b) => b[1].total - a[1].total).slice(0, TOP_GROUPS).map(x => x[0]);
        const compColors = ['#e74c3c', '#3498db', '#f1c40f', '#9b59b6', '#e67e22']; const finalColors = []; let colorPtr = 0; const seriesData = [];
        selected.forEach(comp => {{
            if (comp.toLowerCase().includes('schneider') || comp.toLowerCase() === 'se') {{ finalColors.push('#3DCD58'); }} else {{ finalColors.push(compColors[colorPtr % compColors.length]); colorPtr++; }}
            seriesData.push({{ name: comp, value: [] }});
        }});
        let globalMax = 0;
        topGroups.forEach((grp, gIdx) => {{
            selected.forEach((comp, cIdx) => {{
                const val = groupData[grp][comp] || 0; seriesData[cIdx].value.push(val); if(val > globalMax) globalMax = val;
            }});
        }});
        const indicators = topGroups.map(g => ({{ name: g, max: globalMax * 1.1 }}));
        radarChartInst.setOption({{ color: finalColors, tooltip: {{ trigger: 'item' }}, legend: {{ data: selected, bottom: 0 }}, radar: {{ indicator: indicators, shape: 'polygon' }}, series: [{{ type: 'radar', data: seriesData, areaStyle: {{ opacity: 0.1 }} }}] }}, true);
    }}

    function showNetworkMap() {{
        if(!isStdMode) return;
        const idxs = getIndices();
        if (idxs.grp === -1 || idxs.comp === -1) {{ alert("Missing Group or Company columns."); return; }}
        globalNetworkEdges = []; globalNodeStats = {{}};
        let uniqueCompanies = new Set(); let uniqueGroups = new Set();

        filteredIndices.forEach(i => {{
            const row = fullData[i];
            let comp = (row[idxs.comp]||"").trim();
            const grp = (row[idxs.grp]||"").trim();
            const w = getWeight(row[idxs.func]);
            const person = getFullName(row, idxs);

            if(!comp || !grp || comp.toLowerCase() === 'unknown') return;
            
            globalNetworkEdges.push({{ c: comp, g: grp, w: w, p: person }});
            
            if(!globalNodeStats[comp]) globalNodeStats[comp] = 0; globalNodeStats[comp] += w;
            if(!globalNodeStats[grp]) globalNodeStats[grp] = 0; globalNodeStats[grp] += w;

            uniqueCompanies.add(comp); uniqueGroups.add(grp);
        }});

        

        if(globalNetworkEdges.length === 0) {{ alert("No data available for network map."); return; }}
        
        const dataList = $('#networkDatalist'); dataList.empty();
        let allItems = [];
        uniqueCompanies.forEach(c => allItems.push({{name: c, type: 'Company'}}));
        uniqueGroups.forEach(g => allItems.push({{name: g, type: 'Group'}}));
        allItems.sort((a,b) => a.name.localeCompare(b.name));
        allItems.forEach(item => {{ dataList.append(`<option value="${{item.name}}">${{item.type}}</option>`); }});

        $('#networkModal').show(); $('#networkMessage').show();
        
        const chartDom = document.getElementById('networkContainer');
        if(networkChartInst) networkChartInst.dispose();
        networkChartInst = echarts.init(chartDom);
        networkChartInst.setOption({{ title: {{ text: '' }}, series: [] }});
        window.addEventListener('resize', () => networkChartInst.resize());
        
        window.searchNetworkNode = function() {{
            const val = $('#networkSearchInput').val(); 
            if(!val) return;
            $('#networkMessage').hide();
            let subsetEdges = [];
            let isGroupSearch = uniqueGroups.has(val);
            let isCompanySearch = uniqueCompanies.has(val);
            if (isGroupSearch) {{
                let groupEdges = globalNetworkEdges.filter(e => e.g === val);
                groupEdges.sort((a,b) => b.w - a.w);
                subsetEdges = groupEdges;
            }} else if (isCompanySearch) {{
                subsetEdges = globalNetworkEdges.filter(e => e.c === val);
            }} else {{ 
                alert("Name not found in current view."); 
                return; 
            }}
            let localNodeStats = {{}};
            subsetEdges.forEach(e => {{
                if(!localNodeStats[e.c]) localNodeStats[e.c] = 0; 
                localNodeStats[e.c] += e.w;
                if(!localNodeStats[e.g]) localNodeStats[e.g] = 0; 
                localNodeStats[e.g] += e.w;
            }});





            let companyTopMembers = {{}}; 
            let groupTopCompanies = {{}}; 
            subsetEdges.forEach(e => {{
                if(e.p && e.p !== "Unknown Name") {{
                    if (!companyTopMembers[e.c]) companyTopMembers[e.c] = {{}};
                    if (!companyTopMembers[e.c][e.p]) companyTopMembers[e.c][e.p] = 0;
                    companyTopMembers[e.c][e.p] += e.w;
                }}
            }});


            let visibleGroups = new Set(subsetEdges.map(e => e.g));
            globalNetworkEdges.forEach(e => {{
                if (visibleGroups.has(e.g)) {{
                    if (!groupTopCompanies[e.g]) groupTopCompanies[e.g] = {{}};
                    if (!groupTopCompanies[e.g][e.c]) groupTopCompanies[e.g][e.c] = 0;
                    groupTopCompanies[e.g][e.c] += e.w;
                }}
            }});

            function getDescHtml(title, itemsMap) {{
                if(!itemsMap) return "No details available";
                let sorted = Object.entries(itemsMap).sort((a,b) => b[1] - a[1]).slice(0, 5);
                let html = `<div style="font-weight:bold; border-bottom:1px solid #777; margin-bottom:5px; padding-bottom:3px;">${{title}}</div>`;
                sorted.forEach(x => {{ html += `<div style="display:flex; justify-content:space-between; font-size:11px;"><span>${{x[0]}}</span><span style="font-weight:bold;">${{x[1]}}</span></div>`; }});
                if(sorted.length === 0) html += `<div style="font-size:11px;">No data</div>`;
                return html;
            }}


            let nodesMap = new Map(); 
            let links = [];

            

            subsetEdges.forEach(e => {{
                const isSE = e.c.toLowerCase().includes('schneider') || e.c.toLowerCase() === 'se';
                
                if(!nodesMap.has('C:'+e.c)) {{
                    let val = localNodeStats[e.c] || 1;
                    let size = Math.min(70, Math.max(20, Math.log(val + 1) * 15)); 
                    let tooltiphtml = getDescHtml("Top Members (In this View)", companyTopMembers[e.c]);
                    
                    nodesMap.set('C:'+e.c, {{ 
                        id: 'C:'+e.c, 
                        name: e.c, 
                        category: isSE ? 0 : 1, 
                        symbolSize: size, 
                        value: val, 
                        desc: tooltiphtml 
                    }});
                }}
                if(!nodesMap.has('G:'+e.g)) {{
                    let val = localNodeStats[e.g] || 1;
                    let size = Math.min(70, Math.max(20, Math.log(val + 1) * 15));
                    let tooltiphtml = getDescHtml("Top 5 Companies", groupTopCompanies[e.g]);
                    
                    nodesMap.set('G:'+e.g, {{ 
                        id: 'G:'+e.g, 
                        name: e.g, 
                        category: 2, 
                        symbolSize: size, 
                        value: val, 
                        desc:tooltiphtml 
                    }});
                }}
                
                links.push({{ source: 'C:'+e.c, target: 'G:'+e.g }});
            }});


            const nodes = Array.from(nodesMap.values());
            const option = {{
                tooltip: {{ 
                    trigger: 'item', 
                    formatter: function(params) {{
                        if (params.dataType === 'node') {{ 
                            return `<div style="text-align:left;"><strong>${{params.data.name}}</strong><br/><br/>${{params.data.desc}}</div>`; 
                        }}
                        return params.name; 
                    }} 
                }},
                legend: {{ data: [{{name: 'Schneider Electric'}}, {{name: 'Competitors'}}, {{name: 'Groups'}}] }},
                series: [{{
                    type: 'graph', 
                    layout: 'force', 
                    data: nodes, 
                    links: links,
                    categories: [
                        {{ name: 'Schneider Electric', itemStyle: {{ color: '#3DCD58' }} }}, 
                        {{ name: 'Competitors', itemStyle: {{ color: '#e74c3c' }} }}, 
                        {{ name: 'Groups', itemStyle: {{ color: '#2d3436' }} }}
                    ],
                    roam: true, 
                    label: {{ show: true, position: 'right' }}, 
                    force: {{ repulsion: 800, edgeLength: 150, gravity: 0.1 }}
                }}]
            }};

            networkChartInst.setOption(option, true);
        }};
        window.resetNetworkMap = function() {{ $('#networkSearchInput').val(''); $('#networkMessage').show(); networkChartInst.clear(); }};
    }}

    

    function calculatePower() {{
        if(!isStdMode) return;
        const idxs = getIndices();
        if (idxs.comp === -1 || idxs.func === -1) {{ alert("Missing columns."); return; }}
        const roleDefs = [{{ key: 'Secretary', color: '#6c5ce7' }}, {{ key: 'Assistant Sec', color: '#a29bfe' }}, {{ key: 'Chair', color: '#e17055' }}, {{ key: 'Vice-Chair', color: '#fab1a0' }}, {{ key: 'Convenor', color: '#ffeaa7' }}, {{ key: 'Liaison', color: '#74b9ff' }}, {{ key: 'Nat. Part', color: '#00cec9' }}, {{ key: 'Member', color: '#55efc4' }}, {{ key: 'Other', color: '#b2bec3' }}];
        let stats = {{}};
        filteredIndices.forEach(i => {{
            const row = fullData[i]; let comp = (row[idxs.comp]||"").trim();
            if(!comp) return;
            if(!stats[comp]) {{ stats[comp] = {{ score:0, breakdown: {{}} }}; roleDefs.forEach(r => stats[comp].breakdown[r.key] = 0); }}
            const funcStr = (row[idxs.func]||"").toLowerCase(); const w = getWeight(row[idxs.func]);
            let type = 'Other';
            if (funcStr.includes('secretary')) type = 'Secretary'; else if (funcStr.includes('assistant sec')) type = 'Assistant Sec'; else if (funcStr.includes('vice-chair')) type = 'Vice-Chair'; else if (funcStr.includes('chair')) type = 'Chair'; else if (funcStr.includes('convenor')) type = 'Convenor'; else if (funcStr.includes('liaison')) type = 'Liaison'; else if (funcStr.includes('national participant')) type = 'Nat. Part'; else if (funcStr.includes('nc official') || funcStr.includes('observer') || funcStr.includes('member')) type = 'Member';
            stats[comp].score += w; stats[comp].breakdown[type]++;
        }});
        const sorted = Object.entries(stats).sort((a,b) => b[1].score - a[1].score);
        powerExportData = sorted.map((item, i) => [i+1, item[0], item[1].score]);
        const top20 = sorted.slice(0, 20);
        let html = '<table id="powerTable" class="dataTable"><thead><tr><th>Rank</th><th>Company</th><th>Score</th><th>Action</th></tr></thead><tbody>';
        top20.forEach((item, i) => {{ html += `<tr><td>${{i+1}}</td><td><strong>${{item[0]}}</strong></td><td><span class="badge badge-green">${{item[1].score}}</span></td><td><button class="btn btn-link" onclick="addFilter('${{item[0].replace(/'/g, "\\\\'")}}', ${{idxs.comp}}, 'Company')">Filter</button></td></tr>`; }});
        let btnHtml = `<div style="text-align:right; margin-bottom:10px;"><button class="btn btn-secondary" onclick="exportDataToExcel(['Rank', 'Company', 'Score'], powerExportData, 'Power_Rankings')"><i class="fas fa-file-download"></i> Export Full List</button></div>`;
        $('#powerTableContainer').html(btnHtml + html + '</tbody></table>'); makeModalTable('powerTable');
        if(powerChartInst) powerChartInst.destroy();
        const datasets = roleDefs.map(def => ({{ label: def.key, data: top20.map(i => i[1].breakdown[def.key]), backgroundColor: def.color }}));
        powerChartInst = new Chart(document.getElementById('powerChart'), {{ type: 'bar', data: {{ labels: top20.map(i => i[0]), datasets: datasets }}, options: {{ responsive: true, maintainAspectRatio: false, scales: {{ x: {{ stacked: true }}, y: {{ stacked: true }} }} }} }} );
        $('#powerModal').show();
    }}







    function calculateGeo() {{
        if(!isStdMode) return;
        const idxs = getIndices();
        if (idxs.nc === -1) {{ alert("No 'NC' column."); return; }}

        $('#geoModal').show();

        if (geoChartInst) geoChartInst.dispose();
        geoChartInst = echarts.init(document.getElementById('geoContainer'));

        const countryStats = {{}};
        geoHeatmapMax = 0;

        filteredIndices.forEach(i => {{
            const row = fullData[i];
            const code = (row[idxs.nc]||"").trim().toUpperCase();
            if(!code || code.length > 3) return;

            const mapName = isoMap[code] || code;
            const w = getWeight(row[idxs.func]);

            if(!countryStats[mapName]) countryStats[mapName] = {{ total: 0, comps: {{}} }};
            countryStats[mapName].total += w;
            geoHeatmapMax = Math.max(geoHeatmapMax, countryStats[mapName].total);

            if(idxs.comp !== -1) {{
                const comp = (row[idxs.comp]||"").trim();
                if(comp && comp.toLowerCase() !== 'unknown') {{
                    if(!countryStats[mapName].comps[comp]) countryStats[mapName].comps[comp] = 0;
                    countryStats[mapName].comps[comp] += w;
                }}
            }}
        }});




        const mapSeriesData = [];
        Object.keys(countryStats).forEach(country => {{
            const d = countryStats[country];
            const top = Object.entries(d.comps).sort((a,b)=>b[1]-a[1]).slice(0,5);

            let top5Html = top.map(([c,s]) =>
                `<div><span style="color:#dfe6e9; font-size:11px;">●</span> ${{c}}: <strong>${{s}}</strong></div>`
            ).join("");

            if(!top5Html) top5Html = "<div>No company details</div>";

            mapSeriesData.push({{ name: country, value: d.total, top5: top5Html }});
        }});

        


        if (!mapDataLoaded) {{
            $('#geoLoader').show();
            geoChartInst.showLoading();

            $.getJSON('https://s3-us-west-2.amazonaws.com/s.cdpn.io/95368/world.json', function(mapJson) {{
                echarts.registerMap('world', mapJson);
                mapDataLoaded = true;
                geoChartInst.hideLoading();
                $('#geoLoader').hide();
                renderGeoMap(mapSeriesData);
            }}).fail(function() {{
                alert("Could not load map data (internet required).");
                geoChartInst.hideLoading();
                $('#geoLoader').hide();
            }});
        }} else {{
            renderGeoMap(mapSeriesData);
        }}
    }}

    



    function renderGeoMap(data) {{
        geoChartInst.setOption({{
            title: {{ text: 'Power Heatmap', left: 'center' }},
            tooltip: {{
                trigger: 'item',
                formatter: function(params) {{
                    if(!params.data) return params.name;
                    return `
                        <div style="font-size:14px; font-weight:bold; margin-bottom:5px;">${{params.name}}</div>
                        <div style="border-bottom:1px solid #555; padding-bottom:5px; margin-bottom:5px;">
                            Total Power: <span style="color:#3DCD58; font-weight:bold;">${{params.value}}</span>
                        </div>
                        <div style="font-size:12px; color:#ccc; margin-bottom:3px;">Top 5 Companies:</div>
                        ${{params.data.top5}}
                    `;
                }}
            }},
            visualMap: {{
                left: 20, bottom: 40, min: 0, max: geoHeatmapMax,
                text: ['High', 'Low'], calculable: true,
                inRange: {{ color: ['#e0f2f1', '#3DCD58', '#004d2e'] }}
            }},
            series: [{{
                type: 'map', mapType: 'world',
                roam: true,
                emphasis: {{ label: {{ show: true }} }},
                data: data
            }}]
        }}, true);

        window.addEventListener('resize', () => geoChartInst.resize());
    }}

    


    function calculateLandscape() {{
        if(!isStdMode) return;
        const idxs = getIndices();
        if (idxs.comp === -1 || idxs.grp === -1) {{ alert("Missing columns."); return; }}
        const grps = {{}};
        filteredIndices.forEach(i => {{
            const row = fullData[i]; const grp = (row[idxs.grp]||"").trim(); let comp = (row[idxs.comp]||"").trim(); const desc = idxs.desc !== -1 ? (row[idxs.desc]||"") : "";
            if(!grp) return;
            if(!grps[grp]) grps[grp] = {{ se:0, comps:{{}}, desc: desc }};
            let w = getWeight(row[idxs.func]);
            if(comp.toLowerCase().includes('schneider') || comp.toLowerCase() === 'se') grps[grp].se += w;
            else if(comp.length > 0) {{ if(!grps[grp].comps[comp]) grps[grp].comps[comp] = 0; grps[grp].comps[comp] += w; }}
        }});
        let items = [];
        Object.keys(grps).forEach(g => {{ 
            let maxC = 0, topC = "None"; Object.entries(grps[g].comps).forEach(([c, s]) => {{ if(s > maxC) {{ maxC = s; topC = c; }} }});
            let diff = grps[g].se - maxC;
            if (Math.abs(diff) >= SIGNIFICANT_GAP_THRESHOLD) items.push({{ grp: g, desc: grps[g].desc, competitor: topC, seScore: grps[g].se, cScore: maxC, diff: diff }});
        }});
        items.sort((a,b) => Math.abs(b.diff) - Math.abs(a.diff));
        landscapeExportData = items.map(item => [item.grp, item.desc, item.competitor, item.seScore, item.cScore, item.diff, (item.diff > 0 ? 'Dominating' : 'High Risk')]);
        let html = '<table id="landscapeTable" class="dataTable"><thead><tr><th>Group</th><th>Top Competitor</th><th>Score Comparison</th><th>Status</th></tr></thead><tbody>';
        items.forEach(item => {{ 
            let isWin = item.diff > 0; let statusStyle = isWin ? 'color:#3DCD58; font-weight:bold;' : 'color:#d63031; font-weight:bold;';
            html += `<tr><td><strong>${{item.grp}}</strong><br><span style="font-size:11px;color:#888">${{item.desc}}</span></td><td>${{item.competitor}}</td><td>SE: <strong>${{item.seScore}}</strong> vs ${{item.cScore}}</td><td style="${{statusStyle}}">${{isWin?'DOMINATING':'HIGH RISK'}} (${{(isWin?'+':'')+item.diff}})</td></tr>`; 
        }});
        let btnHtml = `<div style="text-align:right; margin-bottom:10px;"><button class="btn btn-secondary" onclick="exportDataToExcel(['Group', 'Description', 'Top Competitor', 'SE Score', 'Competitor Score', 'Gap', 'Status'], landscapeExportData, 'Competitive_Landscape')"><i class="fas fa-file-download"></i> Export Full List</button></div>`;
        $('#landscapeTableContainer').html(btnHtml + html + '</tbody></table>'); makeModalTable('landscapeTable'); $('#landscapeModal').show();
    }}

    function calculateBench() {{
    if(!isStdMode) return;
    const idxs = getIndices();
    if (idxs.lname === -1 && idxs.fname === -1) {{ alert("Missing Name columns."); return; }}
    if (idxs.nc === -1) {{ alert("Missing NC column."); return; }}
    if (idxs.comp === -1 || idxs.func === -1) {{ alert("Missing Company/Role columns."); return; }}

    function filterSchneiderAndCountry(code) {{
        addFilter('schneider', idxs.comp, 'Company');
       

        addFilter(code, idxs.nc, 'NC');
    }}

    const countryStats = {{}};
    let benchGeoMax = 0;

    filteredIndices.forEach(i => {{
        const row = fullData[i];
        const comp = (row[idxs.comp] || "").toLowerCase();
        if(!(comp.includes('schneider') || comp === 'se')) return;

        const code = (row[idxs.nc] || "").trim().toUpperCase();
        if(!code || code.length > 3) return;

        const mapName = isoMap[code] || code;
        const person = getFullName(row, idxs);
        const w = getWeight(row[idxs.func]);

        if(!countryStats[mapName]) countryStats[mapName] = {{ total: 0, people: {{}}, code: code }};
        countryStats[mapName].total += w;
        benchGeoMax = Math.max(benchGeoMax, countryStats[mapName].total);

        if(person && person !== "Unknown Name" && w > 0) {{
            if(!countryStats[mapName].people[person]) countryStats[mapName].people[person] = 0;
            countryStats[mapName].people[person] += w;
        }}
    }});

    


    const mapSeriesData = [];
    const personTotals = {{}};

    Object.keys(countryStats).forEach(countryName => {{
        const d = countryStats[countryName];

    
        
        const topPeople = Object.entries(d.people)
            .sort((a,b) => b[1] - a[1])
            .slice(0, 5);

        topPeople.forEach(([p, s]) => {{
            personTotals[p] = (personTotals[p] || 0) + s;
        }});

        let peopleHtml = topPeople.map(([p,s]) =>
            `<div style="display:flex; justify-content:space-between; gap:12px;">
                <span style="white-space:nowrap; overflow:hidden; text-overflow:ellipsis; max-width:260px;">${{p}}</span>
                <strong>${{s}}</strong>
             </div>`
        ).join("");

        if(!peopleHtml) peopleHtml = "<div>No people details</div>";

        
        mapSeriesData.push({{
            name: countryName,
            value: d.total,
            code: d.code,
            peopleHtml: peopleHtml
        }});
    }});

    const sortedPeople = Object.entries(personTotals).sort((a,b) => b[1] - a[1]).slice(0, 20);
    benchExportData = sortedPeople.map(([name, power]) => [name, power]);

    

    let html = `
      <table id="benchTable" class="dataTable">
        <thead><tr>
          <th>Expert Name</th>
          <th>Total Strength</th>
          <th>Action</th>
        </tr></thead>
        <tbody>
    `;

    sortedPeople.forEach(([name, power]) => {{
        const safe = name.replace(/'/g, "\\\\'");
        html += `
          <tr>
            <td><strong>${{name}}</strong></td>
            <td><strong>${{power}}</strong></td>
            <td>
              <button class="btn btn-link" onclick="addFilter('${{safe}}', 'all', 'All')">Search Person</button>
            </td>
          </tr>
        `;
    }});

    

    html += `</tbody></table>`;

    const btnHtml = `
      <div style="text-align:right; margin-bottom:10px;">
        <button class="btn btn-secondary"
          onclick="exportDataToExcel(['Expert Name','Total Strength'], benchExportData, 'Expert_Workload_SE')">
          <i class="fas fa-file-download"></i> Export Full List
        </button>
      </div>
    `;

    $('#benchTableContainer').html(btnHtml + html);
    makeModalTable('benchTable');

    


    $('#benchModal').show();

    if (benchGeoChartInst) benchGeoChartInst.dispose();
    benchGeoChartInst = echarts.init(document.getElementById('benchGeoContainer'));

    function renderBenchGeoMap() {{
        benchGeoChartInst.setOption({{
            title: {{ text: 'Schneider Expert Strength by Country (NC)', left: 'center' }},
            tooltip: {{
                trigger: 'item',
                formatter: function(params) {{
                    if(!params.data) return params.name;

                    return `
                        <div style="font-size:14px; font-weight:bold; margin-bottom:6px;">${{params.name}}</div>
                        <div style="margin-bottom:8px;">
                            Total Strength: <span style="color:#3DCD58; font-weight:bold;">${{params.value}}</span>
                        </div>
                        <div style="font-size:12px; color:#ccc; margin-bottom:4px;">Top People:</div>
                        ${{params.data.peopleHtml}}
                        <div style="margin-top:8px; font-size:11px; color:#aaa;"><i class="fas fa-mouse-pointer"></i> Double-click to filter</div>
                    `;
                }}
            }},
            visualMap: {{
                left: 20, bottom: 40,
                min: 0, max: benchGeoMax,
                text: ['High', 'Low'],
                calculable: true,
                inRange: {{ color: ['#e0f2f1', '#3DCD58', '#004d2e'] }}
            }},
            series: [{{
                type: 'map',
                map: 'world',
                roam: true,
                emphasis: {{ label: {{ show: true }} }},
                data: mapSeriesData
            }}]
        }}, true);

        benchGeoChartInst.off('dblclick');
        benchGeoChartInst.on('dblclick', function(params) {{
            if(!params.data) return;
            const code = params.data.code || (nameToIso[params.name.toLowerCase()] || "");
            if(code) filterSchneiderAndCountry(code);
        }});

        window.addEventListener('resize', () => benchGeoChartInst.resize());
    }}

    if (!mapDataLoaded) {{
        benchGeoChartInst.showLoading();
        $.getJSON('https://s3-us-west-2.amazonaws.com/s.cdpn.io/95368/world.json', function(mapJson) {{
            echarts.registerMap('world', mapJson);
            mapDataLoaded = true;
            benchGeoChartInst.hideLoading();
            renderBenchGeoMap();
        }}).fail(function() {{
            benchGeoChartInst.hideLoading();
            alert("Could not load map data (internet required).");
        }});
    }} else {{
        renderBenchGeoMap();
    }}
}}

    function closeModal(id) {{ document.getElementById(id).style.display = 'none'; }}
    window.onclick = (e) => {{ if($(e.target).hasClass('modal')) $('.modal').hide(); }};

    //INTERACTIVE CHART EXPORT
    window.exportInteractiveChart = function(engine, chartInstance, filename) {{
        if (!chartInstance) {{
            alert("Chart is not fully loaded yet.");
            return;
        }}

        let chartConfigStr = "";
        let libraryScript = "";
        let initScript = "";

        // ECharts
        if (engine === 'echarts') {{
            const currentOption = chartInstance.getOption();
            chartConfigStr = JSON.stringify(currentOption);
            libraryScript = '<script src="https://cdn.jsdelivr.net/npm/echarts@5.4.3/dist/echarts.min.js"></' + 'script>';
            
            
            let tooltipPatch = "";
            if (filename.includes('Network')) {{
                tooltipPatch = `option.tooltip = {{ trigger: 'item', formatter: function(params) {{ if (params.dataType === 'node') return "<div style='text-align:left;'><strong>" + params.data.name + "</strong><br/><br/>" + (params.data.desc || "") + "</div>"; return params.name; }} }};`;
            }} else if (filename.includes('Geo') || filename.includes('Workload')) {{
                tooltipPatch = `option.tooltip = {{ trigger: 'item', formatter: function(params) {{ if(!params.data) return params.name; return "<div style='font-size:14px; font-weight:bold; margin-bottom:5px;'>" + params.name + "</div><div style='border-bottom:1px solid #555; padding-bottom:5px; margin-bottom:5px;'>Total Score: <span style='color:#3DCD58; font-weight:bold;'>" + params.value + "</span></div>" + (params.data.top5 || params.data.peopleHtml ? "<div style='font-size:12px; color:#ccc; margin-bottom:3px;'>Top Entities:</div>" + (params.data.top5 || params.data.peopleHtml) : ""); }} }};`;
            }}

            const isMap = chartConfigStr.includes('"map":"world"') || chartConfigStr.includes('"mapType":"world"');
            
            if (isMap) {{
                initScript = "var myChart = echarts.init(document.getElementById('chart-container')); " +
                             "myChart.showLoading(); " +
                             "fetch('https://s3-us-west-2.amazonaws.com/s.cdpn.io/95368/world.json').then(r => r.json()).then(mapJson => {{ " +
                             "    echarts.registerMap('world', mapJson); " +
                             "    myChart.hideLoading(); " +
                             "    var option = " + chartConfigStr + "; " +
                             "    " + tooltipPatch + " " +
                             "    myChart.setOption(option, true); " +
                             "    window.addEventListener('resize', () => myChart.resize()); " +
                             "}});";
            }} else {{
                initScript = "var myChart = echarts.init(document.getElementById('chart-container')); " +
                             "var option = " + chartConfigStr + "; " +
                             "    " + tooltipPatch + " " +
                             "myChart.setOption(option, true); " +
                             "window.addEventListener('resize', () => myChart.resize());";
            }}
        }} 
        // Chart.js
        else if (engine === 'chartjs') {{
            const safeConfig = {{
                type: 'bar',
                data: chartInstance.data,
                options: {{ responsive: true, maintainAspectRatio: false, scales: {{ x: {{ stacked: true }}, y: {{ stacked: true }} }} }}
            }};
            chartConfigStr = JSON.stringify(safeConfig);
            libraryScript = '<script src="https://cdn.jsdelivr.net/npm/chart.js"></' + 'script>';
            

            initScript = "var container = document.getElementById('chart-container'); " +
                         "var wrapper = document.createElement('div'); " +
                         "wrapper.style.width = '90vw'; wrapper.style.height = '85vh'; wrapper.style.position = 'relative'; wrapper.style.marginTop = '80px'; " +
                         "var canvas = document.createElement('canvas'); " +
                         "wrapper.appendChild(canvas); " +
                         "container.appendChild(wrapper); " +
                         "var ctx = canvas.getContext('2d'); " +
                         "new Chart(ctx, " + chartConfigStr + ");";
        }}

        const displayTitle = filename.replace(/_/g, ' ');

        const htmlContent = "<!DOCTYPE html>\\n" +
        "<html lang='en'>\\n" +
        "<head>\\n" +
        "    <meta charset='UTF-8'>\\n" +
        "    <meta name='viewport' content='width=device-width, initial-scale=1.0'>\\n" +
        "    <title>" + displayTitle + "</title>\\n" +
        "    " + libraryScript + "\\n" +
        "    <link href='https://fonts.googleapis.com/css2?family=Inter:wght@400;600&display=swap' rel='stylesheet'>\\n" +
        "    <style>\\n" +
        "        body {{ margin: 0; padding: 0; font-family: 'Inter', sans-serif; background: #FBFCFD; overflow: hidden; }}\\n" +
        "        #chart-container {{ width: 100vw; height: 100vh; display: flex; justify-content: center; align-items: center; padding: 30px; box-sizing: border-box; }}\\n" +
        "        .header {{ position: absolute; top: 15px; left: 15px; z-index: 10; background: rgba(255,255,255,0.95); padding: 12px 20px; border-radius: 8px; box-shadow: 0 4px 12px rgba(0,0,0,0.1); border: 1px solid #EDF2F7; }}\\n" +
        "        .header h3 {{ margin: 0; color: #2D3436; font-size: 16px; }}\\n" +
        "        .header p {{ margin: 4px 0 0 0; color: #636E72; font-size: 12px; }}\\n" +
        "    </style>\\n" +
        "</head>\\n" +
        "<body>\\n" +
        "    <div class='header'>\\n" +
        "        <h3>" + displayTitle + "</h3>\\n" +
        "        <p>Interactive View - Data Visualization snapshot captured from Data Manager.</p>\\n" +
        "    </div>\\n" +
        "    <div id='chart-container'></div>\\n" +
        "    <script>\\n" +
        "        " + initScript + "\\n" +
        "    </" + "script>\\n" +
        "</body>\\n" +
        "</html>";

        const blob = new Blob([htmlContent], {{ type: 'text/html' }});
        const url = URL.createObjectURL(blob);
        const a = document.createElement('a');
        a.href = url;
        a.download = filename + '.html';
        document.body.appendChild(a);
        a.click();
        document.body.removeChild(a);
        URL.revokeObjectURL(url);
    }};

    $(document).ready(function() {{
        initTable(currentTableName);
        
        $('#addFilter').click(() => {{ const v = $('#customSearch').val(), i = $('#columnSelect').val(), n = $('#columnSelect option:selected').text(); i === 'all' ? addFilter(v, 'all') : addFilter(v, parseInt(i), n); }});
        $('#excludeBtn').click(() => {{
  const v = $('#customSearch').val();
  const i = $('#columnSelect').val();
  const n = $('#columnSelect option:selected').text();

  if (i === 'all') addFilter(v, 'all', 'All', false, true);
  else addFilter(v, parseInt(i), n, false, true);
}});

        $('#customSearch').keypress(e => {{ if(e.which === 13) $('#addFilter').click(); }});
        $('#clearAll').click(() => {{ activeFilters = []; renderFilters(); applyFilters(); }});
        
        $('#topExportBtn').click(() => {{
             exportDataToExcel(fullColumns.map(c=>c.title), filteredIndices.map(i=>fullData[i]), currentTableName.replace(/ /g, "_") + "_Export");
        }});
        
        if(isStdMode) {{
            $('#powerBtn').click(calculatePower); 
            $('#radarBtn').click(initRadar); 
            $('#landscapeBtn').click(calculateLandscape); 
            $('#benchBtn').click(calculateBench); 
            $('#networkBtn').click(showNetworkMap);
            $('#geoBtn').click(calculateGeo);
        }}
    }});
    </script>
    """

if __name__ == "__main__":
    app = ctk.CTk()
    gui = App(app)
    app.mainloop()