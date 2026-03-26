"""
Data Manager (Internal)
Developed by: Sean
Primary users: Internal SE teams for committee list processing and competitive landscape analysis

Changelog:
- v3.2: Original committee data tool (single-table exports)
- v4.0: UI refresh + dashboard export 
- v5.8: multi-structure detection
- v6.4: RapidFuzz grouping + threshold slider (tuned on Q4 exports)
- v7.1: Dashboard analytics modals + export improvements + network viz
- v8.0: Community template support (Microsoft form export)
- v8.8: Chart export + UI polish (Patrick's feedback)
- v9.4: Community mode redesign 
- v9.9: Community mode performance optimizations
- v10.3: Tutorial addition + voting

Known Issues:
- Keep openpyxl read_only=True (large files spike memory otherwise, learned this the hard way)
- 'SESA' is treated as required identifier in Community mode (legacy template format)
- Fuzzy threshold of 0.85 was tuned on 2025-Q4 committee exports; don't change without re-testing

TODO:
- finalize_export rebuilds tables by re-reading all files (acceptable for now but slow on 50+ files)
- Consider caching parsed files to avoid double-reads

Performance Notes:
- Tested with 45+ files (~10MB total) → ~3s load time on typical laptops (T14S)
- Dashboard generation adds ~2s for large datasets
"""

import customtkinter as ctk
import os
import glob
import sys
import threading
import webbrowser
import json
import time
from datetime import datetime
import tkinter as tk
from tkinter import messagebox, scrolledtext, filedialog
import openpyxl
from rapidfuzz import fuzz  


# CONFIGURATION & CONSTANTS

DEBUG = False  # Set to True to see detailed header/column logs during file processing

MAX_COMPANIES_COMPARE = 5

# Tuned on 2025-Q4 committee exports:
# - 0.90 missed too many merges ("Schneider Electric" vs "Schneider Electric Japan")
# - 0.80 merged unrelated short names ("Schneider" merged with random "Schindler" entities)
# - 0.85 is the sweet spot but still requires manual review on edge cases
DEFAULT_FUZZY_THRESHOLD = 0.85

CONFIG = {
    # Colors (green theme approved by Martial & Patrick)
    "BRAND": {
        "MAIN_COLOR": "#3DCD58", 
        "ERROR_COLOR": "#c0392b",
    },


    # Search synonyms for dashboard filters (by Pauline's request, can be expanded as needed this one for sustainability)
    "SEARCH_SYNONYMS": {
        "tc 111": ["sustainability", "environmentally conscious design", "ecd", "environmental impact assessment", "chemical", "substance", "material declaration", "environmental declaration", "circular economy", "circular", "recycled", "reused", "renewable", "materials", "lifetime", "durability", "recovery", "recoverability", "end-of-life", "eol", "climate change", "resource depletion", "biodiversity"],
        "tc 111/jahg": ["digitalization", "digitization", "data", "exchange", "carbon", "footprint", "co2", "impact", "calculation"],
        "tc 111/jwg 14": ["test", "method", "substances", "plastics"],
        "tc 111/jwg 16": ["communication"],
        "tc 111/jwg ecd-62430": ["ecodesign", "eco-design", "assessment", "guide", "material efficiency", "circular", "circular economy"],
        "tc 111/mt 20": ["ecodesign", "eco-design", "material efficiency", "circular", "circular economy", "environmentally", "conscious", "design"],
        "tc 111/mt 21": ["glossary", "terms", "definitions", "iev", "dictionary"],
        "tc 111/mt 63000": ["material", "declaration", "restricted", "substance"],
        "tc 111/sdbt 62474": ["material", "declaration", "communication"],
        "tc 111/wg 3": ["substances", "benzotriazole", "ultraviolet", "plastics", "gas chromatography-mass spectrometry", "gcms", "gc-ms", "tetrabromobisphenol a (tbbpa)", "spectrometry", "lc-ms", "lc-dad", "polycyclic aromatic hydrocarbons (pahs)", "polymer", "phthalates", "pyrolyzer", "thermal desorption", "py-td-gc-ms", "screening", "polybrominated biphenyls", "aas", "afs", "icp-oes", "icp-ms", "chlorinated paraffins", "lead", "mercury", "cadmium", "chromium", "bromine", "phosphorus", "chlorine", "tin", "antimony", "fluorescence"],
        "tc 111/wg 5": ["assessment", "circular content", "recycled materials", "recycled", "materials", "reuse", "re-use", "second", "hand", "part", "component"],
        "tc 111/wg 15": ["product category rules", "pcr", "lca", "electrical", "electronic", "products", "systems", "eee"],
        "tc 111/wg 17": ["ghg", "greenhouse gas", "saved", "avoided", "emissions", "quantification", "communication"],
        "tc 111/wg 18": ["sustainable management", "waste", "electrical electronic equipment", "eee", "e-waste", "assessment", "material recyclability", "recyclability rate", "material recoverability", "recoverability rate"],
        "tc 111/wg 19": ["material", "declaration", "communication"],
        "tc 111/wg 23": ["assessment", "durability", "repair", "reuse", "upgrade"],
        "tc 111/wg 111/855/np": ["digitalization", "digitization", "carbon", "footprint", "cfp", "format", "guidance", "exchanging", "data", "classes", "properties", "identification", "common data dictionary", "cdd"],
        "tc 111/jwg24": ["digitalization", "digitization", "carbon", "footprint", "cfp", "data", "classes", "properties", "identification", "common data dictionary", "cdd"],
        "jtc 10": ["ecodesign", "eco-design", "assessment", "harmonized", "harmonization"],
        "jtc 10/wg 01": ["terminology", "material efficiency", "ecodesign", "eco-design", "environmentally", "conscious", "design"],
        "jtc 10/wg 02": ["durability", "assessment", "upgradability"],
        "jtc 10/wg 03": ["repairability", "reusability", "re-use", "reuse", "upgradability", "second", "hand", "dismantle"],
        "jtc 10/wg 04": ["remanufacturing", "refurbish", "re-manufacture", "second", "hand"],
        "jtc 10/wg 05": ["recyclability", "recoverability", "recycled", "content", "end-of-life", "eol", "components", "re-use", "reuse"],
        "jtc 10/wg 06": ["declaration", "communication"],
        "jtc 10/wg 08": ["ecodesign", "eco-design", "circular economy"],
        "tc 111x": ["environmental", "impacts", "reducing", "social", "economic", "safety", "performance", "requirement", "aspect"],
        "tc 111x/wg 01": ["environmental", "database", "data"],
        "tc 111x/wg 02": ["data", "digitalization", "digitization", "environmental"],
        "tc 111x/wg 03": ["continuous improvement", "environmental performance", "ecodesign", "eco-design"],
        "tc 111x/wg 04": ["eol", "end-of-life", "end of life", "treatment", "collection", "transportation", "storage", "handling"],
        "tc 111x/wg 05": ["materials", "communication", "restrictions", "rohs", "evaluation", "assessment", "substance", "eee"],
        "tc 111x/wg 06": ["weee", "recycling", "standards", "reuse", "re-use", "waste", "e-waste", "logistic", "treatment", "collection", "de-pollution", "end-of-life", "eol"],
        "tc 111x/wg 07": ["reuse", "re-use", "waste", "e-waste", "logistic", "treatment", "collection", "second", "hand", "de-pollution", "end-of-life", "eol", "recycling", "recovery", "recycle", "recover", "component"],
        "tc 111x/wg 08": ["lca", "environmentally conscious design", "ecodesign", "eco-design", "environmental impact", "reduction"],
        "tc 111x/wg 09": ["marking", "eee", "waste", "e-waste", "disposal", "collection", "recovery"],
        "tc 111x/wg 10": ["sbp", "update"],
        "tc 111x/wg 11": ["reuse", "material efficiency", "critical raw", "treatment", "recycle"],
        "tc 111x/wg 12": ["material", "recycle", "treatment"],
        "tc 22x/wg 9": ["material efficiency", "circular economy", "ecodesign", "eco-design", "psr", "environmental assessment", "environmental impact", "carbon footprint", "declaration", "repair", "repairability", "reuse", "reusability", "re-use", "upgrade", "upgradability", "dismantle", "recycle", "recover", "recyclability", "recoverability", "recycled content", "end-of-life", "eol", "components", "durability", "remanufacturing", "refurbish", "re-manufacture", "second", "hand", "communication", "power electronics"]
    },

    # Role weights (approved by Martial & Patrick, last reviewed Jan 2025)
    "ROLE_WEIGHTS": {
        "Secretary": "75",
        "Assistant Sec": "50",
        "Chair": "60",
        "Vice-Chair": "40",
        "Liaison": "15",
        "Convenor": "20",
        "Nat. Part": "5",
        "Member": "5",
    },

    # Analytics settings
    "TOP_GROUPS": 8,  # Dashboard default limit, I choose 8 because too for more than 8 it looks messy and unreadable
    "SIGNIFICANT_GAP_THRESHOLD": 15,  # When changed to 20 almost nothing shows up , when changed to 5 almost everything shows up, 15 is a good balance for now based on 2025-Q4 data

    # Fuzzy grouping - common internal shorthand patterns
    # Schneider has tons of variants: "Schneider Electric", "Schneider Electric Espana", "Schneider Electric Indonesia" etc.
    # Kept explicit because fuzzy matching alone is inconsistent with very short tokens
    "CANONICAL_ALIASES": {
        "schneider electric": ["schneider", "schneider electric", "schneider-elec", "schneider elect."],
    },

    "PRINT_TRACEBACKS": False,  # yes this is a bool, don't change, backend logging uses it

    # FROM IEC WEBSITE so if there is changes in the participant visit IEC WEB to get the ISOCODE (I dont think there will be changes but just in case)
    "ISO_MAP": {
        "AF": "Afghanistan", "AL": "Albania", "DZ": "Algeria", "AO": "Angola",
        "AR": "Argentina", "AU": "Australia", "AT": "Austria", "AZ": "Azerbaijan",
        "BD": "Bangladesh", "BE": "Belgium", "BJ": "Benin", "BR": "Brazil",
        "BG": "Bulgaria", "CA": "Canada", "CL": "Chile", "CN": "China",
        "CO": "Colombia", "HR": "Croatia", "CZ": "Czech Republic", "DK": "Denmark",
        "EG": "Egypt", "FI": "Finland", "FR": "France", "DE": "Germany",
        "GH": "Ghana", "GR": "Greece", "HU": "Hungary", "IN": "India",
        "ID": "Indonesia", "IR": "Iran", "IE": "Ireland", "IL": "Israel",
        "IT": "Italy", "JP": "Japan", "KZ": "Kazakhstan", "KE": "Kenya",
        "KR": "South Korea", "MY": "Malaysia", "MX": "Mexico", "MA": "Morocco",
        "NL": "Netherlands", "NZ": "New Zealand", "NG": "Nigeria", "NO": "Norway",
        "PK": "Pakistan", "PE": "Peru", "PH": "Philippines", "PL": "Poland",
        "PT": "Portugal", "RO": "Romania", "RU": "Russia", "SA": "Saudi Arabia",
        "RS": "Serbia", "SG": "Singapore", "SK": "Slovakia", "ZA": "South Africa",
        "ES": "Spain", "SE": "Sweden", "CH": "Switzerland", "TH": "Thailand",
        "TN": "Tunisia", "TR": "Turkey", "UA": "Ukraine", "GB": "United Kingdom",
        "UK": "United Kingdom", "US": "United States of America", "UZ": "Uzbekistan",
        "VN": "Vietnam",
    },
}

C_MAIN = CONFIG["BRAND"]["MAIN_COLOR"]
C_ERR = CONFIG["BRAND"]["ERROR_COLOR"]

# CUSTOM EXCEPTIONS

class DataManagerError(Exception):

    def __init__(self, user_message: str, detail: str | None = None):
        super().__init__(user_message)
        self.user_message = user_message
        self.detail = detail




class ColumnMismatchError(DataManagerError):
    pass


class ExportError(DataManagerError):
    pass


# UI SETUP

ctk.set_appearance_mode("light")
ctk.set_default_color_theme("green")

# UTILITY HELPERS

def clean_cell_value(val):
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    v = str(val).strip().replace('\xa0', ' ').replace('\n', ' ')
    if v in ["#VALUE!", "#REF!", "#N/A", "#DIV/0!"]:
        return ""
    return v

def safe_int(value, default=0):
    try:
        return int(value)
    except (ValueError, TypeError):
        return default

def get_stripped_text(widget):
    return widget.get().strip()

def extract_headers(ws, header_rows=1):
    raw_headers = []
    for r in ws.iter_rows(min_row=1, max_row=header_rows, values_only=True):
        raw_headers.append(r)
    
    if not raw_headers:
        return []
    
    local_cols = []
    max_c = len(raw_headers[0])
    
    for c_idx in range(max_c):
        parts = []
        for r_idx in range(header_rows):
            if r_idx < len(raw_headers) and c_idx < len(raw_headers[r_idx]):
                v = raw_headers[r_idx][c_idx]
                if v:
                    parts.append(str(v).strip().replace('\n', ' '))
        local_cols.append(" - ".join(parts) if parts else f"Column_{c_idx+1}")
    
    if DEBUG:
        print(f"[DEBUG] Extracted {len(local_cols)} headers from {header_rows} row(s)")
    
    return local_cols

def extract_form_sub_label(header: str, base_phrases: list[str]) -> str:
    h = (header or "").strip().replace("\n", " ")
    hl = h.lower()

    if "[" in h and "]" in h:
        left = h.find("[")
        right = h.rfind("]")
        if right > left:
            return h[left+1:right].strip()

    for bp in base_phrases:
        bpl = bp.lower()
        pos = hl.find(bpl)
        if pos != -1:
            dot = h.find(".", pos + len(bp))
            if dot != -1:
                return h[dot+1:].strip(" .:-\t")

    for sep in [" - ", ":", "–", "—"]:
        if sep in h:
            return h.split(sep, 1)[1].strip()
    return h


# UI COMPONENTS


class tooltip:

    def __init__(self, widget, msg):
        self.widget = widget
        self.msg = msg
        self.win = None
        self._after_id = None
        self.widget.bind("<Enter>", self.on_enter)
        self.widget.bind("<Leave>", self.on_leave)

    def on_enter(self, event=None):
        # Cancel any previous pending show
        if self._after_id:
            self.widget.after_cancel(self._after_id)
        self._after_id = self.widget.after(500, lambda e=event: self._show(e))

    def _show(self, event=None):
        self._after_id = None
        if self.win:
            return
        try:
            # Use the cursor's screen coordinates instead of bbox("insert")
            # which is unreliable on non-text widgets and causes flickering
            x = self.widget.winfo_rootx() + 20
            y = self.widget.winfo_rooty() + self.widget.winfo_height() + 4

            self.win = tk.Toplevel(self.widget)
            self.win.wm_overrideredirect(True)
            self.win.wm_geometry(f"+{x}+{y}")
            self.win.wm_attributes("-topmost", True)

            lbl = tk.Label(self.win, text=self.msg, justify='left',
                           background="#2d3436", fg="white", relief='solid', borderwidth=0,
                           font=("Segoe UI", 9))
            lbl.pack(ipadx=8, ipady=5)
        except Exception:
            return

    def on_leave(self, event=None):
        # Cancel pending show if mouse leaves before delay fires
        if self._after_id:
            self.widget.after_cancel(self._after_id)
            self._after_id = None
        if self.win:
            self.win.destroy()
            self.win = None




class TableStructureNamerDialog(ctk.CTkToplevel):
    def __init__(self, parent, buckets_info, callback):
        super().__init__(parent)
        self.callback = callback
        self.buckets_info = buckets_info 
        self.result = None
        
        self.title("Identify Tables")
        self.geometry("600x500")
        self.grab_set()
        self.protocol("WM_DELETE_WINDOW", self.on_cancel)
        self.bind("<Escape>", lambda e: self.on_cancel())
        

        ctk.CTkLabel(self, text="Different file formats detected", font=("Segoe UI", 16, "bold")).pack(pady=10)
        ctk.CTkLabel(self, text="Give each structure a tab name (what you want to see in the dashboard).", font=("Segoe UI", 12)).pack(pady=(0,10))
        

        self.scroll = ctk.CTkScrollableFrame(self, fg_color="transparent")
        self.scroll.pack(fill="both", expand=True, padx=20, pady=10)
        

        self.entries = {}
        self.header_rows = {}
        
        for i, (sig, data) in enumerate(buckets_info.items()):
            frame = ctk.CTkFrame(self.scroll, fg_color="white")
            frame.pack(fill="x", pady=5)
        


            file_count = len(set(r[-1] for r in data['rows'])) 
            col_preview = ", ".join(data['cols'][:5])
            if len(data['cols']) > 5: col_preview += "..."
            
            info_text = f"Group {i+1}: {file_count} File(s) | Cols: {len(data['cols'])}\nPreview: [{col_preview}]"
            
            ctk.CTkLabel(frame, text=info_text, anchor="w", justify="left", text_color="gray", font=("Consolas", 11)).pack(fill="x", padx=10, pady=5)
            
            ent = ctk.CTkEntry(frame, placeholder_text=f"Enter Name for Table {i+1}...")
            ent.pack(fill="x", padx=10, pady=(0, 10))
            self.entries[sig] = ent


            hdr_var = ctk.StringVar(value="1")
            self.header_rows[sig] = hdr_var

            hdr_frame = ctk.CTkFrame(frame, fg_color="transparent")
            hdr_frame.pack(fill="x", padx=10, pady=(0, 10))

            ctk.CTkLabel(
            hdr_frame,
                text="Header rows to combine:",
                font=("Segoe UI", 11),
                text_color="#636e72"
            ).pack(side="left")

            hdr_menu = ctk.CTkOptionMenu(
                hdr_frame,
                values=["1", "2", "3", "4", "5"],
                variable=hdr_var,
                width=70
            )
            hdr_menu.pack(side="left", padx=(10, 0))
            
        btn_frame = ctk.CTkFrame(self, fg_color="transparent")
        btn_frame.pack(fill="x", pady=15, padx=20)
        
        ctk.CTkButton(btn_frame, text="Cancel", fg_color="#b2bec3", text_color="#2d3436",
              command=self.on_cancel).pack(side="left", expand=True, fill="x", padx=(0, 8))

        ctk.CTkButton(btn_frame, text="Generate Dashboard", fg_color=C_MAIN,
              command=self.on_submit).pack(side="left", expand=True, fill="x")
        

    def on_cancel(self):
        self.result = None
        try:
            self.callback(None)
        finally:
            self.destroy()

    def on_submit(self):
        final_map = {}
        used_names = set()
        


        for sig, ent in self.entries.items():
            name = get_stripped_text(ent)
            if not name:
                messagebox.showwarning("Missing Name", "Please provide a name for all tables.")
                return
            if name in used_names:
                messagebox.showwarning("Duplicate Name", f"Table name '{name}' is used twice.")
                return
            used_names.add(name)
            final_map[name] = {
                "cols": self.buckets_info[sig]["cols"],
                "rows": self.buckets_info[sig]["rows"],
                "files": self.buckets_info[sig].get("files", []),
                "header_rows": int(self.header_rows[sig].get())
            }
        
        self.result = final_map
        self.destroy()
        self.callback(self.result)

# MAIN APPLICATION

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Data Manager v10.7") 
        self.root.geometry("1150x800")
        
        # Determine paths
        if getattr(sys, 'frozen', False):
            self.app_dir = os.path.dirname(sys.executable)
        else:
            self.app_dir = os.path.dirname(os.path.abspath(__file__))
        

        self.data_dir = os.path.join(self.app_dir, 'Data')

        self.out_file = os.path.join(self.app_dir, 'Tool_Result.html')
        
        self.master_cols = []
        self.master_rows = []
        
        # State variables
        self.mode = ctk.StringVar(value="standard")
        self.skip_rows = ctk.StringVar(value="1")
        self.file_count = 0
        self.file_states = {}



        self.settings_visible = False
        self.var_fuzzy = ctk.StringVar(value="off")
        self.var_fuzzy_threshold = ctk.DoubleVar(value=0.85) 
        self.var_fuzzy_col = ctk.StringVar(value="Company")  
        self.var_export_fmt = ctk.StringVar(value="xlsx")
        self.weight_inputs = {}

        self.default_weights = dict(CONFIG["ROLE_WEIGHTS"])



        self.vcmd = (self.root.register(self.check_num), '%P')
        self.init_view()
        self.refresh_file_list()


    def init_view(self):
      

        self.nav = ctk.CTkFrame(self.root, width=260, corner_radius=0, fg_color="#006039")
        self.nav.pack(side="left", fill="y")
        self.nav.pack_propagate(False) 
        
        self.lbl_logo = ctk.CTkLabel(self.nav, text="Data Manager", 
                                     font=ctk.CTkFont(family="Segoe UI", size=26, weight="bold"), 
                                     text_color="white")
        self.lbl_logo.pack(pady=(50, 20))
        

        self.runBtn = ctk.CTkButton(
            self.nav, 
            text="PROCESS DATA", 
            font=ctk.CTkFont(family="Segoe UI", size=13, weight="bold"),
            fg_color=C_MAIN, 
            text_color="white",
            hover_color="#2eb84a",
            height=45,
            corner_radius=8,
            command=self.on_run_click
        )
        self.runBtn.pack(pady=(15, 5), padx=25, fill="x")
        tooltip(self.runBtn, "Start processing the selected files.\nThe system will merge, clean, and analyse your data\nbased on the chosen mode and settings.")
        self.pbar = ctk.CTkProgressBar(self.nav, height=10, corner_radius=5, progress_color=C_MAIN, fg_color="#004d2e")
        self.pbar.set(0)


        self.pbar.pack_forget()
        self.btn_settings = ctk.CTkButton(
            self.nav, 
            text="SETTINGS", 
            font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"),
            fg_color="transparent",
            border_width=1,
            border_color="#aeb6bf",
            text_color="#aeb6bf",
            hover_color="#004d2e",
            height=30,
            corner_radius=8,
            command=self.toggle_settings
        )
        self.btn_settings.pack(pady=5, padx=25, fill="x")
        tooltip(self.btn_settings, "Toggle the configuration panel.\nConfigure Smart Grouping (fuzzy matching),\nposition weights, and export format (xlsx / csv).")
        self.settings_frame = ctk.CTkScrollableFrame(self.nav, fg_color="#004d2e", height=320, label_text="Configuration", label_text_color="white")
        

        # Op Dashboard
        self.openBtn = ctk.CTkButton(
            self.nav, 
            text="VIEW DASHBOARD", 
            font=ctk.CTkFont(family="Segoe UI", size=13, weight="bold"),
            fg_color="transparent",
            border_width=2,
            border_color="#aeb6bf",
            text_color="#aeb6bf",
            hover_color="#004d2e",
            height=45,
            corner_radius=8,
            state="disabled", 
            command=self.launch_report
        )
        self.openBtn.pack(side="bottom", pady=20, padx=25, fill="x")
        tooltip(self.openBtn, "Open the generated HTML dashboard in your browser.\nEnabled only after a successful processing run.")

        self.btn_help = ctk.CTkButton(
            self.nav,
            text="❓  TUTORIAL",
            font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"),
            fg_color="transparent",
            border_width=1,
            border_color="#3DCD58",
            text_color="#3DCD58",
            hover_color="#004d2e",
            height=30,
            corner_radius=8,
            command=self.show_tutorial
        )
        self.btn_help.pack(side="bottom", pady=(0, 5), padx=25, fill="x")
        tooltip(self.btn_help, "Open the step-by-step tutorial for the Data Manager.")

        ctk.CTkLabel(self.nav, text="v10.7 | Internal Use", text_color="#8FBC8F", font=("Segoe UI", 11)).pack(side="bottom", pady=5)






        # Main
        self.main_area = ctk.CTkFrame(self.root, fg_color="#F5F7FA", corner_radius=0)
        self.main_area.pack(side="right", fill="both", expand=True)       
        self.container = ctk.CTkFrame(self.main_area, fg_color="transparent")
        self.container.pack(fill="both", expand=True, padx=30, pady=30)
        self.top_bar = ctk.CTkFrame(self.container, fg_color="transparent")
        self.top_bar.pack(fill="x", pady=(0, 15))
        self.cv_led = tk.Canvas(self.top_bar, width=15, height=15, bg="#f0f2f5", highlightthickness=0)
        self.led = self.cv_led.create_oval(2, 2, 13, 13, fill="gray", outline="")
        self.cv_led.pack(side="left", padx=(0, 10))
        self.lbl_path = ctk.CTkLabel(self.top_bar, text=f"SOURCE: {self.data_dir}", 
                                     font=("Consolas", 12, "bold"), text_color="#636e72")
        self.lbl_path.pack(side="left")
       
        self.mainGrid = ctk.CTkFrame(self.container, fg_color="transparent")
        self.mainGrid.pack(fill="x", expand=False, pady=(0, 20))
        self.mainGrid.grid_columnconfigure(0, weight=3) 
        self.mainGrid.grid_columnconfigure(1, weight=2) 



        # Left 
        self.configCard = ctk.CTkFrame(self.mainGrid, fg_color="white", corner_radius=15)
        self.configCard.grid(row=0, column=0, sticky="nsew", padx=(0, 15))
        
        ctk.CTkLabel(self.configCard, text="Import Settings", font=("Segoe UI", 16, "bold"), text_color="#2d3436").pack(anchor="w", padx=20, pady=15)
        

        self.rb_std = ctk.CTkRadioButton(self.configCard, text="Standard Mode (Committee List Dataset)",  
                                         variable=self.mode, value="standard", font=("Segoe UI", 13),
                                         fg_color=C_MAIN, command=self.update_inputs)
        self.rb_std.pack(anchor="w", padx=20, pady=(5, 10))
        tooltip(self.rb_std, "Standard mode. Choose this for committee list files\nwith single-row headers. This mode enables advanced analytics.")
        self.rb_gen = ctk.CTkRadioButton(self.configCard, text="Normal Mode", 
                                         variable=self.mode, value="normal", font=("Segoe UI", 13),
                                         fg_color=C_MAIN, command=self.update_inputs)
        self.rb_gen.pack(anchor="w", padx=20, pady=(5, 10))
        
        tooltip(self.rb_gen, "Generic mode. Choose this for standard 1-row files\nOR multi-row headers. Choosing this option will not give access to advanced analytics.")

        self.rb_sesa = ctk.CTkRadioButton(self.configCard, text="Community Mode (MS Form Export)", 
                                          variable=self.mode, value="sesa", font=("Segoe UI", 13),
                                          fg_color=C_MAIN, command=self.update_inputs)
        self.rb_sesa.pack(anchor="w", padx=20, pady=(5, 10))
        tooltip(self.rb_sesa, "Community mode. Use with the Standardization Microsoft Form export.\nGenerates a searchable & filterable directory by Expertise, Standard Group,\nExternal Association, and Internal Communities.")


        self.spin_box = ctk.CTkFrame(self.configCard, fg_color="transparent")
        self.spin_box.pack(anchor="w", padx=45, pady=(0, 20))
        
        ctk.CTkLabel(self.spin_box, text="Default Header Rows for preview:", font=("Segoe UI", 12), text_color="#636e72").pack(side="left", padx=(0, 10))
        
        self.ent_rows = ctk.CTkEntry(self.spin_box, width=50, textvariable=self.skip_rows,
                                     justify="center", font=("Segoe UI", 13), border_color="#dfe6e9",
                                     validate='key', validatecommand=self.vcmd)
        self.ent_rows.pack(side="left")
        self.ent_rows.configure(state="disabled") 


        # Right 
        self.fileCard = ctk.CTkFrame(self.mainGrid, fg_color="white", corner_radius=15)
        self.fileCard.grid(row=0, column=1, sticky="nsew")
        
        self.file_header_frame = ctk.CTkFrame(self.fileCard, fg_color="transparent")
        self.file_header_frame.pack(fill="x", padx=15, pady=10)

        self.lbl_files = ctk.CTkLabel(self.file_header_frame, text="Pending Files (0)", font=("Segoe UI", 14, "bold"), text_color="#2d3436")
        self.lbl_files.pack(side="left")

        self.btn_refresh = ctk.CTkButton(self.file_header_frame, text="↻", width=30, height=24,
                                         fg_color="#dfe6e9", text_color="#2d3436", hover_color="#b2bec3",
                                         font=("Segoe UI", 14), command=self.refresh_file_list)
        self.btn_refresh.pack(side="right")
        tooltip(self.btn_refresh, "Refresh file list")
        
        self.file_scroll = ctk.CTkScrollableFrame(self.fileCard, fg_color="transparent", height=120)
        self.file_scroll.pack(fill="both", expand=True, padx=5, pady=(0, 10))


        # Logs CMD lke
        ctk.CTkLabel(self.container, text="System Logs", font=("Segoe UI", 14, "bold"), text_color="#2d3436").pack(anchor="w", pady=(0, 5))
        
        self.log_frame = ctk.CTkFrame(self.container, fg_color="white", corner_radius=15)
        self.log_frame.pack(fill="both", expand=True)
        
        self.txt_log = scrolledtext.ScrolledText(self.log_frame, height=6, bg="#1e272e", fg="#d2dae2", 
                                                 font=("Consolas", 11), insertbackground="white", relief="flat", padx=15, pady=15, bd=0)
        self.txt_log.pack(fill="both", expand=True, padx=5, pady=5)


    #Tutorial
    def show_tutorial(self):
        win = ctk.CTkToplevel(self.root)
        win.title("Data Manager – Tutorial")
        win.geometry("720x620")
        win.resizable(False, False)
        win.grab_set()

        STEPS = [
            {
                "title": "👋  Welcome to Data Manager",
                "body": (
                    "This tool lets you process, clean, and analyse standardisation committee "
                    "data exported from IEC, CENELEC, or the Schneider community portal.\n\n"
                    "This short tutorial walks you through every section of the interface.\n\n"
                    "Use the  ◀  Previous  /  Next ▶  buttons below to navigate."
                ),
            },
            {
                "title": "📂  Step 1 – Place Your Files",
                "body": (
                    "Put all your Excel (.xlsx) source files inside the  /Data  folder located "
                    "in the same directory as the application.\n\n"
                    "• The 'Pending Files' panel (top-right) lists every file found.\n"
                    "• Tick or untick individual files to include/exclude them.\n"
                    "• The  ↻  Refresh button re-scans the folder if you add files while the "
                    "app is already running."
                ),
            },
            {
                "title": "⚙️  Step 2 – Choose Your Import Mode",
                "body": (
                    "Three modes are available in the 'Import Settings' panel:\n\n"
                    "🟢  Standard Mode  — For IEC/CENELEC committee list exports. Unlocks ALL "
                    "advanced analytics (Rankings, Radar, Heatmap …). All files must share the "
                    "same column layout.\n\n"
                    "🔵  Normal Mode  — For any other tabular Excel file. Supports merge, clean, "
                    "filter, and export. Analytics are disabled.\n\n"
                    "🟡  Community Mode  — For Microsoft Form exports from the Standardisation "
                    "Community. Generates a searchable expert directory with cards, map, and filters."
                ),
            },
            {
                "title": "🔧  Step 3 – Configure Settings (Optional)",
                "body": (
                    "Click  SETTINGS  to expand the configuration panel:\n\n"
                    "• Smart Grouping (Fuzzy Match)  — Merges similar company names automatically "
                    "(e.g. 'Schneider Electric Japan' → 'Schneider Electric'). Use the slider to "
                    "control strictness: >90 % is safe, <80 % is aggressive.\n\n"
                    "• Position Weights  — Assign point values to each role (Chair, Secretary, "
                    "Member …). These weights drive every analytics score.\n\n"
                    "• Export Format  — Choose  xlsx  (default) or  csv  for the output file."
                ),
            },
            {
                "title": "▶  Step 4 – Process Data",
                "body": (
                    "Click  PROCESS DATA  to start.\n\n"
                    "The app will:\n"
                    "  1. Read and merge all selected files.\n"
                    "  2. Apply deduplication and fuzzy grouping (if enabled).\n"
                    "  3. Compute all analytics scores.\n"
                    "  4. Generate a self-contained  Tool_Result.html  file.\n\n"
                    "Progress is shown in the System Logs panel at the bottom. "
                    "Processing typically takes ~10 seconds for a standard dataset."
                ),
            },
            {
                "title": "✅  You're All Set!",
                "body": (
                    "You now know everything you need to get started with the Data Manager.\n\n"
                    "Quick reminders:\n"
                    "• Hover over any button or radio button to see a quick tooltip.\n"
                    "• Filters in the dashboard update analytics in real time.\n"
                    "• For access issues contact: martial.patra@se.com or pauline.mourlon@se.com\n\n"
                    "Click  Close  to dismiss this window and start exploring!"
                ),
            },
        ]

        current = [0]

        
        hdr = ctk.CTkFrame(win, fg_color="#006039", corner_radius=0)
        hdr.pack(fill="x")
        ctk.CTkLabel(hdr, text="Data Manager  –  Tutorial",
                     font=ctk.CTkFont(family="Segoe UI", size=18, weight="bold"),
                     text_color="white").pack(pady=18, padx=25, anchor="w")

        
        prog_frame = ctk.CTkFrame(win, fg_color="#f0f2f5", corner_radius=0, height=4)
        prog_frame.pack(fill="x")
        self._tut_progbar = ctk.CTkProgressBar(prog_frame, height=5, corner_radius=0,
                                                progress_color="#3DCD58", fg_color="#dfe6e9")
        self._tut_progbar.pack(fill="x")

        
        content = ctk.CTkFrame(win, fg_color="white", corner_radius=0)
        content.pack(fill="both", expand=True, padx=0, pady=0)

        lbl_step = ctk.CTkLabel(content, text="", font=ctk.CTkFont(family="Segoe UI", size=11),
                                 text_color="#3DCD58")
        lbl_step.pack(anchor="w", padx=30, pady=(20, 0))

        lbl_title = ctk.CTkLabel(content, text="", font=ctk.CTkFont(family="Segoe UI", size=20, weight="bold"),
                                  text_color="#2d3436", wraplength=640, justify="left")
        lbl_title.pack(anchor="w", padx=30, pady=(5, 0))

        lbl_body = ctk.CTkLabel(content, text="", font=ctk.CTkFont(family="Segoe UI", size=13),
                                 text_color="#636e72", wraplength=640, justify="left")
        lbl_body.pack(anchor="w", padx=30, pady=(15, 0))

        
        footer = ctk.CTkFrame(win, fg_color="#f8f9fa", corner_radius=0, height=60)
        footer.pack(fill="x", side="bottom")
        footer.pack_propagate(False)

        btn_prev = ctk.CTkButton(footer, text="◀  Previous", width=130,
                                  fg_color="#dfe6e9", text_color="#2d3436", hover_color="#b2bec3",
                                  font=ctk.CTkFont(family="Segoe UI", size=13))
        btn_prev.pack(side="left", padx=20, pady=12)

        btn_next = ctk.CTkButton(footer, text="Next  ▶", width=130,
                                  fg_color="#3DCD58", text_color="white", hover_color="#2eb84a",
                                  font=ctk.CTkFont(family="Segoe UI", size=13, weight="bold"))
        btn_next.pack(side="right", padx=20, pady=12)

        lbl_counter = ctk.CTkLabel(footer, text="", font=("Segoe UI", 12), text_color="#636e72")
        lbl_counter.pack(side="right", padx=10)

        def refresh_view():
            idx = current[0]
            n = len(STEPS)
            step = STEPS[idx]
            lbl_step.configure(text=f"STEP {idx + 1} OF {n}")
            lbl_title.configure(text=step["title"])
            lbl_body.configure(text=step["body"])
            lbl_counter.configure(text=f"{idx + 1} / {n}")
            self._tut_progbar.set((idx + 1) / n)
            btn_prev.configure(state="normal" if idx > 0 else "disabled")
            is_last = idx == n - 1
            btn_next.configure(text="Close" if is_last else "Next  ▶",
                               fg_color="#c0392b" if is_last else "#3DCD58",
                               hover_color="#a93226" if is_last else "#2eb84a")

        def go_next():
            if current[0] < len(STEPS) - 1:
                current[0] += 1
                refresh_view()
            else:
                win.destroy()

        def go_prev():
            if current[0] > 0:
                current[0] -= 1
                refresh_view()

        btn_next.configure(command=go_next)
        btn_prev.configure(command=go_prev)
        refresh_view()

    def toggle_settings(self):
        if not self.settings_visible:
            self.btn_settings.configure(fg_color="#004d2e", text="⚙️  CLOSE SETTINGS")
            self.settings_frame.pack(after=self.btn_settings, pady=10, padx=20, fill="x")
            self.render_settings_content()
            self.settings_visible = True
        else:
            self.settings_frame.pack_forget()
            self.btn_settings.configure(fg_color="transparent", text="⚙️  SETTINGS")
            self.settings_visible = False

    def render_settings_content(self):
        for w in self.settings_frame.winfo_children():
            w.destroy()

        # Weights
        ctk.CTkLabel(self.settings_frame, text="Power Weights", font=("Segoe UI", 11, "bold"), text_color="#bdc3c7").pack(anchor="w", pady=(5,5))
        self.weight_inputs = {}
        for role, val in self.default_weights.items():
            r_frame = ctk.CTkFrame(self.settings_frame, fg_color="transparent", height=25)
            r_frame.pack(fill="x", pady=2)
            
            lbl = ctk.CTkLabel(r_frame, text=role, font=("Segoe UI", 11), text_color="white", width=120, anchor="w")
            lbl.pack(side="left")
            
            entry = ctk.CTkEntry(r_frame, width=50, height=22, font=("Consolas", 11), justify="center")
            entry.insert(0, val)
            entry.configure(state="disabled", text_color="gray") 
            entry.pack(side="right")
            self.weight_inputs[role] = entry

        ctk.CTkFrame(self.settings_frame, height=2, fg_color="#bdc3c7").pack(fill="x", pady=10)





        # Fuzzy
        ctk.CTkLabel(self.settings_frame, text="Advanced Logic (RapidFuzz)", font=("Segoe UI", 11, "bold"), text_color="#bdc3c7").pack(anchor="w", pady=(0,5))
    
        fuzzy_switch = ctk.CTkSwitch(self.settings_frame, text="Enable Smart Grouping", variable=self.var_fuzzy, 
                                     onvalue="on", offvalue="off", command=self.toggle_fuzzy_slider,
                                     font=("Segoe UI", 11), text_color="white", progress_color=C_MAIN)
        fuzzy_switch.pack(anchor="w", pady=5)
        
        self.ent_fuzzy_col = ctk.CTkEntry(self.settings_frame, textvariable=self.var_fuzzy_col, height=24, font=("Segoe UI", 11))
        self.ent_fuzzy_col.pack(fill="x", padx=25, pady=(0, 10))
        


        # Slider
        self.fuzzy_container = ctk.CTkFrame(self.settings_frame, fg_color="transparent")
        self.fuzzy_container.pack(fill="x", pady=(0, 10))
        
        self.lbl_fuzzy_val = ctk.CTkLabel(self.fuzzy_container, text=f"Match Sensitivity: {int(self.var_fuzzy_threshold.get()*100)}%", 
                                          font=("Segoe UI", 10), text_color="#dfe6e9")
        self.lbl_fuzzy_val.pack(anchor="w", padx=25)
        
        self.slider_fuzzy = ctk.CTkSlider(self.fuzzy_container, from_=0.5, to=1.0, variable=self.var_fuzzy_threshold,
                                          command=self.update_fuzzy_lbl, height=16, progress_color=C_MAIN)
        self.slider_fuzzy.pack(fill="x", padx=20)
        self.toggle_fuzzy_slider()

        # Export Format
        ctk.CTkLabel(self.settings_frame, text="Export Format", font=("Segoe UI", 11, "bold"), text_color="#bdc3c7").pack(anchor="w", pady=(5,5))
        ctk.CTkOptionMenu(self.settings_frame, values=["xlsx", "csv"], variable=self.var_export_fmt,
                          fg_color="#1e272e", button_color="#1e272e", width=100).pack(anchor="w", pady=5)






    def toggle_fuzzy_slider(self):
        if self.var_fuzzy.get() == "on":
            self.slider_fuzzy.configure(state="normal", progress_color=C_MAIN)
            self.ent_fuzzy_col.configure(state="normal", border_color="#555")
            self.lbl_fuzzy_val.configure(text_color="#dfe6e9")
        else:
            self.slider_fuzzy.configure(state="disabled", progress_color="gray")
            self.ent_fuzzy_col.configure(state="disabled", border_color="#333")
            self.lbl_fuzzy_val.configure(text_color="gray")

    def update_fuzzy_lbl(self, val):
        self.lbl_fuzzy_val.configure(text=f"Strictness: {int(val*100)}%")

    def get_current_weights(self):
        if not self.weight_inputs: return self.default_weights.copy()
        current = {}
        for role, entry_widget in self.weight_inputs.items():
            try:
                val = get_stripped_text(entry_widget)
                if val.isdigit(): current[role] = val
                else: current[role] = self.default_weights[role]
            except:
                current[role] = self.default_weights[role]
        return current

    def check_num(self, val):
        return val.isdigit() or val == ""

    def update_inputs(self):
        if self.mode.get() == "normal":
            self.ent_rows.configure(state="normal", border_color=C_MAIN)
            self.ent_rows.focus_set()
        else:
            self.ent_rows.configure(state="disabled", border_color="#dfe6e9")

    def toggle_led(self, color):
        self.cv_led.itemconfig(self.led, fill=color)

    def log(self, txt):
        self.txt_log.insert(tk.END, f"> {txt}\n")
        self.txt_log.see(tk.END)

    def refresh_file_list(self):
        self.log("scanning folder...") 
        self.file_states.clear()
        
        if not os.path.exists(self.data_dir):
            os.makedirs(self.data_dir)
            self.log(f"Created input folder: {self.data_dir}")
            self.toggle_led("gray")
            self.lbl_files.configure(text="Pending Files (0)")
        else:
            files = glob.glob(os.path.join(self.data_dir, '*.xlsx'))
            self.file_count = len(files)
            
            for w in self.file_scroll.winfo_children(): w.destroy()

            if self.file_count > 0:
                self.toggle_led(C_MAIN)
                self.lbl_files.configure(text=f"Pending Files ({self.file_count})")
                
                self.var_select_all = ctk.StringVar(value="on")
                
             
                ctrl_frame = ctk.CTkFrame(self.file_scroll, fg_color="transparent")
                ctrl_frame.pack(fill="x", pady=(0, 5))
                
                switch = ctk.CTkSwitch(ctrl_frame, text="Select All", variable=self.var_select_all, 
                                       onvalue="on", offvalue="off", command=self.toggle_select_all,
                                       font=("Segoe UI", 11, "bold"), height=20, progress_color=C_MAIN)
                switch.pack(side="left", padx=5)
                
                files.sort()
                for f in files:
                    fname = os.path.basename(f)
                    kb = os.path.getsize(f) / 1024
                  
                    f_var = ctk.StringVar(value="on")
                    self.file_states[f] = f_var
                    
                    r = ctk.CTkFrame(self.file_scroll, fg_color="transparent")
                    r.pack(fill="x", pady=2)
                    
                    chk = ctk.CTkCheckBox(r, text=fname, variable=f_var, onvalue="on", offvalue="off",
                                          font=("Segoe UI", 12), checkbox_height=18, checkbox_width=18,
                                          hover_color=C_MAIN, fg_color=C_MAIN)
                    chk.pack(side="left", padx=5)
                    
                    ctk.CTkLabel(r, text=f"{kb:.1f} KB", font=("Segoe UI", 11), text_color="gray").pack(side="right", padx=5)

                self.log(f"Ready. Found {self.file_count} files.")
            else:
                self.toggle_led("gray")
                self.lbl_files.configure(text="Pending Files (0)")
                self.log("Folder is empty.")
                ctk.CTkLabel(self.file_scroll, text="No .xlsx files found", text_color="gray").pack(pady=10)

    def toggle_select_all(self):
        val = self.var_select_all.get()
        for var in self.file_states.values(): var.set(val)

    def set_ui_idle(self, success=False):
        self.runBtn.configure(state="normal", text="▶  PROCESS DATA")
        self.pbar.pack_forget() 
        
        if success:
            self.toggle_led(C_MAIN)
            self.anim_btn() 
        else:
            self.toggle_led(C_ERR)
            self.openBtn.configure(state="disabled", border_color="#aeb6bf", text_color="#aeb6bf", text="🌐  OPEN DASHBOARD")

    def anim_btn(self, step=0):
        self.openBtn.configure(state="normal", border_color=C_MAIN, text_color=C_MAIN)
        cols = ["#e8f5e9", "transparent"] 
        if step < 6: 
            c = cols[step % 2]
            self.openBtn.configure(fg_color=c)
            self.root.after(300, lambda: self.anim_btn(step + 1))
        else:
            self.openBtn.configure(fg_color=C_MAIN, text_color="white", text="✔  OPEN DASHBOARD")

    def on_run_click(self):
        self.txt_log.delete(1.0, tk.END)
        selected_files = [f for f, var in self.file_states.items() if var.get() == "on"]
        
        if not selected_files:
            messagebox.showwarning("Selection Error", "Please select at least one file to process.")
            return


        weights_config = self.get_current_weights()
        fuzzy_enabled = (self.var_fuzzy.get() == "on")
        fuzzy_thresh = self.var_fuzzy_threshold.get()
        export_fmt = self.var_export_fmt.get()
        target_col_name = get_stripped_text(self.var_fuzzy_col) or "Company"

        self.runBtn.configure(state="disabled", text="PROCESSING...")
        self.openBtn.configure(state="disabled", fg_color="transparent", text_color="#aeb6bf", text="🌐  OPEN DASHBOARD", border_color="#aeb6bf")


        self.pbar.pack(pady=(0, 20), padx=25, fill="x", after=self.runBtn)
        self.pbar.set(0)

        self.toggle_led("#f1c40f") 
        
        t = threading.Thread(
            target=self.worker_process_files, 
            args=(selected_files, weights_config, fuzzy_enabled, fuzzy_thresh, export_fmt, target_col_name), 
            daemon=True
        )

        t.start()



    def run_fuzzy_logic(self, rows, headers, threshold, target_col_name):
        col_idx = -1
        target = target_col_name.lower()
        
        lower_headers = [h.lower() for h in headers]
        for i, h in enumerate(lower_headers):
            if target in h:
                col_idx = i
                break
        
        if col_idx == -1:
            self.log(f"WARN: Grouping skipped, col '{target_col_name}' not found.")
            return

        rf_threshold = threshold * 100
    
        
        unique_vals = list(set([str(r[col_idx]).strip() for r in rows if r[col_idx] and str(r[col_idx]).strip() != ""]))
        unique_vals.sort(key=len)
        
        mapping = {}
        self.log(f"analyzing {len(unique_vals)} unique values for grouping...")
        
        for i, short_val in enumerate(unique_vals):
            s_lower = short_val.lower()
         
            # Common internal shorthand: "SE", "S.E." etc.
            # Eventhough rightnow there is no SE nor S.E. in the dataset, just in case.
            if "schneider" in s_lower or s_lower in ["se", "s.e."]:
                mapping[short_val] = "Schneider Electric"
                continue
            
            if short_val in mapping: continue 
            if len(short_val) < 2:
                mapping[short_val] = short_val
                continue

            mapping[short_val] = short_val
            
            for long_val in unique_vals[i+1:]:
                if long_val in mapping: continue

                l_lower = long_val.lower()
                if "schneider" in l_lower or l_lower in ["se", "s.e."]:
                    mapping[long_val] = "Schneider Electric"
                    continue

                if l_lower.startswith(s_lower):
                    mapping[long_val] = short_val
                    continue

                len_ratio = len(short_val) / len(long_val)
                if len_ratio < 0.6: continue 

                score = fuzz.token_sort_ratio(s_lower, l_lower)
                if score >= rf_threshold:
                    mapping[long_val] = short_val
       
        count = 0
        for r in rows:
            orig = str(r[col_idx]).strip()
            if orig in mapping and mapping[orig] != orig:
                r[col_idx] = mapping[orig]
                count += 1
        
        self.log(f"Grouping done. Merged {count} records.")


    def worker_process_files(self, target_files, weights, is_fuzzy, fuzzy_thresh, export_fmt, target_col_name):
        try:
            t0 = time.time()  
            curr_mode = self.mode.get()

            if curr_mode == "sesa":
                self.log("Community (SESA) mode: reading Google Form export...")
                self.root.after(0, lambda: self.pbar.set(0.2))

                people = []
                errors = []

                for fpath in target_files:
                    fname = os.path.basename(fpath)
                    try:
                        wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
                        ws = wb.active
                        headers = [clean_cell_value(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]

                        
                        expertise_cols, std_cols, assoc_cols, community_cols = [], [], [], []
                        for idx, h in enumerate(headers):
                            hl = h.lower()
                            if "which expertise" in hl:
                                label = extract_form_sub_label(h, ["Which Expertise?"])
                                expertise_cols.append((idx, label))
                            elif "which standard group" in hl:
                                label = extract_form_sub_label(h, ["Which standard group?"])
                                std_cols.append((idx, label))
                            elif "external association" in hl:
                                label = extract_form_sub_label(h, ["External Association"])
                                assoc_cols.append((idx, label))
                            elif "internal communities" in hl:
                                label = extract_form_sub_label(h, ["Internal Communities"])
                                community_cols.append((idx, label))

                        def find_col(keyword):
                            kw = keyword.lower()
                            for i, h in enumerate(headers):
                                if kw in h.lower():
                                    return i
                            return -1

                        idx_fname     = find_col("first name")
                        idx_lname     = find_col("last name")
                        idx_sesa      = find_col("sesa")
                        idx_entity    = find_col("entity")
                        idx_role      = find_col("position")
                        idx_postype   = find_col("type of position")
                        idx_std_inv   = find_col("involved in standardization")
                        idx_role_desc = find_col("role description")   
                        idx_nc        = find_col("national committee")  

                        for row in ws.iter_rows(min_row=2, values_only=True):
                            if not any(row):
                                continue
                            cv = lambda i: clean_cell_value(row[i]) if i >= 0 and i < len(row) else ""

                            expertise = {label: cv(i) for i, label in expertise_cols if cv(i) and cv(i).lower() not in ("none","","n/a")}
                            std_groups = {label: cv(i) for i, label in std_cols if cv(i) and cv(i).lower() not in ("none","","n/a")}
                            associations = {label: cv(i) for i, label in assoc_cols if cv(i) and cv(i).lower() not in ("none","","n/a")}
                            communities = {label: cv(i) for i, label in community_cols if cv(i) and cv(i).lower() not in ("none","","n/a")}

                            person = {
                                "first_name":          cv(idx_fname),
                                "family_name":         cv(idx_lname),
                                "sesa":                cv(idx_sesa),
                                "entity":              cv(idx_entity),
                                "role":                cv(idx_role),
                                "position_type":       cv(idx_postype),
                                "std_involved":        cv(idx_std_inv),
                                "role_desc":           cv(idx_role_desc),
                                "national_committee":  cv(idx_nc),
                                "expertise":           expertise,
                                "std_groups":          std_groups,
                                "associations":        associations,
                                "communities":         communities,
                                "source_file":         fname,
                            }
                            people.append(person)

                        wb.close()
                        self.log(f"Loaded {fname} → {len(people)} people")

                    except Exception as e:
                        self.log(f"Error reading {os.path.basename(fpath)}: {e}")
                        errors.append(str(e))

                self.root.after(0, lambda: self.pbar.set(0.7))

                if not people:
                    self.log("No valid SESA data found.")
                    self.root.after(0, lambda: self.set_ui_idle(False))
                    return

                all_expertise   = sorted(set(k for p in people for k in p["expertise"]))
                all_std_groups  = sorted(set(k for p in people for k in p["std_groups"]))
                all_assocs      = sorted(set(k for p in people for k in p["associations"]))
                all_communities = sorted(set(k for p in people for k in p["communities"]))
                all_entities    = sorted(set(p["entity"] for p in people if p["entity"]))
                all_pos_types   = sorted(set(p["position_type"] for p in people if p["position_type"]))

                self.log("Generating SESA Community Dashboard...")
                html_out = render_sesa_html(people, all_expertise, all_std_groups, all_assocs, all_communities, all_entities, all_pos_types)

                with open(self.out_file, 'w', encoding='utf-8') as f:
                    f.write(html_out)

                self.root.after(0, lambda: self.pbar.set(1.0))
                elapsed = time.time() - t0
                self.log(f"Done — SESA dashboard ready ({elapsed:.1f}s)")
                self.root.after(0, lambda: self.set_ui_idle(True))
                self.root.after(0, lambda: messagebox.showinfo("Success", f"Community dashboard generated!\n{len(people)} people loaded."))
                return
            


            is_std = (curr_mode == "standard")

            h_count = 1
            if not is_std:
                try:
                    h_val = safe_int(self.skip_rows.get(), 1)
                    if h_val > 0: h_count = h_val
                except: pass

            self.log(f"{'Standard' if is_std else 'Normal'} mode: detecting file structures...")
            target_files.sort()
            
            
            buckets = {}
            first_sig_standard = None

            for i, fpath in enumerate(target_files):
                prog = (i + 1) / len(target_files)
                self.root.after(0, lambda p=prog: self.pbar.set(p))
                fname = os.path.basename(fpath)
                self.log(f"Reading {fname} ({i+1}/{len(target_files)})...")
                
                wb = None
                try:
                    wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
                    ws = wb.active
                    
                    if ws.max_row is not None and ws.max_row < h_count: 
                        self.log("Skipping empty sheet")
                        continue

                    local_cols = extract_headers(ws, h_count)
                    if not local_cols: continue
                    
                    sig = tuple(local_cols)
                    
                    if is_std:
                        if first_sig_standard is None: first_sig_standard = sig
                        elif sig != first_sig_standard:
                            raise ColumnMismatchError(
                            user_message=f"'{fname}' has different columns. Standard mode requires identical headers across all files.",
                            detail="Ensure the header row is identical in every file, or switch to Normal Mode."
                        )
                    
                    if sig not in buckets:
                        buckets[sig] = { 'cols': list(local_cols), 'rows': [], 'files': [], 'seen_rows': set() }

                    buckets[sig]['files'].append(fpath)

                    for r in ws.iter_rows(min_row=h_count+1, values_only=True):
                        if not any(r):
                            continue

                        row_data = []
                        for idx in range(len(local_cols)):
                            val = r[idx] if idx < len(r) else None
                            row_data.append(clean_cell_value(val))

                        row_data.append(fname)

                        sig_row = tuple(row_data)
                        if sig_row in buckets[sig]['seen_rows']:
                            continue
                        buckets[sig]['seen_rows'].add(sig_row)

                        buckets[sig]['rows'].append(row_data)

                except Exception as e:
                    self.log(f"Error reading {fname}: {e}")
                    if is_std: raise e 
                finally:
                    if wb: wb.close()

            if not buckets:
                self.log("No valid data found."); self.root.after(0, lambda: self.set_ui_idle(False)); return

            self.log("Processing buckets...")
            
            for sig in buckets:
                b_cols = buckets[sig]['cols']
                b_rows = buckets[sig]['rows']
                if is_fuzzy:
                    self.run_fuzzy_logic(b_rows, b_cols, fuzzy_thresh, target_col_name)
                    # Second dedup pass: fuzzy grouping may have unified previously distinct names,
                    # creating new duplicates that the initial seen_rows check could not anticipate.
                    seen_post_fuzzy = set()
                    deduped = []
                    for row in b_rows:
                        k = tuple(row)
                        if k not in seen_post_fuzzy:
                            seen_post_fuzzy.add(k)
                            deduped.append(row)
                    buckets[sig]['rows'] = deduped
            
            self.bucket_data_payload = (buckets, weights, is_std, export_fmt, is_fuzzy, fuzzy_thresh, target_col_name)
            self.root.after(0, self.prompt_table_names)
            
        except DataManagerError as e:
            self.log(f"{e.__class__.__name__}: {e.user_message}")
            if e.detail:
                self.log(f"Details: {e.detail}")
            messagebox.showerror("Processing error", e.user_message)
            self.root.after(0, lambda: self.set_ui_idle(False))
        except Exception as e:
            self.log(f"Unexpected error: {e}")
            if CONFIG["PRINT_TRACEBACKS"]:
                import traceback; traceback.print_exc()
            messagebox.showerror("Processing error", "Unexpected error while processing. Check the logs for details.")
            self.root.after(0, lambda: self.set_ui_idle(False))

    def prompt_table_names(self):
        buckets, weights, is_std, export_fmt, is_fuzzy, fuzzy_thresh, target_col_name = self.bucket_data_payload
        
        if len(buckets) > 25:
            messagebox.showerror("Limit Exceeded", f"Found {len(buckets)} different file structures. The limit is 25. Please cleanup your folder.")
            self.set_ui_idle(False)
            return

        if is_std or len(buckets) == 1:
            named_buckets = {}
            for sig, data in buckets.items():
                named_buckets["Main Table" if len(buckets)==1 else "Data"] = data
            self.finalize_export(named_buckets)
        else:
            TableStructureNamerDialog(self.root, buckets, self.finalize_export)

    def finalize_export(self, named_buckets):
        if not named_buckets:
            self.log("Operation cancelled."); self.set_ui_idle(False); return
            
        try:
            t0 = time.time()
            # TODO: This rebuild re-reads all files; acceptable for now but slow on 50+ files
            _, weights, is_std, export_fmt, is_fuzzy, fuzzy_thresh, target_col_name = self.bucket_data_payload
            self.log("Generating HTML Dashboard...")
            
            rebuilt = {}
            for tab_name, data in named_buckets.items():
                h_count = safe_int(data.get("header_rows"), 1)

                new_cols = None
                new_rows = []
                # Cross-file dedup for finalize_export (mirrors the bucket-level seen_rows in initial read)
                seen_rows = set()

                for fpath in data.get("files", []):
                    wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
                    ws = wb.active
                    fname = os.path.basename(fpath)

                    local_cols = extract_headers(ws, h_count)
                    if not local_cols:
                        wb.close()
                        continue

                    if new_cols is None:
                        new_cols = local_cols

                    for r in ws.iter_rows(min_row=h_count + 1, values_only=True):
                        if not any(r):
                            continue

                        row_data = []
                        for idx in range(len(local_cols)):
                            val = r[idx] if idx < len(r) else None
                            row_data.append(clean_cell_value(val))

                        row_data.append(fname)

                        sig_row = tuple(row_data)
                        if sig_row in seen_rows:
                            continue
                        seen_rows.add(sig_row)
                        new_rows.append(row_data)

                    wb.close()

                if is_fuzzy:
                    self.run_fuzzy_logic(new_rows, new_cols or [], fuzzy_thresh, target_col_name)
                    # Second dedup pass after fuzzy: merged names may produce new duplicates
                    seen_post_fuzzy = set()
                    deduped = []
                    for row in new_rows:
                        k = tuple(row)
                        if k not in seen_post_fuzzy:
                            seen_post_fuzzy.add(k)
                            deduped.append(row)
                    new_rows = deduped

                rebuilt[tab_name] = {"cols": new_cols or [], "rows": new_rows}

            named_buckets = rebuilt

            html_out = render_dashboard_html(named_buckets, is_std, weights, export_fmt)
            
            with open(self.out_file, 'w', encoding='utf-8') as f:
                f.write(html_out)
            

            elapsed = time.time() - t0
            self.log(f"Done — dashboard updated ({elapsed:.1f}s)"); self.set_ui_idle(True)
            messagebox.showinfo("Success", "Processing complete!")
        except DataManagerError as e:
            self.log(f"{e.__class__.__name__}: {e.user_message}")
            if e.detail:
                self.log(f"Details: {e.detail}")
            messagebox.showerror("Export error", e.user_message)
            self.set_ui_idle(False)
        except Exception as e:
            self.log(f"Unexpected export error: {e}")
            if CONFIG["PRINT_TRACEBACKS"]:
                import traceback; traceback.print_exc()
            messagebox.showerror("Export error", "Unexpected error while exporting. Check the logs for details.")
            self.set_ui_idle(False)

    def launch_report(self):
        if os.path.exists(self.out_file):
            webbrowser.open('file://' + os.path.abspath(self.out_file))
            self.log("Opening Dashboard...")
        else:
            messagebox.showerror("Error", "File not found.")



# DASHBOARD (HTML/CSS/JS)

HTML_HEAD = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Schneider Data Manager</title>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet">
    <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css">
    <link rel="stylesheet" href="https://cdn.datatables.net/1.13.6/css/jquery.dataTables.min.css">
"""

HTML_SCRIPTS = """
    <script src="https://code.jquery.com/jquery-3.7.0.min.js"></script>
    <script src="https://cdn.datatables.net/1.13.6/js/jquery.dataTables.min.js"></script>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/xlsx/0.18.5/xlsx.full.min.js"></script>
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <script src="https://cdn.jsdelivr.net/npm/echarts@5.4.3/dist/echarts.min.js"></script>
"""



def render_sesa_html(people, all_expertise, all_std_groups, all_assocs, all_communities, all_entities, all_pos_types):
    import json as _json

    people_js = _json.dumps(people, ensure_ascii=False)

    def make_checkbox_group(title, icon, items, key, color):
        if not items:
            return ""
        html = f'<div class="filter-group"><div class="filter-title"><span class="ficon">{icon}</span>{title}</div><div class="filter-items" id="fg-{key}">'
        for item in items:
            safe = item.replace('"', '&quot;')
            html += f'<label class="chip"><input type="checkbox" data-key="{key}" value="{safe}" onchange="applyFilters()"><span>{item}</span></label>'
        html += '</div></div>'
        return html

    expertise_html   = make_checkbox_group("Expertise Domain", "🔬", all_expertise,   "expertise",   "#6c5ce7")
    std_html         = make_checkbox_group("Standard Group",   "📋", all_std_groups,  "std_groups",  "#0984e3")
    assoc_html       = make_checkbox_group("External Association", "🤝", all_assocs,  "associations","#00b894")
    community_html   = make_checkbox_group("Internal Communities", "🏘️", all_communities, "communities","#e17055")

    entity_opts = "<option value=''>All Entities</option>" + "".join(f"<option>{e}</option>" for e in all_entities)
    postype_opts = "<option value=''>All Position Types</option>" + "".join(f"<option>{p}</option>" for p in all_pos_types)

    

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>SESA Community Directory</title>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet">
<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css">
<script src="https://cdnjs.cloudflare.com/ajax/libs/xlsx/0.18.5/xlsx.full.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/echarts@5.4.3/dist/echarts.min.js"></script>
<style>
  *, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: 'Inter', sans-serif; background: #F5F7FA; color: #2d3436; display: flex; height: 100vh; overflow: hidden; }}

  /* Sidebar */
  #sidebar {{
    width: 320px; min-width: 280px; background: #fff; border-right: 1px solid #e0e0e0;
    display: flex; flex-direction: column; overflow: hidden;
  }}
  #sidebar-header {{
    background: #006039; color: white; padding: 18px 20px; flex-shrink: 0;
  }}
  #sidebar-header img {{ height: 28px; margin-bottom: 8px; display: block; }}
  #sidebar-header h2 {{ font-size: 16px; font-weight: 700; }}
  #sidebar-header p  {{ font-size: 12px; opacity: 0.8; margin-top: 2px; }}

  #filter-scroll {{ flex: 1; overflow-y: auto; padding: 14px; }}

  .search-box {{
    display: flex; align-items: center; background: #f0f4f8; border-radius: 8px;
    padding: 8px 12px; margin-bottom: 14px; gap: 8px;
  }}
  .search-box input {{
    border: none; background: transparent; outline: none; font-size: 13px;
    width: 100%; color: #2d3436;
  }}
  .search-box i {{ color: #b2bec3; font-size: 13px; }}

  .dropdowns {{ display: flex; flex-direction: column; gap: 8px; margin-bottom: 14px; }}
  .dropdowns select {{
    border: 1px solid #dfe6e9; border-radius: 6px; padding: 7px 10px;
    font-size: 12px; color: #2d3436; outline: none; background: white; cursor: pointer;
  }}

  .filter-group {{ margin-bottom: 14px; }}
  .filter-title {{
    font-size: 12px; font-weight: 600; color: #636e72; text-transform: uppercase;
    letter-spacing: 0.5px; margin-bottom: 7px; display: flex; align-items: center; gap: 6px;
  }}
  .ficon {{ font-size: 14px; }}
  .filter-items {{ display: flex; flex-wrap: wrap; gap: 5px; }}

  .chip label, .chip {{ display: inline-flex; }}
  .chip input {{ display: none; }}
  .chip span {{
    display: inline-block; padding: 4px 10px; border-radius: 20px; font-size: 11.5px;
    border: 1.5px solid #dfe6e9; cursor: pointer; color: #636e72; background: #fafafa;
    transition: all 0.15s; user-select: none; white-space: nowrap;
  }}
  .chip input:checked + span {{
    background: #006039; border-color: #006039; color: white; font-weight: 500;
  }}
  .chip span:hover {{ border-color: #3DCD58; color: #006039; }}

  .clear-btn {{
    width: 100%; margin-top: 8px; padding: 9px; background: #f0f4f8; border: none;
    border-radius: 8px; font-size: 12px; color: #636e72; cursor: pointer; font-weight: 500;
  }}
  .clear-btn:hover {{ background: #dfe6e9; }}

  /* Main */
  #main {{
    flex: 1; display: flex; flex-direction: column; overflow: hidden;
  }}
  #topbar {{
    background: white; border-bottom: 1px solid #e0e0e0; padding: 12px 20px;
    display: flex; align-items: center; justify-content: space-between; flex-shrink: 0;
  }}
  #topbar h1 {{ font-size: 18px; font-weight: 700; color: #006039; }}
  #result-count {{ font-size: 13px; color: #636e72; }}
  .export-btn {{
    background: #3DCD58; color: white; border: none; padding: 8px 18px;
    border-radius: 8px; font-size: 13px; font-weight: 600; cursor: pointer;
    display: flex; align-items: center; gap: 7px;
  }}
  .export-btn:hover {{ background: #2eb84a; }}

  /* Cards */
  #cards-wrap {{ flex: 1; overflow-y: auto; padding: 16px 20px; }}
  #no-results {{
    text-align: center; padding: 60px 20px; color: #b2bec3; font-size: 15px; display: none;
  }}
  #no-results i {{ font-size: 40px; display: block; margin-bottom: 12px; }}

  .person-card {{
    background: white; border-radius: 12px; padding: 16px 20px; margin-bottom: 12px;
    border: 1px solid #e0e0e0; display: flex; gap: 16px; align-items: flex-start;
    transition: box-shadow 0.15s;
  }}
  .person-card:hover {{ box-shadow: 0 4px 16px rgba(0,0,0,0.08); }}

  .avatar {{
    width: 46px; height: 46px; border-radius: 50%; background: #006039;
    color: white; font-size: 17px; font-weight: 700; display: flex; align-items: center;
    justify-content: center; flex-shrink: 0;
  }}
  .card-body {{ flex: 1; min-width: 0; }}
  .card-name {{ font-size: 15px; font-weight: 600; color: #2d3436; display: flex; align-items: center; flex-wrap: wrap; gap: 8px; }}
  .card-role-desc {{
    font-size: 11px; font-weight: 500; color: #006039; background: #e8f5ec;
    border: 1px solid #b2dfca; border-radius: 20px; padding: 2px 9px; white-space: nowrap;
  }}
  .card-meta {{ font-size: 12px; color: #636e72; margin: 2px 0 10px; }}
  .card-meta span {{ margin-right: 10px; }}

  .tag-row {{ display: flex; flex-wrap: wrap; gap: 5px; margin-bottom: 5px; }}
  .tag {{
    display: inline-flex; align-items: center; gap: 4px; padding: 3px 9px;
    border-radius: 20px; font-size: 11px; font-weight: 500; white-space: nowrap;
  }}
  .tag-exp    {{ background: #f3f0ff; color: #6c5ce7; border: 1px solid #e0d9ff; }}
  .tag-std    {{ background: #e8f4fd; color: #0984e3; border: 1px solid #c8e6f9; }}
  .tag-assoc  {{ background: #e8fdf5; color: #00b894; border: 1px solid #c3f0df; }}
  .tag-comm   {{ background: #fff4f2; color: #e17055; border: 1px solid #ffd5cc; }}
  .tag-role-lbl {{ font-weight: 400; opacity: 0.75; }}

  .section-lbl {{ font-size: 10.5px; color: #b2bec3; font-weight: 600; text-transform: uppercase;
                  letter-spacing: 0.3px; margin-bottom: 3px; margin-top: 6px; }}

  /* Map button */
  .map-btn {{
    background: #2d3436; color: white; border: none; padding: 8px 18px;
    border-radius: 8px; font-size: 13px; font-weight: 600; cursor: pointer;
    display: flex; align-items: center; gap: 7px;
  }}
  .map-btn:hover {{ background: #1e272e; }}

  /* NC active filter pill shown in topbar */
  #ncActiveBadge {{
    display: none; align-items: center; gap: 6px;
    background: #e8f5ec; border: 1.5px solid #3DCD58; border-radius: 20px;
    padding: 4px 12px; font-size: 12px; font-weight: 600; color: #006039;
  }}
  #ncActiveBadge button {{
    background: none; border: none; cursor: pointer; color: #006039;
    font-size: 14px; line-height: 1; padding: 0; margin-left: 2px;
  }}

  /* Map modal overlay */
  #communityMapModal {{
    display: none; position: fixed; inset: 0;
    background: rgba(0,0,0,0.55); z-index: 1000;
    align-items: center; justify-content: center;
  }}
  #communityMapModal.open {{ display: flex; }}
  #mapModalBox {{
    background: white; border-radius: 14px; width: 88vw; height: 82vh;
    display: flex; flex-direction: column; overflow: hidden;
    box-shadow: 0 20px 60px rgba(0,0,0,0.35);
  }}
  #mapModalHeader {{
    background: #006039; color: white; padding: 14px 20px;
    display: flex; align-items: center; justify-content: space-between; flex-shrink: 0;
  }}
  #mapModalHeader h2 {{ font-size: 16px; font-weight: 700; margin: 0; }}
  #mapModalHeader p  {{ font-size: 12px; opacity: 0.75; margin: 2px 0 0; }}
  #mapCloseBtn {{
    background: rgba(255,255,255,0.2); border: none; color: white;
    border-radius: 8px; padding: 6px 14px; font-size: 13px; cursor: pointer; font-weight: 600;
  }}
  #mapCloseBtn:hover {{ background: rgba(255,255,255,0.35); }}
  #communityMapContainer {{ flex: 1; width: 100%; }}
  #mapHint {{
    text-align: center; font-size: 11.5px; color: #b2bec3; padding: 6px 0;
    flex-shrink: 0; background: #fafafa; border-top: 1px solid #f0f0f0;
  }}

  /* Expert Network Modal */
  #expertNetModal {{
    display: none; position: fixed; inset: 0;
    background: rgba(0,0,0,0.55); z-index: 1000;
    align-items: center; justify-content: center;
  }}
  #expertNetModal.open {{ display: flex; }}
  #expertNetBox {{
    background: white; border-radius: 14px; width: 92vw; height: 88vh;
    display: flex; flex-direction: column; overflow: hidden;
    box-shadow: 0 20px 60px rgba(0,0,0,0.35);
  }}
  #expertNetHeader {{
    background: #2d3436; color: white; padding: 14px 20px;
    display: flex; align-items: center; justify-content: space-between; flex-shrink: 0; flex-wrap: wrap; gap: 10px;
  }}
  #expertNetHeader h2 {{ font-size: 16px; font-weight: 700; margin: 0; }}
  #expertNetHeader p  {{ font-size: 12px; opacity: 0.75; margin: 2px 0 0; }}
  #expertNetControls {{ display: flex; align-items: center; gap: 10px; flex-wrap: wrap; }}
  #expertNetControls select, #expertNetControls label {{
    border: 1px solid rgba(255,255,255,0.25); border-radius: 6px; padding: 6px 10px;
    font-size: 12px; background: rgba(255,255,255,0.12); color: white; outline: none; cursor: pointer;
  }}
  #expertNetControls select option {{ background: #2d3436; color: white; }}
  #expertNetCloseBtn {{
    background: rgba(255,255,255,0.2); border: none; color: white;
    border-radius: 8px; padding: 6px 14px; font-size: 13px; cursor: pointer; font-weight: 600; white-space: nowrap;
  }}
  #expertNetCloseBtn:hover {{ background: rgba(255,255,255,0.35); }}
  #expertNetContainer {{ flex: 1; width: 100%; }}
  #expertNetHint {{
    text-align: center; font-size: 11.5px; color: #b2bec3; padding: 6px 0;
    flex-shrink: 0; background: #fafafa; border-top: 1px solid #f0f0f0;
  }}
  .net-btn {{
    background: #4a5568; color: white; border: none; padding: 8px 18px;
    border-radius: 8px; font-size: 13px; font-weight: 600; cursor: pointer;
    display: flex; align-items: center; gap: 7px;
  }}
  .net-btn:hover {{ background: #2d3748; }}
</style>
</head>
<body>

<!-- SIDEBAR -->
<div id="sidebar">
  <div id="sidebar-header">
    <h2>Community Directory</h2>
    <p>Standardization Network</p>
  </div>
  <div id="filter-scroll">
    <div class="search-box">
      <i class="fas fa-search"></i>
      <input type="text" id="nameSearch" placeholder="Search by name, SESA, role..." oninput="applyFilters()">
    </div>
    <div class="dropdowns">
      <select id="entityFilter" onchange="applyFilters()">{entity_opts}</select>
      <select id="posTypeFilter" onchange="applyFilters()">{postype_opts}</select>
    </div>
    {expertise_html}
    {std_html}
    {assoc_html}
    {community_html}
    <button class="clear-btn" onclick="clearAll()">✕ &nbsp;Clear all filters</button>
  </div>
</div>

<!-- MAIN -->
<div id="main">
  <div id="topbar">
    <div style="display:flex; align-items:center; gap:14px; flex-wrap:wrap;">
      <div>
        <h1>SESA Community Explorer</h1>
        <span id="result-count">Loading...</span>
      </div>
      <div id="ncActiveBadge">
        🌍 <span id="ncActiveLabel"></span>
        <button onclick="clearNcFilter()" title="Clear country filter">✕</button>
      </div>
    </div>
    <div style="display:flex; gap:10px;">
      <button class="map-btn" onclick="openCommunityMap()">🌍 Community Map</button>
      <button class="net-btn" onclick="openExpertNetwork()">🕸 Expert Network</button>
      <button class="export-btn" onclick="exportExcel()">
        <i class="fas fa-file-excel"></i> Export to Excel
      </button>
    </div>
  </div>
  <div id="cards-wrap">
    <div id="no-results"><i class="fas fa-user-slash"></i>No people match the selected filters.</div>
    <div id="cards-container"></div>
  </div>
</div>

<!-- COMMUNITY MAP MODAL -->
<div id="communityMapModal">
  <div id="mapModalBox">
    <div id="mapModalHeader">
      <div>
        <h2>🌍 Community Map</h2>
        <p>People per country — double-click a country to filter the directory</p>
      </div>
      <button id="mapCloseBtn" onclick="closeCommunityMap()">✕ Close</button>
    </div>
    <div id="communityMapContainer"></div>
    <div id="mapHint">Double-click a highlighted country to filter · Scroll to zoom · Drag to pan</div>
  </div>
</div>

<!-- EXPERT NETWORK MODAL -->
<div id="expertNetModal">
  <div id="expertNetBox">
    <div id="expertNetHeader">
      <div>
        <h2>🕸 Expert Network</h2>
        <p>Experts connected by shared expertise domains — click a node to filter the directory</p>
      </div>
      <div id="expertNetControls">
        <select id="netDomainFilter" onchange="buildExpertNetwork()">
          <option value="">All Expertise Domains</option>
        </select>
        <select id="netColorBy" onchange="buildExpertNetwork()">
          <option value="entity">Colour by Entity</option>
          <option value="nc">Colour by National Committee</option>
          <option value="postype">Colour by Position Type</option>
        </select>
        <button id="expertNetCloseBtn" onclick="closeExpertNetwork()">✕ Close</button>
      </div>
    </div>
    <div id="expertNetContainer"></div>
    <div id="expertNetEmpty" style="display:none; flex:1; align-items:center; justify-content:center; font-size:15px; color:#b2bec3; padding:40px; text-align:center;">
      😶 No experts match the selected domain filter.
    </div>
    <div id="expertNetHint">Click node to filter · Scroll to zoom · Drag to pan · Hover for details</div>
  </div>
</div>

<script>
const ALL_PEOPLE = {people_js};
let filtered = [];
let activeNcFilter = '';   // set by double-clicking the map

function applyFilters() {{
  const nameQ    = document.getElementById('nameSearch').value.trim().toLowerCase();
  const entityQ  = document.getElementById('entityFilter').value;
  const posTypeQ = document.getElementById('posTypeFilter').value;

  const checked = {{}};
  document.querySelectorAll('.chip input:checked').forEach(cb => {{
    const k = cb.dataset.key;
    if (!checked[k]) checked[k] = [];
    checked[k].push(cb.value);
  }});

  filtered = ALL_PEOPLE.filter(p => {{
   
    if (nameQ) {{
      const blob = [p.first_name, p.family_name, p.sesa, p.role, p.entity, p.role_desc, p.national_committee].join(' ').toLowerCase();
      if (!blob.includes(nameQ)) return false;
    }}
    
    if (entityQ  && p.entity         !== entityQ)  return false;
    if (posTypeQ && p.position_type  !== posTypeQ) return false;
    if (activeNcFilter && (p.national_committee||'').trim().toUpperCase() !== activeNcFilter) return false;

    for (const [key, vals] of Object.entries(checked)) {{
      if (!vals.length) continue;
      const pMap = p[key] || {{}};
      if (!vals.every(v => v in pMap)) return false;
    }}
    return true;
  }});

  const badge = document.getElementById('ncActiveBadge');
  if (activeNcFilter) {{
    badge.style.display = 'flex';
    document.getElementById('ncActiveLabel').textContent = activeNcFilter;
  }} else {{
    badge.style.display = 'none';
  }}

  renderCards();
}}

function renderCards() {{
  const container = document.getElementById('cards-container');
  const noRes = document.getElementById('no-results');
  const cnt = document.getElementById('result-count');

  cnt.textContent = `${{filtered.length}} of ${{ALL_PEOPLE.length}} people`;

  if (!filtered.length) {{
    noRes.style.display = 'block';
    container.innerHTML = '';
    return;
  }}
  noRes.style.display = 'none';

  container.innerHTML = filtered.map(p => {{
    const initials = ((p.first_name||'?')[0] + (p.family_name||'?')[0]).toUpperCase();
    const fullName = [p.first_name, p.family_name].filter(Boolean).join(' ');

    const expTags = Object.entries(p.expertise||{{}}).map(([k,v]) =>
      `<span class="tag tag-exp">🔬 ${{k}} <span class="tag-role-lbl">(${{v}})</span></span>`).join('');
    const stdTags = Object.entries(p.std_groups||{{}}).map(([k,v]) =>
      `<span class="tag tag-std">📋 ${{k}} <span class="tag-role-lbl">(${{v}})</span></span>`).join('');
    const assocTags = Object.entries(p.associations||{{}}).map(([k,v]) =>
      `<span class="tag tag-assoc">🤝 ${{k}} <span class="tag-role-lbl">(${{v}})</span></span>`).join('');
    const commTags = Object.entries(p.communities||{{}}).map(([k,v]) =>
      `<span class="tag tag-comm">🏘️ ${{k}} <span class="tag-role-lbl">(${{v}})</span></span>`).join('');

    const expSection  = expTags  ? `<div class="section-lbl">Expertise Domain</div><div class="tag-row">${{expTags}}</div>`  : '';
    const stdSection  = stdTags  ? `<div class="section-lbl">Standard Groups</div><div class="tag-row">${{stdTags}}</div>`   : '';
    const aSection    = assocTags? `<div class="section-lbl">External Associations</div><div class="tag-row">${{assocTags}}</div>` : '';
    const cSection    = commTags ? `<div class="section-lbl">Internal Communities</div><div class="tag-row">${{commTags}}</div>`   : '';

    return `<div class="person-card">
      <div class="avatar">${{initials}}</div>
      <div class="card-body">
        <div class="card-name">${{fullName}}${{p.role_desc ? `<span class="card-role-desc">${{p.role_desc}}</span>` : ''}}</div>
        <div class="card-meta">
          <span>ID: ${{p.sesa||'—'}}</span>
          <span>Entity: ${{p.entity||'—'}}</span>
          <span>Position: ${{p.position_type||'—'}}${{p.national_committee ? ` &nbsp;·&nbsp; NC: <strong>${{p.national_committee}}</strong>` : ''}}</span>
        </div>
        ${{expSection}}${{stdSection}}${{aSection}}${{cSection}}
      </div>
    </div>`;
  }}).join('');
}}

function clearNcFilter() {{
  activeNcFilter = '';
  applyFilters();
}}

function clearAll() {{
  document.getElementById('nameSearch').value = '';
  document.getElementById('entityFilter').value = '';
  document.getElementById('posTypeFilter').value = '';
  document.querySelectorAll('.chip input:checked').forEach(cb => cb.checked = false);
  activeNcFilter = '';
  applyFilters();
}}

function exportExcel() {{
  if (!filtered.length) {{ alert('No data to export.'); return; }}
  const rows = [['First Name','Family Name','SESA','Entity','Role','Position Type','Role Description','National Committee',
                  'Expertise (Domain:Role)','Standard Groups (Group:Role)',
                  'External Associations (Assoc:Role)','Internal Communities (Comm:Role)','Source File']];
  filtered.forEach(p => {{
    const fmt = obj => Object.entries(obj||{{}}).map(([k,v])=>`${{k}}: ${{v}}`).join('; ');
    rows.push([p.first_name, p.family_name, p.sesa, p.entity, p.role, p.position_type, p.role_desc||'', p.national_committee||'',
               fmt(p.expertise), fmt(p.std_groups), fmt(p.associations), fmt(p.communities), p.source_file]);
  }});
  const ws = XLSX.utils.aoa_to_sheet(rows);
  const wb = XLSX.utils.book_new();
  XLSX.utils.book_append_sheet(wb, ws, 'SESA Community');
  XLSX.writeFile(wb, 'SESA_Community_Export.xlsx');
}}

const isoMap = {json.dumps(CONFIG["ISO_MAP"])};

let communityChartInst = null;
let mapDataLoaded = false;

function openCommunityMap() {{
  const modal = document.getElementById('communityMapModal');
  modal.classList.add('open');

  const ncCounts = {{}};
  ALL_PEOPLE.forEach(p => {{
    const code = (p.national_committee||'').trim().toUpperCase();
    if (!code) return;
    ncCounts[code] = (ncCounts[code] || 0) + 1;
  }});

  const seriesData = [];
  let maxCount = 0;
  Object.entries(ncCounts).forEach(([code, count]) => {{
    const countryName = isoMap[code] || code;
    maxCount = Math.max(maxCount, count);
    seriesData.push({{ name: countryName, value: count, ncCode: code }});
  }});

  const container = document.getElementById('communityMapContainer');
  if (communityChartInst) {{ communityChartInst.dispose(); }}
  communityChartInst = echarts.init(container);

  const renderMap = () => {{
    communityChartInst.setOption({{
      backgroundColor: '#fafafa',
      tooltip: {{
        trigger: 'item',
        formatter: params => {{
          if (!params.data) return `<strong>${{params.name}}</strong><br/>No members`;
          return `<div style="font-size:14px;font-weight:700;margin-bottom:4px;">${{params.name}}</div>` +
                 `<div>NC Code: <strong style="color:#006039;">${{params.data.ncCode}}</strong></div>` +
                 `<div>People: <strong style="color:#3DCD58;font-size:16px;">${{params.value}}</strong></div>` +
                 `<div style="font-size:11px;color:#b2bec3;margin-top:4px;">Double-click to filter</div>`;
        }}
      }},
      visualMap: {{
        min: 0, max: maxCount || 1,
        left: 20, bottom: 30,
        text: ['More', 'Fewer'],
        calculable: true,
        inRange: {{ color: ['#c8f0d5', '#3DCD58', '#006039'] }},
        textStyle: {{ color: '#636e72' }}
      }},
      series: [{{
        type: 'map',
        map: 'world',
        roam: true,
        emphasis: {{
          label: {{ show: true, fontSize: 11, color: '#2d3436' }},
          itemStyle: {{ areaColor: '#f39c12' }}
        }},
        select: {{ disabled: true }},
        itemStyle: {{ borderColor: '#fff', borderWidth: 0.5, areaColor: '#e8ecef' }},
        data: seriesData
      }}]
    }}, true);

    communityChartInst.on('dblclick', params => {{
      if (!params.data) return;
      const code = params.data.ncCode;
      if (!code) return;
      activeNcFilter = code;
      closeCommunityMap();
      applyFilters();
      document.getElementById('cards-wrap').scrollTop = 0;
    }});

    window.addEventListener('resize', () => communityChartInst && communityChartInst.resize());
  }};

  if (!mapDataLoaded) {{
    communityChartInst.showLoading({{ text: 'Loading map…', color: '#3DCD58' }});
    fetch('https://s3-us-west-2.amazonaws.com/s.cdpn.io/95368/world.json')
      .then(r => r.json())
      .then(mapJson => {{
        echarts.registerMap('world', mapJson);
        mapDataLoaded = true;
        communityChartInst.hideLoading();
        renderMap();
      }})
      .catch(() => {{
        communityChartInst.hideLoading();
        alert('Could not load map data. An internet connection is required.');
      }});
  }} else {{
    renderMap();
  }}
}}

function closeCommunityMap() {{
  document.getElementById('communityMapModal').classList.remove('open');
}}

document.getElementById('communityMapModal').addEventListener('click', e => {{
  if (e.target === document.getElementById('communityMapModal')) closeCommunityMap();
}});


// EXPERT NETWORK

let expertNetChart = null;

function openExpertNetwork() {{
  // Populate domain filter from current filtered set
  const allDomains = [...new Set(filtered.flatMap(p => Object.keys(p.expertise || {{}})))].sort();
  const sel = document.getElementById('netDomainFilter');
  const prev = sel.value;
  sel.innerHTML = '<option value="">All Expertise Domains</option>' +
    allDomains.map(d => `<option value="${{d}}" ${{d === prev ? 'selected' : ''}}>${{d}}</option>`).join('');

  document.getElementById('expertNetModal').classList.add('open');
  buildExpertNetwork();
}}

function closeExpertNetwork() {{
  document.getElementById('expertNetModal').classList.remove('open');
}}

document.getElementById('expertNetModal').addEventListener('click', e => {{
  if (e.target === document.getElementById('expertNetModal')) closeExpertNetwork();
}});

function buildExpertNetwork() {{
  const container  = document.getElementById('expertNetContainer');
  const emptyState = document.getElementById('expertNetEmpty');
  const selectedDomain = document.getElementById('netDomainFilter').value;
  const colorBy        = document.getElementById('netColorBy').value;

  if (expertNetChart) {{ expertNetChart.dispose(); expertNetChart = null; }}

  // Determine which people to include
  let netPeople = filtered;
  if (selectedDomain) {{
    netPeople = filtered.filter(p => selectedDomain in (p.expertise || {{}}));
  }}

  // Cap at 120 nodes to keep the graph readable
  if (netPeople.length > 120) netPeople = netPeople.slice(0, 120);

  if (!netPeople.length) {{
    container.style.display  = 'none';
    emptyState.style.display = 'flex';
    return;
  }}
  container.style.display  = '';
  emptyState.style.display = 'none';

  // Colour palette
  const PALETTE = [
    '#6c5ce7','#0984e3','#00b894','#e17055','#fd79a8',
    '#fdcb6e','#a29bfe','#55efc4','#fab1a0','#74b9ff',
    '#636e72','#3DCD58','#f39c12','#e74c3c','#9b59b6'
  ];
  const colorMap = {{}};
  let ci = 0;
  const getColor = key => {{
    if (!key) return '#b2bec3';
    if (!colorMap[key]) colorMap[key] = PALETTE[ci++ % PALETTE.length];
    return colorMap[key];
  }};

  // Build nodes
  const nodes = netPeople.map((p, i) => {{
    const fullName = [p.first_name, p.family_name].filter(Boolean).join(' ') || p.sesa || 'Unknown';
    const domainCount = Object.keys(p.expertise || {{}}).length;
    const groupKey = colorBy === 'entity' ? (p.entity || '—')
                   : colorBy === 'nc'     ? (p.national_committee || '—')
                   : colorBy === 'postype'? (p.position_type || '—')
                   : (p.entity || '—');
    return {{
      id: String(i),
      name: fullName,
      value: domainCount,
      category: groupKey,
      symbolSize: Math.max(18, Math.min(46, 18 + domainCount * 4)),
      itemStyle: {{ color: getColor(groupKey) }},
      label: {{ show: true, formatter: p => p.data.name.split(' ')[0], fontSize: 10,
                position: 'bottom', color: '#2d3436' }},
      personData: p,
      _groupKey: groupKey
    }};
  }});

  // Build edges: connect nodes that share at least one expertise domain
  const links = [];
  for (let i = 0; i < netPeople.length; i++) {{
    const diKeys = Object.keys(netPeople[i].expertise || {{}});
    for (let j = i + 1; j < netPeople.length; j++) {{
      const djKeys = Object.keys(netPeople[j].expertise || {{}});
      const shared = diKeys.filter(d => djKeys.includes(d));
      if (!shared.length) continue;
      // If a specific domain is selected only keep edges containing that domain
      if (selectedDomain && !shared.includes(selectedDomain)) continue;
      links.push({{
        source: String(i), target: String(j),
        value: shared.length,
        sharedDomains: shared,
        lineStyle: {{ width: Math.min(shared.length * 1.2, 5), opacity: 0.45, color: '#aaa' }}
      }});
    }}
  }}

  // Legend categories
  const categories = [...new Set(nodes.map(n => n._groupKey))].sort()
    .map(k => ({{ name: k, itemStyle: {{ color: getColor(k) }} }}));

  expertNetChart = echarts.init(container);
  expertNetChart.setOption({{
    backgroundColor: '#f8f9fa',
    tooltip: {{
      trigger: 'item',
      enterable: false,
      formatter: params => {{
        if (params.dataType === 'node') {{
          const p = params.data.personData;
          const domains = Object.entries(p.expertise || {{}})
            .map(([k, v]) => `• ${{k}} <em style="opacity:.7">(${{v}})</em>`).join('<br/>');
          return `<div style="max-width:260px">
            <strong style="font-size:13px">${{params.name}}</strong><br/>
            <span style="font-size:11px;color:#636e72">
              ${{p.entity || '—'}} · ${{p.position_type || '—'}} · NC: ${{p.national_committee || '—'}}
            </span>
            ${{domains ? `<hr style="margin:5px 0;border-color:#eee"/><div style="font-size:11px">${{domains}}</div>` : ''}}
            <div style="font-size:10px;color:#b2bec3;margin-top:4px;">Click to filter directory</div>
          </div>`;
        }}
        if (params.dataType === 'edge') {{
          return `<strong>Shared:</strong> ${{params.data.sharedDomains.join(', ')}}`;
        }}
        return '';
      }}
    }},
    legend: {{
      type: 'scroll', orient: 'vertical',
      right: 10, top: 'middle',
      data: categories.map(c => c.name),
      textStyle: {{ fontSize: 11, color: '#2d3436' }},
      pageTextStyle: {{ color: '#2d3436' }},
      icon: 'circle'
    }},
    series: [{{
      type: 'graph',
      layout: 'force',
      data: nodes,
      links: links,
      categories: categories,
      roam: true,
      zoom: 0.85,
      draggable: true,
      force: {{
        repulsion: netPeople.length < 30 ? 280 : 180,
        gravity: 0.06,
        edgeLength: netPeople.length < 30 ? [100, 250] : [60, 160],
        layoutAnimation: true
      }},
      lineStyle: {{ curveness: 0.1 }},
      emphasis: {{
        focus: 'adjacency',
        lineStyle: {{ width: 4, opacity: 0.85 }},
        label: {{ show: true, fontSize: 11, fontWeight: 'bold' }}
      }},
      scaleLimit: {{ min: 0.4, max: 5 }}
    }}]
  }});

  expertNetChart.on('click', params => {{
    if (params.dataType !== 'node') return;
    const p = params.data.personData;
    const fullName = [p.first_name, p.family_name].filter(Boolean).join(' ');
    closeExpertNetwork();
    document.getElementById('nameSearch').value = fullName;
    applyFilters();
    document.getElementById('cards-wrap').scrollTop = 0;
  }});

  window.addEventListener('resize', () => expertNetChart && expertNetChart.resize());
}}




applyFilters();
</script>
</body>
</html>"""


# Dashboard rendering

def render_dashboard_html(named_buckets, is_std, weights_dict=None, export_fmt='xlsx'):
    tables_js = {}
    
    for t_name, data in named_buckets.items():
        js_cols = [{"title": c} for c in data['cols']]
        
        if "📄 Source File" not in data['cols']:
             js_cols.append({"title": "📄 Source File"})
        f_counts = {}
        for r in data['rows']:
            fn = r[-1] 
            f_counts[fn] = f_counts.get(fn, 0) + 1
        
        tables_js[t_name] = {
            "columns": js_cols,
            "data": data['rows'],
            "file_labels": list(f_counts.keys()),
            "file_counts": list(f_counts.values())
        }

    # Get CSS and JS 
    css_block = _get_dashboard_css()
    js_block = _get_dashboard_js(tables_js, is_std, export_fmt, weights_dict, CONFIG.get("SEARCH_SYNONYMS", {}))
    

    logo_url = "https://raw.githubusercontent.com/Seantuy/Internship-Schneider-Electric/main/Schneider-Electric-Logo.jpg"
    
    # search options
    first_table_name = list(tables_js.keys())[0]
    first_cols = tables_js[first_table_name]['columns']
    col_opts = '<option value="all">Global Search...</option>'
    for i, c in enumerate(first_cols):
        col_opts += f'<option value="{i}">{c["title"]}</option>'

    #html anal
    if is_std:
        analytics_html = """
            <div class="analytics-toolbar">
                <span class="analytics-label">Analytics</span>
                <div class="analytics-pills">
                    <button class="pill-btn" id="powerBtn" title="Company Rankings – Ranks the top 20 entities by total weighted Power Score across all committees.">
                        <span class="pill-icon" style="background:#6c5ce7;">⚡</span>Company Rankings
                    </button>
                    <button class="pill-btn" id="radarBtn" title="Competitor Radar – Spider chart comparing up to 5 companies across their 8 most active shared committees.">
                        <span class="pill-icon" style="background:#e74c3c;">🎯</span>Competitor Radar
                    </button>
                    <button class="pill-btn" id="landscapeBtn" title="Gaps & Leads – Highlights committees where SE leads or trails its top competitor by more than 15 points.">
                        <span class="pill-icon" style="background:#0984e3;">⚖</span>Gaps &amp; Leads
                    </button>
                    <button class="pill-btn" id="networkBtn" title="Network Map – Force-directed graph showing connections between companies and the committees they participate in.">
                        <span class="pill-icon" style="background:#00b894;">🕸</span>Network Map
                    </button>
                    <button class="pill-btn" id="geoBtn" title="Geo Heatmap – World map coloured by national Power Score. Hover a country to see the top 5 contributing companies.">
                        <span class="pill-icon" style="background:#2d3436;">🌍</span>Geo Heatmap
                    </button>
                    <button class="pill-btn" id="benchBtn" title="Expert Workload – SE expert breakdown: role count, weighted strength, and global geographic footprint map.">
                        <span class="pill-icon" style="background:#e67e22;">🛡</span>Expert Workload
                    </button>
                    <button class="pill-btn" id="votingBtn" title="Voting Simulator – Simulate IEC (1 country = 1 vote) or CENELEC (weighted two-step) votes across all member countries.">
                        <span class="pill-icon" style="background:#8e44ad;">🗳</span>Voting
                    </button>
                </div>
            </div>
        """
    else:
        analytics_html = """
            <div class="empty-state">
                <i class="fas fa-info-circle"></i>
                <span>Advanced Analytics are disabled in Normal Mode (Multiple file structures detected).</span>
            </div>
        """

    # Weights Legend 
    if weights_dict is None: weights_dict = {}
    weight_visual_html = """
    <div class="card" style="margin-top: 20px; padding: 15px;">
        <div class="sidebar-title">Position Weight</div>
        <table style="width:100%; font-size:12px; color: var(--text-muted);">
    """
    sorted_weights = sorted(weights_dict.items(), key=lambda x: int(x[1]), reverse=True)
    for role, score in sorted_weights:
        weight_visual_html += f"""
        <tr>
            <td style="padding: 4px 0;">{role}</td>
            <td style="text-align:right; font-weight:bold; color:var(--text-main);">{score}</td>
        </tr>
        """
    weight_visual_html += "</table></div>"

    # Tab navigation
    tabs_nav = '<div class="tabs-nav">'
    for i, t_name in enumerate(tables_js.keys()):
        active = "active" if i == 0 else ""
        tabs_nav += f'<button class="tab-btn {active}" onclick="switchTab(\'{t_name}\')">{t_name}</button>'
    tabs_nav += '</div>'

    # Assemble the page
    html = HTML_HEAD + f"""
    <style>{css_block}</style>
</head>
<body>
    <nav class="navbar">
        <div class="logo-area">
            <img src="{logo_url}" alt="SE">
            <div><h1>Data Manager</h1></div>
        </div>
        <div style="display:flex; align-items:center; gap:18px;">
            <button onclick="openTutorial()" title="Open tutorial guide"
              style="display:flex; align-items:center; gap:8px; padding:8px 16px; border:2px solid #3DCD58; border-radius:8px; background:transparent; color:#3DCD58; font-size:13px; font-weight:700; cursor:pointer; font-family:Segoe UI,sans-serif; transition:background .2s;"
              onmouseover="this.style.background='#f0fdf4'" onmouseout="this.style.background='transparent'">
              ❓ Tutorial
            </button>
            <div style="font-size: 13px; color: var(--text-muted); font-weight: 500;">
                {datetime.now().strftime('%d %b %Y, %H:%M')}
            </div>
        </div>
    </nav>
    <div class="container">
        <aside>
            <div class="card">
                <div class="sidebar-title">Dataset Overview</div>
                <div class="stat-item"><span class="stat-label">Total Records</span><span class="stat-val" id="statTotal">0</span></div>
                <div class="stat-item"><span class="stat-label">Visible</span><span class="stat-val" style="color: var(--se-green);" id="statFiltered">0</span></div>
                <div class="stat-item"><span class="stat-label">Active Files</span><span class="stat-val" id="statFiles">0</span></div>
                
                <div class="sidebar-title" style="margin-top: 20px; margin-bottom: 5px;">Active Files</div>
                <div id="fileListContainer" class="file-list-container"></div>
            </div>
            {weight_visual_html}
        </aside>
        <main>
            <div class="card">
                <div class="search-wrapper">
                    <div class="input-group">
                        <select id="columnSelect">{col_opts}</select>
                        <input type="text" id="customSearch" placeholder="Type keywords to filter...">
                    </div>
                    <button class="btn btn-primary" id="addFilter">Apply</button>
                    <button class="btn btn-secondary" id="excludeBtn" title="Exclude"><i class="fas fa-ban"></i></button>
                    <button class="btn btn-secondary" id="clearAll" title="Reset"><i class="fas fa-undo"></i></button>
                </div>
                <div class="filter-logic-hint">
                    <i class="fas fa-lightbulb"></i> <strong>Tip:</strong> Filters in the SAME column are combined with <b>OR</b>. Filters in DIFFERENT columns are combined with <b>AND</b>.
                </div>
                <div class="filters-area" id="filtersContainer"><span style="color: var(--text-muted); font-size: 13px; padding-left: 5px;">No active filters</span></div>
                {analytics_html}
            </div>
            
            {tabs_nav}
            
            <div class="card" style="padding: 0; overflow: hidden; border-top-left-radius: 0;">
                 <div style="padding: 15px 25px; border-bottom: 1px solid var(--border-subtle); display: flex; justify-content: space-between; align-items: center; background: #fff;">
                    <div id="resultsInfo" style="font-weight: 600; font-size: 14px; color: var(--text-main);">Loading...</div>
                    <div>
                        <button class="btn btn-secondary" id="topExportBtn" style="padding: 6px 15px; font-size: 12px;">
                             <i class="fas fa-file-download"></i> Export Table
                        </button>
                    </div>
                 </div>
                <div style="padding: 0 25px 25px 25px;">
                    <table id="dataTable" class="display" style="width:100%"></table>
                </div>
            </div>
        </main>
    </div>
    
    <div id="detailsModal" class="modal"><div class="modal-content"><div class="modal-header"><h2>Record Details</h2><span class="close-btn" onclick="closeModal('detailsModal')">&times;</span></div><div class="modal-body" id="modalBody"></div></div></div>

    <!-- ===== TUTORIAL MODAL ===== -->
    <div id="tutorialModal" style="display:none; position:fixed; inset:0; background:rgba(0,0,0,0.6); z-index:9999; align-items:center; justify-content:center;">
      <div style="background:#fff; border-radius:16px; width:min(720px,95vw); max-height:90vh; display:flex; flex-direction:column; overflow:hidden; box-shadow:0 20px 60px rgba(0,0,0,0.4);">

        <!-- Header -->
        <div style="background:#006039; padding:22px 28px; display:flex; align-items:center; justify-content:space-between;">
          <div style="display:flex; align-items:center; gap:14px;">
            <span style="font-size:28px;">📖</span>
            <div>
              <div style="color:#fff; font-size:20px; font-weight:700; font-family:Segoe UI,sans-serif;">Data Manager – Tutorial</div>
              <div style="color:#8FBC8F; font-size:13px; font-family:Segoe UI,sans-serif;" id="tut-step-label">Step 1 of 7</div>
            </div>
          </div>
          <span onclick="closeTutorial()" style="color:#8FBC8F; font-size:24px; cursor:pointer; line-height:1;" title="Close tutorial">&times;</span>
        </div>

        <!-- Progress bar -->
        <div style="height:5px; background:#e0e0e0;">
          <div id="tut-progress" style="height:5px; background:#3DCD58; width:14.3%; transition:width .3s;"></div>
        </div>

        <!-- Body -->
        <div style="padding:30px 32px; flex:1; overflow-y:auto; font-family:Segoe UI,sans-serif;">
          <div id="tut-title" style="font-size:22px; font-weight:700; color:#2d3436; margin-bottom:16px;"></div>
          <div id="tut-body" style="font-size:14px; color:#555; line-height:1.8; white-space:pre-line;"></div>
        </div>

        <!-- Footer -->
        <div style="padding:16px 28px; background:#f8f9fa; display:flex; align-items:center; justify-content:space-between; border-top:1px solid #eee;">
          <button id="tut-prev" onclick="tutNav(-1)"
            style="padding:9px 22px; border:none; border-radius:8px; background:#dfe6e9; color:#2d3436; font-size:13px; font-weight:600; cursor:pointer;">
            ◀ Previous
          </button>
          <span id="tut-counter" style="font-size:12px; color:#888;"></span>
          <button id="tut-next" onclick="tutNav(1)"
            style="padding:9px 22px; border:none; border-radius:8px; background:#3DCD58; color:#fff; font-size:13px; font-weight:700; cursor:pointer;">
            Next ▶
          </button>
        </div>
      </div>
    </div>

    <div id="radarModal" class="modal">
        <div class="modal-content" style="max-width:1100px; height:85vh;">
            <div class="modal-header">
                <h2>Competitor Radar Comparison</h2>
                <div>
                    <button class="btn btn-secondary" style="margin-right: 15px; padding: 6px 15px; font-size: 12px;" onclick="exportInteractiveChart('echarts', radarChartInst, 'Interactive_Radar_Chart')"><i class="fas fa-file-download"></i> Export Interactive Chart</button>
                    <span class="close-btn" onclick="closeModal('radarModal')" style="cursor:pointer; font-size:24px;">&times;</span>
                </div>
            </div>
            <div class="modal-body" style="display: flex; flex-direction: column;">
                <div style="background: #f1f2f6; padding: 15px; border-radius: 8px; margin-bottom: 15px;">
                    <div style="font-size: 12px; font-weight: bold; color: #636e72; margin-bottom: 8px;">SELECT COMPANIES TO COMPARE (MAX 5):</div>
                    <div id="radarCompanyList" class="checkbox-list"></div>
                    <div style="margin-top: 10px; font-size: 11px; color: #636e72; font-style: italic;">
                        * The chart automatically selects the Top 8 Domains where the selected companies are most active combined.
                    </div>
                </div>
                <div id="radarChartContainer" style="flex-grow: 1; width: 100%; min-height: 400px;"></div>
            </div>
        </div>
    </div>

    <div id="networkModal" class="modal">
        <div class="modal-content">
            <div class="modal-header">
                <h2><span class="badge badge-green">Top 10</span> Map of Connections</h2>
                <span class="close-btn" onclick="closeModal('networkModal')" style="cursor:pointer; font-size:24px;">&times;</span>
            </div>
            <div class="modal-body" style="padding:0; overflow:hidden; display:flex; flex-direction:column;">
                <div style="padding: 15px; background: #fff; border-bottom: 1px solid #eee; display: flex; gap: 10px; align-items: center;">
                    <input list="networkDatalist" id="networkSearchInput" placeholder="Type to search Company or Group..." style="flex-grow: 1; padding: 10px; border: 1px solid #ddd; border-radius: 8px;">
                    <datalist id="networkDatalist"></datalist>
                    <button class="btn btn-primary" onclick="searchNetworkNode()">Search</button>
                    <button class="btn btn-secondary" onclick="resetNetworkMap()">Reset</button>
                    <button class="btn btn-secondary" onclick="exportInteractiveChart('echarts', networkChartInst, 'Interactive_Network_Map')"><i class="fas fa-file-download"></i> Export Viz</button>
                </div>
                <div id="networkMessage" style="position: absolute; top: 50%; left: 50%; transform: translate(-50%, -50%); text-align: center; color: #636E72; z-index: 10;">
                    <i class="fas fa-search" style="font-size: 3rem; margin-bottom: 10px; color: #dfe6e9;"></i>
                    <h3 style="margin: 0;">Map is empty</h3>
                    <p>Search for a Company or Group to visualize connections.</p>
                </div>
                <div id="networkContainer" style="width: 100%; height: 70vh;"></div>
            </div>
        </div>
    </div>
    
    <div id="geoModal" class="modal">
        <div class="modal-content" style="max-width:1300px; height:90vh;">
            <div class="modal-header">
                <h2>Geographic Heatmap</h2>
                <div>
                    <button class="btn btn-secondary" style="margin-right: 15px; padding: 6px 15px; font-size: 12px;" onclick="exportInteractiveChart('echarts', geoChartInst, 'Interactive_Geo_Heatmap')"><i class="fas fa-file-download"></i> Export Interactive Chart</button>
                    <span class="close-btn" onclick="closeModal('geoModal')" style="cursor:pointer; font-size:24px;">&times;</span>
                </div>
            </div>
            <div class="modal-body" style="display:flex; flex-direction:column; padding:15px;">
                <div id="geoLoader" class="loader"></div>
                <div id="geoContainer" style="width: 100%; flex-grow:1; min-height: 500px; border-radius: 12px; overflow: hidden;"></div>
            </div>
        </div>
    </div>
    
    <div id="powerModal" class="modal">
        <div class="modal-content" style="max-width:1000px; height:80vh;">
            <div class="modal-header">
                <h2>Company Power Rankings</h2>
                <div>
                    <button class="btn btn-secondary" style="margin-right: 15px; padding: 6px 15px; font-size: 12px;" onclick="exportInteractiveChart('chartjs', powerChartInst, 'Interactive_Power_Rankings')"><i class="fas fa-file-download"></i> Export Interactive Chart</button>
                    <span class="close-btn" onclick="closeModal('powerModal')" style="cursor:pointer; font-size:24px;">&times;</span>
                </div>
            </div>
            <div class="modal-body">
                <div style="height:300px; margin-bottom:20px;"><canvas id="powerChart"></canvas></div>
                <div id="powerTableContainer"></div>
            </div>
        </div>
    </div>

    <div id="landscapeModal" class="modal">
        <div class="modal-content" style="max-width:1100px; height:80vh;">
            <div class="modal-header">
                <h2>Gaps & Leads</h2>
                <span class="close-btn" onclick="closeModal('landscapeModal')" style="cursor:pointer; font-size:24px;">&times;</span>
            </div>
            <div class="modal-body">
                <div class="alert" style="background:#f1f2f6; padding:10px; border-radius:8px; margin-bottom:15px; font-size:12px; color:#636e72;">
                    <strong>Logic:</strong> Shows groups where the difference between SE Score and the Top Competitor Score is significant (Gap > SIGNIFICANT_GAP_THRESHOLD or Lead > SIGNIFICANT_GAP_THRESHOLD).
                </div>
                <div id="landscapeTableContainer"></div>
            </div>
        </div>
    </div>
    
    <div id="benchModal" class="modal">
        <div class="modal-content" style="max-width:1000px; height:80vh;">
            <div class="modal-header">
                <h2>Expert Workload Analysis</h2>
                <div>
                    <button class="btn btn-secondary" style="margin-right: 15px; padding: 6px 15px; font-size: 12px;" onclick="exportInteractiveChart('echarts', benchGeoChartInst, 'Interactive_Workload_Map')"><i class="fas fa-file-download"></i> Export Interactive Chart</button>
                    <span class="close-btn" onclick="closeModal('benchModal')" style="cursor:pointer; font-size:24px;">&times;</span>
                </div>
            </div>
            <div class="modal-body">
                <div style="margin-bottom: 20px; font-size:12px; color:#636e72; background:#f1f2f6; padding:10px; border-radius:8px;">
                      <strong>Metrics:</strong> "Strength" is the total weighted power score derived from all roles held by the expert.
                </div>
                <div id="benchGeoWrap" style="height:520px; margin-bottom:20px; border-radius:12px; overflow:hidden;">
                    <div id="benchGeoContainer" style="width:100%; height:100%;"></div>
                </div>
                <div id="benchTableContainer"></div>
            </div>
        </div>
    </div>

    <!-- Voting Modal -->
    <div id="votingModal" class="modal">
        <div class="modal-content" style="max-width:1340px; height:92vh;">
            <div class="modal-header">
                <h2>🗳 Voting Predictor</h2>
                <div style="display:flex; align-items:center; gap:10px;">
                    <button class="btn btn-secondary" style="padding:6px 15px; font-size:12px;" onclick="exportVotingResults()"><i class="fas fa-file-download"></i> Export</button>
                    <span class="close-btn" onclick="closeModal('votingModal')" style="cursor:pointer; font-size:24px;">&times;</span>
                </div>
            </div>
            <div class="modal-body" style="display:flex; flex-direction:column; height:calc(92vh - 65px); overflow:hidden; gap:10px;">

                <!-- Top control bar -->
                <div style="display:flex; gap:10px; align-items:center; flex-wrap:wrap; flex-shrink:0; padding:10px 14px; background:#f8f9fa; border-radius:10px; border:1px solid #edf2f7;">
                    <div>
                        <label style="font-size:12px; font-weight:600; color:#636e72; display:block; margin-bottom:3px;">Committee</label>
                        <select id="votingCommitteeSelect" onchange="loadVotingFromCommittee()" style="padding:6px 12px; font-size:13px; border-radius:6px; border:1px solid #ddd; min-width:180px;">
                            <option value="">— select —</option>
                        </select>
                    </div>
                    <div>
                        <label style="font-size:12px; font-weight:600; color:#636e72; display:block; margin-bottom:3px;">Mode</label>
                        <div style="background:#edf2f7; border-radius:8px; padding:3px; display:flex; gap:3px;">
                            <button id="modeIEC"     onclick="setVotingMode('IEC')"     style="padding:5px 16px; border-radius:6px; border:none; cursor:pointer; font-weight:600; font-size:13px; background:#3DCD58; color:white;">IEC</button>
                            <button id="modeCENELEC" onclick="setVotingMode('CENELEC')" style="padding:5px 16px; border-radius:6px; border:none; cursor:pointer; font-weight:600; font-size:13px; background:transparent; color:#636e72;">CENELEC</button>
                        </div>
                    </div>
                    <div style="margin-left:auto; display:flex; gap:6px; align-items:flex-end; padding-bottom:1px;">
                        <button onclick="setAllVotingParticipants('yes')"     style="padding:6px 11px; border-radius:6px; border:1px solid #27ae60; color:#27ae60; background:white; cursor:pointer; font-size:12px; font-weight:600;">✓ All Yes</button>
                        <button onclick="setAllVotingParticipants('no')"      style="padding:6px 11px; border-radius:6px; border:1px solid #c0392b; color:#c0392b; background:white; cursor:pointer; font-size:12px; font-weight:600;">✗ All No</button>
                        <button onclick="setAllVotingParticipants('abstain')" style="padding:6px 11px; border-radius:6px; border:1px solid #bdc3c7; color:#636e72; background:white; cursor:pointer; font-size:12px; font-weight:600;">— Reset</button>
                    </div>
                </div>

                <!-- Main two-column layout -->
                <div style="display:grid; grid-template-columns:1fr 390px; gap:12px; flex:1; min-height:0;">

                    <!-- Left: country list -->
                    <div style="display:flex; flex-direction:column; min-height:0;">
                        <!-- column headers -->
                        <div style="display:grid; grid-template-columns:1fr 120px 90px; gap:8px; padding:5px 12px; font-size:11px; font-weight:700; color:#888; text-transform:uppercase; letter-spacing:.5px; border-bottom:2px solid #edf2f7; flex-shrink:0;">
                            <span>Country</span>
                            <span>SE influence in NC</span>
                            <span style="text-align:right;">Vote</span>
                        </div>
                        <div id="votingCountryList" style="overflow-y:auto; flex:1;"></div>
                        <div id="votingNoDataMsg" style="display:none; padding:30px; text-align:center; color:#888; font-size:13px;">
                            Select a committee above to load participating countries and SE influence data.
                        </div>
                    </div>

                    <!-- Right: results -->
                    <div style="overflow-y:auto;">
                        <div id="votingResultsPanel"></div>
                    </div>
                </div>

            </div>
        </div>
    </div>
"""
    html += HTML_SCRIPTS + js_block + """
</body>
</html>
    """
    return html

def _get_dashboard_css():
    return """
        :root {
            --se-green: #3DCD58;
            --se-dark: #006039;
            --bg-body: #F5F7FA;
            --bg-card: #FFFFFF;
            --text-main: #2D3436;
            --text-muted: #636E72;
            --border-subtle: #EDF2F7;
            --shadow-sm: 0 2px 4px rgba(0,0,0,0.02);
            --shadow-md: 0 4px 12px rgba(0,0,0,0.05);
            --radius: 12px;
        }
        * { box-sizing: border-box; outline: none; }
        body { font-family: 'Inter', sans-serif; background: var(--bg-body); margin: 0; padding: 0; color: var(--text-main); -webkit-font-smoothing: antialiased; }
        .navbar { background: var(--bg-card); height: 70px; padding: 0 40px; display: flex; align-items: center; justify-content: space-between; box-shadow: var(--shadow-sm); position: sticky; top: 0; z-index: 100; }
        .logo-area { display: flex; align-items: center; gap: 15px; }
        .logo-area img { height: 35px; }
        .logo-area h1 { font-size: 18px; font-weight: 700; color: var(--se-dark); margin: 0; letter-spacing: -0.5px; }
        .container { max-width: 1600px; margin: 30px auto; padding: 0 30px; display: grid; grid-template-columns: 300px 1fr; gap: 30px; }
        .card { background: var(--bg-card); border-radius: var(--radius); box-shadow: var(--shadow-md); padding: 25px; margin-bottom: 25px; border: 1px solid var(--border-subtle); }
        .sidebar-title { font-size: 12px; font-weight: 700; color: var(--text-muted); text-transform: uppercase; margin-bottom: 15px; letter-spacing: 1px; }
        .stat-item { display: flex; justify-content: space-between; align-items: center; margin-bottom: 12px; }
        .stat-label { color: var(--text-muted); font-size: 14px; }
        .stat-val { font-weight: 700; font-size: 16px; color: var(--se-dark); }
        
        .file-list-container { max-height: 300px; overflow-y: auto; margin: 15px -10px 0 -10px; padding: 0 10px; }
        .file-item { 
            display: flex; justify-content: space-between; align-items: center; 
            padding: 8px 10px; margin-bottom: 4px; border-radius: 6px; 
            cursor: pointer; transition: background 0.2s; font-size: 13px; color: var(--text-main);
        }
        .file-item:hover { background-color: #f1f2f6; }
        .file-item.active { background-color: #e8f5e9; color: var(--se-dark); font-weight: 600; }
        .file-count { font-size: 11px; color: var(--text-muted); background: #dfe6e9; padding: 2px 6px; border-radius: 10px; }
        .file-item:hover .file-count { background: #fff; }

        
        .tabs-nav { display: flex; gap: 5px; margin-bottom: 0; padding-left: 10px; }
        .tab-btn { 
            padding: 10px 25px; background: #e0e0e0; border: none; 
            border-radius: 8px 8px 0 0; cursor: pointer; font-weight: 600; 
            color: #666; transition: 0.2s; font-size: 13px;
        }
        .tab-btn.active { 
            background: white; color: var(--se-dark); 
            border-top: 3px solid var(--se-green); 
            box-shadow: 0 -2px 5px rgba(0,0,0,0.05);
        }
        .tab-btn:hover:not(.active) { background: #d0d0d0; }

        .search-wrapper { display: flex; gap: 10px; margin-bottom: 10px; }
        .input-group { flex-grow: 1; display: flex; gap: 10px; }
        select, input { padding: 12px 16px; border: 1px solid #E2E8F0; border-radius: 8px; font-family: inherit; font-size: 14px; background: #FAFAFA; transition: 0.2s; }
        select:focus, input:focus { border-color: var(--se-green); background: #fff; box-shadow: 0 0 0 3px rgba(61, 205, 88, 0.1); }
        #columnSelect { width: 180px; }
        #customSearch { flex-grow: 1; }
        .analytics-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 15px; margin-top: 20px; }
        .action-card { background: #fff; border: 1px solid var(--border-subtle); border-radius: 10px; padding: 15px; display: flex; align-items: center; gap: 15px; cursor: pointer; transition: all 0.2s; }
        .action-card:hover { transform: translateY(-3px); box-shadow: var(--shadow-md); border-color: var(--se-green); }
        .icon-box { width: 40px; height: 40px; border-radius: 8px; display: flex; align-items: center; justify-content: center; font-size: 18px; color: white; }
        .bg-purple { background: #6c5ce7; }
        .bg-red { background: #e74c3c; }
        .bg-yellow { background: #f1c40f; }
        .bg-orange { background: #e67e22; }
        .bg-teal { background: #00b894; }
        .bg-blue { background: #0984e3; }
        .bg-green-dark { background: #2d3436; }
        .action-text h4 { margin: 0; font-size: 14px; font-weight: 600; color: var(--text-main); }
        .action-text p { margin: 2px 0 0 0; font-size: 11px; color: var(--text-muted); }
        .filters-area { margin-bottom: 20px; display: flex; flex-wrap: wrap; gap: 8px; min-height: 24px; }
        .chip { background: #E8F5E9; color: var(--se-dark); padding: 5px 12px; border-radius: 20px; font-size: 12px; font-weight: 600; display: flex; align-items: center; gap: 8px; }
        .chip i { cursor: pointer; opacity: 0.5; }
        .chip i:hover { opacity: 1; }
        table.dataTable { border-collapse: collapse !important; width: 100% !important; }
        table.dataTable thead th { background: white !important; color: var(--text-muted) !important; font-weight: 600 !important; font-size: 12px !important; text-transform: uppercase; border-bottom: 2px solid var(--border-subtle) !important; padding: 15px 10px !important; }
        table.dataTable tbody td { padding: 12px 10px !important; font-size: 13px; color: var(--text-main); border-bottom: 1px solid var(--border-subtle); cursor: pointer; user-select: none; }
        table.dataTable tbody td:last-child { cursor: default; }
        table.dataTable tbody td:not(:last-child):active { background-color: #e8f9ed !important; }
        table.dataTable tbody tr:hover { background-color: #FAFAFA !important; }
        .dataTables_info, .dataTables_paginate { font-size: 12px; margin-top: 15px; }
        .paginate_button.current { background: var(--se-green) !important; color: white !important; border: none !important; border-radius: 4px; }
        .btn { border: none; padding: 10px 20px; border-radius: 8px; font-weight: 600; font-size: 13px; cursor: pointer; transition: 0.2s; display: inline-flex; align-items: center; justify-content: center; gap: 8px; }
        .btn:hover { opacity: 0.9; transform: translateY(-1px); }
        .btn-primary { background: var(--se-green); color: white; }
        .btn-secondary { background: white; border: 1px solid #CBD5E0; color: var(--text-main); }
        .btn-secondary:hover { background: #F7FAFC; }
        .btn-link { background: transparent; color: var(--se-green); padding: 0; }
        .full-width { width: 100%; }
        .mb-3 { margin-bottom: 15px; }
        .modal { display: none; position: fixed; z-index: 1000; left: 0; top: 0; width: 100%; height: 100%; background: rgba(0,0,0,0.4); backdrop-filter: blur(4px); }
        .modal-content { background: white; margin: 3vh auto; width: 95%; max-width: 1400px; max-height: 85vh; border-radius: 16px; overflow: hidden; display: flex; flex-direction: column; }
        .modal-header { padding: 20px 30px; border-bottom: 1px solid var(--border-subtle); display: flex; justify-content: space-between; align-items: center; }
        .modal-header h2 { margin: 0; font-size: 18px; color: var(--text-main); display: flex; align-items: center; gap: 10px; }
        .close-btn { font-size: 24px; color: var(--text-muted); cursor: pointer; }
        .modal-body { padding: 20px; overflow-y: auto; flex: 1; min-height: 0; background: #FBFCFD; position: relative; }
        .badge { padding: 4px 8px; border-radius: 6px; font-size: 11px; font-weight: 700; }
        .badge-green { background: #E8F5E9; color: var(--se-dark); }
        .badge-red { background: #FEE2E2; color: #991B1B; }
        .badge-gray { background: #dfe6e9; color: #636e72; }
        .empty-state { text-align: center; color: var(--text-muted); padding: 20px; background: #F7FAFC; border-radius: 8px; border: 1px dashed #CBD5E0; font-size: 13px; }
        .filter-logic-hint { font-size: 11px; color: #636e72; margin-bottom: 10px; font-style: italic; background: #f1f2f6; padding: 5px 10px; border-radius: 4px; display: inline-block; }
        .loader { display: none; border: 4px solid #f3f3f3; border-top: 4px solid #3DCD58; border-radius: 50%; width: 30px; height: 30px; animation: spin 1s linear infinite; margin: 20px auto; }
        
        .checkbox-list { display: grid; grid-template-columns: repeat(auto-fill, minmax(180px, 1fr)); gap: 8px; max-height: 150px; overflow-y: auto; padding: 10px; border: 1px solid #eee; border-radius: 8px; background: #fff; }
        .checkbox-item { display: flex; align-items: center; gap: 8px; font-size: 13px; padding: 4px; border-radius: 4px; }
        .checkbox-item:hover { background: #f1f2f6; }
        .checkbox-item input { margin: 0; padding: 0; width: 16px; height: 16px; }
        
        @keyframes spin { 0% { transform: rotate(0deg); } 100% { transform: rotate(360deg); } }

        /* Analytics pill toolbar */
        .analytics-toolbar {
            display: flex;
            align-items: center;
            gap: 12px;
            padding: 10px 0 4px 0;
            flex-wrap: wrap;
        }
        .analytics-label {
            font-size: 11px;
            font-weight: 700;
            text-transform: uppercase;
            letter-spacing: 1px;
            color: var(--text-muted);
            white-space: nowrap;
        }
        .analytics-pills {
            display: flex;
            flex-wrap: wrap;
            gap: 7px;
        }
        .pill-btn {
            display: inline-flex;
            align-items: center;
            gap: 7px;
            padding: 6px 13px 6px 7px;
            background: #fff;
            border: 1px solid var(--border-subtle);
            border-radius: 20px;
            font-size: 12.5px;
            font-weight: 600;
            color: var(--text-main);
            cursor: pointer;
            transition: all 0.18s;
            white-space: nowrap;
            font-family: inherit;
            box-shadow: 0 1px 3px rgba(0,0,0,0.04);
        }
        .pill-btn:hover {
            border-color: var(--se-green);
            background: #f0fdf4;
            color: var(--se-dark);
            transform: translateY(-1px);
            box-shadow: 0 3px 8px rgba(61,205,88,0.15);
        }
        .pill-icon {
            width: 22px;
            height: 22px;
            border-radius: 50%;
            display: inline-flex;
            align-items: center;
            justify-content: center;
            font-size: 11px;
            flex-shrink: 0;
        }
    """

def _get_dashboard_js(tables_js, is_std, export_fmt, weights_dict, search_synonyms):
    
    weight_js = ""
    mapping = {
        "Secretary": "secretary", "Assistant Sec": "assistant sec", "Chair": "chair",
        "Vice-Chair": "vice-chair", "Liaison": "liaison", "Convenor": "convenor",
        "Nat. Part": "national participant", "Member": "member"
    }
    sorted_keys = sorted(mapping.keys(), key=lambda k: len(mapping[k]), reverse=True)
    
    for k in sorted_keys:
        search_str = mapping[k]
        val = weights_dict.get(k, "0")
        weight_js += f"if (func.includes('{search_str}')) return {val};\n"
    
    mem_val = weights_dict.get("Member", "5")
    weight_js += f"if (func.includes('nc official') || func.includes('observer')) return {mem_val};\n"

    return f"""
    <script>
    const tablesData = {json.dumps(tables_js)};
    const isStdMode = {str(is_std).lower()};
    const exportFormat = "{export_fmt}";

    const searchSynonyms = {json.dumps(search_synonyms)};

    const reverseSynonyms = {{}};
    Object.entries(searchSynonyms).forEach(([committee, keywords]) => {{
        keywords.forEach(kw => {{
            if (!reverseSynonyms[kw]) reverseSynonyms[kw] = [];
            reverseSynonyms[kw].push(committee);
        }});
    }});

    const MAX_COMPANIES_COMPARE = {MAX_COMPANIES_COMPARE};
    const TOP_GROUPS = {CONFIG["TOP_GROUPS"]};
    const SIGNIFICANT_GAP_THRESHOLD = {CONFIG["SIGNIFICANT_GAP_THRESHOLD"]};
    
    let currentTableName = Object.keys(tablesData)[0];

    let fullData = [];
    let fullColumns = [];
    let sourceFileIndex = -1;
    
    let activeFilters = [];
    let filteredByTable = {{}};
    let filteredIndices = [];
    let table;
    
    let powerChartInst, benchChartInst, networkChartInst, geoChartInst, radarChartInst, benchGeoChartInst;
    let mapDataLoaded = false;
    let radarCompanies = [];
    let geoHeatmapMax = 0;
    let powerExportData = [], landscapeExportData = [], benchExportData = [];
    






    // FROM IEC WEBSITE so if there is changes in the participant visit IEC WEB to get the ISOCODE (I dont think there will be changes but just in case)
    const isoMap = {json.dumps(CONFIG["ISO_MAP"])};




    Chart.defaults.font.family = "'Inter', sans-serif";
    Chart.defaults.color = '#636E72';

    function buildNameToIso() {{
        const m = {{}};
        Object.keys(isoMap).forEach(code => {{
            m[isoMap[code].toLowerCase()] = code;
        }});
        return m;
    }}
    const nameToIso = buildNameToIso();
    
    function initTable(tableName) {{
        currentTableName = tableName;
        const tData = tablesData[tableName];
        
        fullData = tData.data;
        fullColumns = tData.columns;
        sourceFileIndex = fullColumns.length - 1; 
        
        if (!filteredByTable[tableName]) {{
            filteredByTable[tableName] = fullData.map((_, idx) => idx);
        }}
        filteredIndices = filteredByTable[tableName];
        
        const sel = $('#columnSelect');
        sel.empty();
        sel.append('<option value="all">Global Search...</option>');
        fullColumns.forEach((c, i) => {{
            sel.append(`<option value="${{i}}">${{c.title}}</option>`);
        }});
        
        $('#statTotal').text(fullData.length.toLocaleString());
        
        updateTableRender();
        updateFileList();
    }}



    
    function switchTab(tableName) {{
        
        $('.tab-btn').removeClass('active');
       
        $(`.tab-btn`).filter(function() {{ return $(this).text() === tableName; }}).addClass('active');
        
        initTable(tableName);
    }}

    function updateTableRender() {{
        let maxCols = Math.min(5, fullColumns.length);

        
        let displayCols = fullColumns.slice(0, maxCols).map(c => ({{ title: c.title }}));
        displayCols.push({{ title: "Details" }});

        
        let tableData = filteredIndices.map(i => {{
            let row = fullData[i];
            let newRow = row.slice(0, maxCols);
            newRow.push(i); 
            return newRow;
        }});

        


if ($.fn.DataTable.isDataTable('#dataTable')) {{
    table.destroy();
    $('#dataTable').empty();
}}

table = $('#dataTable').DataTable({{
    data: tableData,
    columns: displayCols,
    pageLength: 20,
    searching: false,
    lengthChange: false,
    scrollX: true,  
    language: {{ emptyTable: "No data available" }},
    columnDefs: [{{
        targets: -1,
        render: (data) => `<button class="btn btn-link" onclick="showDetails(${{data}})">View</button>`
    }}]
}});

$('#dataTable tbody').off('dblclick', 'td').on('dblclick', 'td', function() {{
    const colIdx = table.cell(this).index().column;
    if (colIdx >= displayCols.length - 1) return;
    const val = ($(this).text() || '').trim();
    if (!val) return;
    const colName = fullColumns[colIdx] ? fullColumns[colIdx].title : 'All';
    addFilter(val, colIdx, colName);
}});

        $('#statFiltered').text(filteredIndices.length.toLocaleString());
        $('#resultsInfo').html(activeFilters.length === 0 ? `All Records` : `Found ${{filteredIndices.length}} matches`);
        renderFilters();
    }}

    function updateFileList() {{
        const container = $('#fileListContainer');
        container.empty();
        
    
        
        const currentCounts = {{}};
        
        

        filteredIndices.forEach(idx => {{ 
            const file = fullData[idx][sourceFileIndex]; 
            currentCounts[file] = (currentCounts[file] || 0) + 1;
        }});

        const tData = tablesData[currentTableName];
        const allFiles = tData.file_labels; 





        
        let activeFileCount = 0;

        allFiles.forEach(file => {{
            const count = currentCounts[file] || 0;
            if (count > 0) activeFileCount++;
            
            const isActive = activeFilters.some(f => f.colIndex === sourceFileIndex && f.value === file);
            
            const itemHtml = `
                <div class="file-item ${{isActive ? 'active' : ''}}" onclick="addFilter('${{file}}', ${{sourceFileIndex}}, 'Source File')">
                    <span style="white-space:nowrap; overflow:hidden; text-overflow:ellipsis; margin-right:5px;" title="${{file}}"><i class="far fa-file-excel"></i> ${{file}}</span>
                    <span class="file-count">${{count}}</span>
                </div>`;
            container.append(itemHtml);
        }});
        
        $('#statFiles').text(activeFileCount.toLocaleString());
    }}

    function checkMatch(textToSearch, searchValue) {{
        if (!searchValue) return false;
        if (textToSearch.includes(searchValue)) return true;
        if (reverseSynonyms[searchValue]) {{
            return reverseSynonyms[searchValue].some(committee => textToSearch.includes(committee));
        }}
        return false;
    }}

    function isSynonymOnlySearch(searchValue) {{
        return !!reverseSynonyms[searchValue] && !searchValue.includes('/');
    }}

    function deduplicateByGroup(rowIndices, tableData, headers) {{
        const grpIdx = headers.findIndex(x => x.includes('group') || x.includes('reference') || x.includes('committee') || x.includes('ref'));
        const compIdx = headers.findIndex(x => x.includes('company') || x.includes('organization') || x.includes('organi'));
        if (grpIdx === -1) return rowIndices;

        const seenGroups = {{}};

        rowIndices.forEach(rowIdx => {{
            const row = tableData[rowIdx];
            const grp = (row[grpIdx] || '').trim().toLowerCase();
            if (!grp) return;
            const comp = compIdx !== -1 ? (row[compIdx] || '').trim().toLowerCase() : '';
            if (!seenGroups[grp] && comp.includes('schneider')) {{
                seenGroups[grp] = rowIdx;
            }}
        }});

        rowIndices.forEach(rowIdx => {{
            const row = tableData[rowIdx];
            const grp = (row[grpIdx] || '').trim().toLowerCase();
            if (!grp) return;
            if (seenGroups[grp] === undefined) {{
                seenGroups[grp] = rowIdx;
            }}
        }});

        const kept = new Set(Object.values(seenGroups));
        return rowIndices.filter(idx => kept.has(idx));
    }}
    
    function applyFilters() {{
        const globalIncludeFilters = activeFilters.filter(f => f.type === 'all');        
        const globalExcludeFilters = activeFilters.filter(f => f.type === 'excludeAll');

        const tableIncludeFilters = activeFilters.filter(f => f.type === 'col' || f.type === 'nonEmpty'); 
        const tableExcludeFilters = activeFilters.filter(f => f.type === 'excludeCol');                  

        
    Object.keys(tablesData).forEach(tName => {{
        const tData = tablesData[tName];
        const data = tData.data;
        const headers = (tData.columns || []).map(c => (c.title || '').toLowerCase());

        const out = [];
        data.forEach((row, rowIdx) => {{
            const rowStr = row.join(' ').toLowerCase();

            if (globalIncludeFilters.length > 0) {{
                const matchesInclude = globalIncludeFilters.some(f =>
                    checkMatch(rowStr, (f.value || "").toLowerCase())
                );
                if (!matchesInclude) return;
            }}

            if (globalExcludeFilters.length > 0) {{
                const matchesExclude = globalExcludeFilters.some(f =>
                    checkMatch(rowStr, (f.value || "").toLowerCase())
                );
                if (matchesExclude) return;
            }}

            out.push(rowIdx);
        }});

        const isSynonymMode = globalIncludeFilters.length > 0 &&
            globalIncludeFilters.every(f => isSynonymOnlySearch((f.value || "").toLowerCase()));

        filteredByTable[tName] = isSynonymMode ? deduplicateByGroup(out, data, headers) : out;
    }});
    const tName = currentTableName;
    const data = tablesData[tName].data;

    let base = filteredByTable[tName] || data.map((_, idx) => idx);

    

    if (tableIncludeFilters.length > 0) {{
        const filtersByCol = {{}};

        tableIncludeFilters.forEach(f => {{
        const key = f.colIndex;
        if (!filtersByCol[key]) filtersByCol[key] = [];
        filtersByCol[key].push(f);
        }});

        const colKeys = Object.keys(filtersByCol);
        const out = [];

        base.forEach(rowIdx => {{
        const row = data[rowIdx];

        const matchesAllCols = colKeys.every(key => {{
            const group = filtersByCol[key];

            


            return group.some(f => {{
            const cellVal = (row[f.colIndex] || "").toString().toLowerCase();
            if (f.type === 'nonEmpty') return cellVal.trim() !== "";
            return checkMatch(cellVal, (f.value || "").toLowerCase());
        }});
      }});

      if (matchesAllCols) out.push(rowIdx);
    }});

    base = out; 
  }}

  if (tableExcludeFilters.length > 0) {{
    const excludeByCol = {{}};

    tableExcludeFilters.forEach(f => {{
      const key = f.colIndex;
      if (!excludeByCol[key]) excludeByCol[key] = [];
      excludeByCol[key].push(f);
    }});

    const colKeys = Object.keys(excludeByCol);
    const out = [];

    base.forEach(rowIdx => {{
      const row = data[rowIdx];

      const hitExclude = colKeys.some(key => {{
        const group = excludeByCol[key];
        return group.some(f => {{
            const cellVal = (row[f.colIndex] || "").toString().toLowerCase();
            return checkMatch(cellVal, (f.value || "").toLowerCase());
        }});
      }});

      if (!hitExclude) out.push(rowIdx);
    }});

    base = out;
  }}

  filteredByTable[tName] = base;
  filteredIndices = filteredByTable[currentTableName] || fullData.map((_, idx) => idx);

  updateTableRender();
  updateFileList();
}}


    function addFilter(val, colIndex = 'all', colName = 'All', isNonEmpty = false, isExclude = false) {{
  if(!isNonEmpty && (!val || val.trim() === '')) return;
  if(!isNonEmpty) val = val.trim();

  let type;
  if (isNonEmpty) type = 'nonEmpty';
  else if (isExclude && colIndex === 'all') type = 'excludeAll';
  else if (isExclude && colIndex !== 'all') type = 'excludeCol';
  else type = (colIndex === 'all' ? 'all' : 'col');

  const exists = activeFilters.some(f =>
    f.colIndex === colIndex &&
    f.type === type &&
    (type === 'nonEmpty' ? true : f.value === val)
  );
  if(exists) return;

  activeFilters.push({{
    type: type,
    value: isNonEmpty ? 'Not Blank' : val,
    colIndex: colIndex,
    colName: colName
  }});

  renderFilters();
  applyFilters();
  $('#customSearch').val('');
}}

    function removeFilter(index) {{ activeFilters.splice(index, 1); renderFilters(); applyFilters(); }}
    
    function renderFilters() {{
  const c = $('#filtersContainer').empty();
  if(activeFilters.length === 0) {{
    c.html('<span style="color: var(--text-muted); font-size: 13px; padding-left: 5px;">No active filters</span>');
    return;
  }}

  activeFilters.forEach((f, idx) => {{
    const isEx = (f.type === 'excludeAll' || f.type === 'excludeCol');
    let label;

    if (f.colIndex === 'all') {{
      label = isEx ? `Global: NOT "${{f.value}}"` : `Global: "${{f.value}}"`;
    }} else {{
      label = isEx
        ? `<strong>${{f.colName}}:</strong> NOT ${{f.value}}`
        : `<strong>${{f.colName}}:</strong> ${{f.value}}`;
    }}

    c.append(`<div class="chip">${{label}} <i class="fas fa-times-circle" onclick="removeFilter(${{idx}})"></i></div>`);
  }});
}}

    window.exportDataToExcel = function(headers, data, filename) {{
        if(!data || data.length === 0) {{ alert('No data to export'); return; }}
        if (exportFormat === 'csv') {{
            let csvContent = "data:text/csv;charset=utf-8,";
            csvContent += headers.join(",") + "\\r\\n";
            data.forEach(function(rowArray) {{
                let row = rowArray.join(",");
                csvContent += row + "\\r\\n";
            }});
            var encodedUri = encodeURI(csvContent);
            var link = document.createElement("a");
            link.setAttribute("href", encodedUri);
            link.setAttribute("download", filename + ".csv");
            document.body.appendChild(link);
            link.click();
        }} else {{
            const ws = XLSX.utils.aoa_to_sheet([headers].concat(data));
            const wb = XLSX.utils.book_new();
            XLSX.utils.book_append_sheet(wb, ws, "Data");
            XLSX.writeFile(wb, filename + ".xlsx");
        }}
    }};

    window.showDetails = function(rowIndex) {{
        let html = '<div style="display:grid; grid-template-columns: 1fr 1fr; gap:20px;">';
        fullColumns.forEach((col, i) => html += `<div style="border-bottom:1px solid #eee; padding-bottom:5px;"><div style="font-size:11px; color:#999; text-transform:uppercase;">${{col.title}}</div><div style="font-weight:500;">${{fullData[rowIndex][i]}}</div></div>`);
        html += '</div>';
        $('#modalBody').html(html); $('#detailsModal').show();
    }};

    

    
    function getWeight(func) {{
        func = (func || "").toLowerCase();
        {weight_js}
        return 0;
    }}

    function getIndices() {{
        const h = fullColumns.map(c => c.title.toLowerCase());
        return {{
            comp: h.findIndex(x => x.includes('company') || x.includes('organization') || x.includes('organi')),
            func: h.findIndex(x => x.includes('function') || x.includes('role')),
            grp: h.findIndex(x => x.includes('group') || x.includes('reference') || x.includes('committee') || x.includes('ref')),
            nc: h.findIndex(x => x === 'nc' || x.includes('national committee')),
            desc: h.findIndex(x => x === 'group name' || x === 'groupname' || x === 'title'),
            lname: h.findIndex(x => x.includes('last_name') || x.includes('lastname') || x === 'last name' || x.includes('surname') || x.includes('nom')),
            fname: h.findIndex(x => x.includes('first_name') || x.includes('firstname') || x === 'first name' || x.includes('prénom'))
        }};
    }}

    function getFullName(row, idxs) {{
        if (idxs.lname === -1 && idxs.fname === -1) return "Unknown Name";
        const l = (idxs.lname !== -1 ? (row[idxs.lname] || "") : "").trim();
        const f = (idxs.fname !== -1 ? (row[idxs.fname] || "") : "").trim();
        return (l + " " + f).trim() || "Unknown Name";
    }}
    
    function makeModalTable(id) {{
        if ($.fn.DataTable.isDataTable('#' + id)) $('#' + id).DataTable().destroy();
        $('#' + id).DataTable({{ pageLength: 10, searching: true, lengthChange: false, language: {{ search: "", searchPlaceholder: "Search table..." }} }});
    }}
    


    function initRadar() {{
        if(!isStdMode) return;
        const idxs = getIndices();
        if (idxs.comp === -1 || idxs.grp === -1) {{ alert("Missing Company or Group columns."); return; }}
        const compStats = {{}};
        filteredIndices.forEach(i => {{
            const row = fullData[i]; const comp = (row[idxs.comp]||"").trim(); const w = getWeight(row[idxs.func]);
            if(comp && comp.toLowerCase() !== 'unknown' && w > 0) {{ if(!compStats[comp]) compStats[comp] = 0; compStats[comp] += w; }}
        }});
        const sortedComps = Object.entries(compStats).sort((a,b) => b[1] - a[1]).map(x => x[0]);
        const listContainer = $('#radarCompanyList'); listContainer.empty();
        let schneiderSelected = false; 
        sortedComps.forEach(c => {{
            let isChecked = ''; const isSchneider = c.toLowerCase().includes('schneider') || c.toLowerCase() === 'se';
            if (isSchneider && !schneiderSelected) {{ isChecked = 'checked'; schneiderSelected = true; }}
            listContainer.append(`<label class="checkbox-item"><input type="checkbox" value="${{c}}" ${{isChecked}} onchange="updateRadarChart()"> ${{c}}</label>`);
        }});
        $('#radarModal').show();
        if(radarChartInst) radarChartInst.dispose();
        radarChartInst = echarts.init(document.getElementById('radarChartContainer'));
        window.addEventListener('resize', () => radarChartInst.resize());
        updateRadarChart();
    }}



    function updateRadarChart() {{
        const idxs = getIndices();
        const selected = []; $('#radarCompanyList input:checked').each(function() {{ selected.push($(this).val()); }});
        if(selected.length > MAX_COMPANIES_COMPARE) {{ alert("Please select maximum MAX_COMPANIES_COMPARE companies."); return; }}
        if(selected.length === 0) {{ radarChartInst.clear(); return; }}
        const groupData = {{}};
        filteredIndices.forEach(i => {{
            const row = fullData[i]; const comp = (row[idxs.comp]||"").trim(); const grp = (row[idxs.grp]||"").trim(); const w = getWeight(row[idxs.func]);
            if(selected.includes(comp) && grp) {{
                if(!groupData[grp]) groupData[grp] = {{ total: 0 }};
                if(!groupData[grp][comp]) groupData[grp][comp] = 0;
                groupData[grp][comp] += w; groupData[grp].total += w;
            }}
        }});



        const topGroups = Object.entries(groupData).sort((a,b) => b[1].total - a[1].total).slice(0, TOP_GROUPS).map(x => x[0]);
        const compColors = ['#e74c3c', '#3498db', '#f1c40f', '#9b59b6', '#e67e22']; const finalColors = []; let colorPtr = 0; const seriesData = [];
        selected.forEach(comp => {{
            if (comp.toLowerCase().includes('schneider') || comp.toLowerCase() === 'se') {{ finalColors.push('#3DCD58'); }} else {{ finalColors.push(compColors[colorPtr % compColors.length]); colorPtr++; }}
            seriesData.push({{ name: comp, value: [] }});
        }});
        let globalMax = 0;
        topGroups.forEach((grp, gIdx) => {{
            selected.forEach((comp, cIdx) => {{
                const val = groupData[grp][comp] || 0; seriesData[cIdx].value.push(val); if(val > globalMax) globalMax = val;
            }});
        }});
        const indicators = topGroups.map(g => ({{ name: g, max: globalMax * 1.1 }}));
        radarChartInst.setOption({{ color: finalColors, tooltip: {{ trigger: 'item' }}, legend: {{ data: selected, bottom: 0 }}, radar: {{ indicator: indicators, shape: 'polygon' }}, series: [{{ type: 'radar', data: seriesData, areaStyle: {{ opacity: 0.1 }} }}] }}, true);
    }}

    function showNetworkMap() {{
        if(!isStdMode) return;
        const idxs = getIndices();
        if (idxs.grp === -1 || idxs.comp === -1) {{ alert("Missing Group or Company columns."); return; }}
        globalNetworkEdges = []; globalNodeStats = {{}};
        let uniqueCompanies = new Set(); let uniqueGroups = new Set();

        filteredIndices.forEach(i => {{
            const row = fullData[i];
            let comp = (row[idxs.comp]||"").trim();
            const grp = (row[idxs.grp]||"").trim();
            const w = getWeight(row[idxs.func]);
            const person = getFullName(row, idxs);

            if(!comp || !grp || comp.toLowerCase() === 'unknown') return;
            
            globalNetworkEdges.push({{ c: comp, g: grp, w: w, p: person }});
            
            if(!globalNodeStats[comp]) globalNodeStats[comp] = 0; globalNodeStats[comp] += w;
            if(!globalNodeStats[grp]) globalNodeStats[grp] = 0; globalNodeStats[grp] += w;

            uniqueCompanies.add(comp); uniqueGroups.add(grp);
        }});

        

        if(globalNetworkEdges.length === 0) {{ alert("No data available for network map."); return; }}
        
        const dataList = $('#networkDatalist'); dataList.empty();
        let allItems = [];
        uniqueCompanies.forEach(c => allItems.push({{name: c, type: 'Company'}}));
        uniqueGroups.forEach(g => allItems.push({{name: g, type: 'Group'}}));
        allItems.sort((a,b) => a.name.localeCompare(b.name));
        allItems.forEach(item => {{ dataList.append(`<option value="${{item.name}}">${{item.type}}</option>`); }});

        $('#networkModal').show(); $('#networkMessage').show();
        
        const chartDom = document.getElementById('networkContainer');
        if(networkChartInst) networkChartInst.dispose();
        networkChartInst = echarts.init(chartDom);
        networkChartInst.setOption({{ title: {{ text: '' }}, series: [] }});
        window.addEventListener('resize', () => networkChartInst.resize());
        
        window.searchNetworkNode = function() {{
            const val = $('#networkSearchInput').val(); 
            if(!val) return;
            $('#networkMessage').hide();
            let subsetEdges = [];
            let isGroupSearch = uniqueGroups.has(val);
            let isCompanySearch = uniqueCompanies.has(val);
            if (isGroupSearch) {{
                let groupEdges = globalNetworkEdges.filter(e => e.g === val);
                groupEdges.sort((a,b) => b.w - a.w);
                subsetEdges = groupEdges;
            }} else if (isCompanySearch) {{
                subsetEdges = globalNetworkEdges.filter(e => e.c === val);
            }} else {{ 
                alert("Name not found in current view."); 
                return; 
            }}
            let localNodeStats = {{}};
            subsetEdges.forEach(e => {{
                if(!localNodeStats[e.c]) localNodeStats[e.c] = 0; 
                localNodeStats[e.c] += e.w;
                if(!localNodeStats[e.g]) localNodeStats[e.g] = 0; 
                localNodeStats[e.g] += e.w;
            }});





            let companyTopMembers = {{}}; 
            let groupTopCompanies = {{}}; 
            subsetEdges.forEach(e => {{
                if(e.p && e.p !== "Unknown Name") {{
                    if (!companyTopMembers[e.c]) companyTopMembers[e.c] = {{}};
                    if (!companyTopMembers[e.c][e.p]) companyTopMembers[e.c][e.p] = 0;
                    companyTopMembers[e.c][e.p] += e.w;
                }}
            }});


            let visibleGroups = new Set(subsetEdges.map(e => e.g));
            globalNetworkEdges.forEach(e => {{
                if (visibleGroups.has(e.g)) {{
                    if (!groupTopCompanies[e.g]) groupTopCompanies[e.g] = {{}};
                    if (!groupTopCompanies[e.g][e.c]) groupTopCompanies[e.g][e.c] = 0;
                    groupTopCompanies[e.g][e.c] += e.w;
                }}
            }});

            function getDescHtml(title, itemsMap) {{
                if(!itemsMap) return "No details available";
                let sorted = Object.entries(itemsMap).sort((a,b) => b[1] - a[1]).slice(0, 5);
                let html = `<div style="font-weight:bold; border-bottom:1px solid #777; margin-bottom:5px; padding-bottom:3px;">${{title}}</div>`;
                sorted.forEach(x => {{ html += `<div style="display:flex; justify-content:space-between; font-size:11px;"><span>${{x[0]}}</span><span style="font-weight:bold;">${{x[1]}}</span></div>`; }});
                if(sorted.length === 0) html += `<div style="font-size:11px;">No data</div>`;
                return html;
            }}


            let nodesMap = new Map(); 
            let links = [];

            

            subsetEdges.forEach(e => {{
                const isSE = e.c.toLowerCase().includes('schneider') || e.c.toLowerCase() === 'se';
                
                if(!nodesMap.has('C:'+e.c)) {{
                    let val = localNodeStats[e.c] || 1;
                    let size = Math.min(70, Math.max(20, Math.log(val + 1) * 15)); 
                    let tooltiphtml = getDescHtml("Top Members (In this View)", companyTopMembers[e.c]);
                    
                    nodesMap.set('C:'+e.c, {{ 
                        id: 'C:'+e.c, 
                        name: e.c, 
                        category: isSE ? 0 : 1, 
                        symbolSize: size, 
                        value: val, 
                        desc: tooltiphtml 
                    }});
                }}
                if(!nodesMap.has('G:'+e.g)) {{
                    let val = localNodeStats[e.g] || 1;
                    let size = Math.min(70, Math.max(20, Math.log(val + 1) * 15));
                    let tooltiphtml = getDescHtml("Top 5 Companies", groupTopCompanies[e.g]);
                    
                    nodesMap.set('G:'+e.g, {{ 
                        id: 'G:'+e.g, 
                        name: e.g, 
                        category: 2, 
                        symbolSize: size, 
                        value: val, 
                        desc:tooltiphtml 
                    }});
                }}
                
                links.push({{ source: 'C:'+e.c, target: 'G:'+e.g }});
            }});


            const nodes = Array.from(nodesMap.values());
            const option = {{
                tooltip: {{ 
                    trigger: 'item', 
                    formatter: function(params) {{
                        if (params.dataType === 'node') {{ 
                            return `<div style="text-align:left;"><strong>${{params.data.name}}</strong><br/><br/>${{params.data.desc}}</div>`; 
                        }}
                        return params.name; 
                    }} 
                }},
                legend: {{ data: [{{name: 'Schneider Electric'}}, {{name: 'Competitors'}}, {{name: 'Groups'}}] }},
                series: [{{
                    type: 'graph', 
                    layout: 'force', 
                    data: nodes, 
                    links: links,
                    categories: [
                        {{ name: 'Schneider Electric', itemStyle: {{ color: '#3DCD58' }} }}, 
                        {{ name: 'Competitors', itemStyle: {{ color: '#e74c3c' }} }}, 
                        {{ name: 'Groups', itemStyle: {{ color: '#2d3436' }} }}
                    ],
                    roam: true, 
                    label: {{ show: true, position: 'right' }}, 
                    force: {{ repulsion: 800, edgeLength: 150, gravity: 0.1 }}
                }}]
            }};

            networkChartInst.setOption(option, true);
        }};
        window.resetNetworkMap = function() {{ $('#networkSearchInput').val(''); $('#networkMessage').show(); networkChartInst.clear(); }};
    }}

    

    function calculatePower() {{
        if(!isStdMode) return;
        const idxs = getIndices();
        if (idxs.comp === -1 || idxs.func === -1) {{ alert("Missing columns."); return; }}
        const roleDefs = [{{ key: 'Secretary', color: '#6c5ce7' }}, {{ key: 'Assistant Sec', color: '#a29bfe' }}, {{ key: 'Chair', color: '#e17055' }}, {{ key: 'Vice-Chair', color: '#fab1a0' }}, {{ key: 'Convenor', color: '#ffeaa7' }}, {{ key: 'Liaison', color: '#74b9ff' }}, {{ key: 'Nat. Part', color: '#00cec9' }}, {{ key: 'Member', color: '#55efc4' }}, {{ key: 'Other', color: '#b2bec3' }}];
        let stats = {{}};
        filteredIndices.forEach(i => {{
            const row = fullData[i]; let comp = (row[idxs.comp]||"").trim();
            if(!comp) return;
            if(!stats[comp]) {{ stats[comp] = {{ score:0, breakdown: {{}} }}; roleDefs.forEach(r => stats[comp].breakdown[r.key] = 0); }}
            const funcStr = (row[idxs.func]||"").toLowerCase(); const w = getWeight(row[idxs.func]);
            let type = 'Other';
            if (funcStr.includes('secretary')) type = 'Secretary'; else if (funcStr.includes('assistant sec')) type = 'Assistant Sec'; else if (funcStr.includes('vice-chair')) type = 'Vice-Chair'; else if (funcStr.includes('chair')) type = 'Chair'; else if (funcStr.includes('convenor')) type = 'Convenor'; else if (funcStr.includes('liaison')) type = 'Liaison'; else if (funcStr.includes('national participant')) type = 'Nat. Part'; else if (funcStr.includes('nc official') || funcStr.includes('observer') || funcStr.includes('member')) type = 'Member';
            stats[comp].score += w; stats[comp].breakdown[type]++;
        }});
        const sorted = Object.entries(stats).sort((a,b) => b[1].score - a[1].score);
        powerExportData = sorted.map((item, i) => [i+1, item[0], item[1].score]);
        const top20 = sorted.slice(0, 20);
        let html = '<table id="powerTable" class="dataTable"><thead><tr><th>Rank</th><th>Company</th><th>Score</th><th>Action</th></tr></thead><tbody>';
        top20.forEach((item, i) => {{ html += `<tr><td>${{i+1}}</td><td><strong>${{item[0]}}</strong></td><td><span class="badge badge-green">${{item[1].score}}</span></td><td><button class="btn btn-link" onclick="addFilter('${{item[0].replace(/'/g, "\\\\'")}}', ${{idxs.comp}}, 'Company')">Filter</button></td></tr>`; }});
        let btnHtml = `<div style="text-align:right; margin-bottom:10px;"><button class="btn btn-secondary" onclick="exportDataToExcel(['Rank', 'Company', 'Score'], powerExportData, 'Power_Rankings')"><i class="fas fa-file-download"></i> Export Full List</button></div>`;
        $('#powerTableContainer').html(btnHtml + html + '</tbody></table>'); makeModalTable('powerTable');
        if(powerChartInst) powerChartInst.destroy();
        const datasets = roleDefs.map(def => ({{ label: def.key, data: top20.map(i => i[1].breakdown[def.key]), backgroundColor: def.color }}));
        powerChartInst = new Chart(document.getElementById('powerChart'), {{ type: 'bar', data: {{ labels: top20.map(i => i[0]), datasets: datasets }}, options: {{ responsive: true, maintainAspectRatio: false, scales: {{ x: {{ stacked: true }}, y: {{ stacked: true }} }} }} }} );
        $('#powerModal').show();
    }}







    function calculateGeo() {{
        if(!isStdMode) return;
        const idxs = getIndices();
        if (idxs.nc === -1) {{ alert("No 'NC' column."); return; }}

        $('#geoModal').show();

        if (geoChartInst) geoChartInst.dispose();
        geoChartInst = echarts.init(document.getElementById('geoContainer'));

        const countryStats = {{}};
        geoHeatmapMax = 0;

        filteredIndices.forEach(i => {{
            const row = fullData[i];
            const code = (row[idxs.nc]||"").trim().toUpperCase();
            if(!code || code.length > 3) return;

            const mapName = isoMap[code] || code;
            const w = getWeight(row[idxs.func]);

            if(!countryStats[mapName]) countryStats[mapName] = {{ total: 0, comps: {{}} }};
            countryStats[mapName].total += w;
            geoHeatmapMax = Math.max(geoHeatmapMax, countryStats[mapName].total);

            if(idxs.comp !== -1) {{
                const comp = (row[idxs.comp]||"").trim();
                if(comp && comp.toLowerCase() !== 'unknown') {{
                    if(!countryStats[mapName].comps[comp]) countryStats[mapName].comps[comp] = 0;
                    countryStats[mapName].comps[comp] += w;
                }}
            }}
        }});




        const mapSeriesData = [];
        Object.keys(countryStats).forEach(country => {{
            const d = countryStats[country];
            const top = Object.entries(d.comps).sort((a,b)=>b[1]-a[1]).slice(0,5);

            let top5Html = top.map(([c,s]) =>
                `<div><span style="color:#dfe6e9; font-size:11px;">●</span> ${{c}}: <strong>${{s}}</strong></div>`
            ).join("");

            if(!top5Html) top5Html = "<div>No company details</div>";

            mapSeriesData.push({{ name: country, value: d.total, top5: top5Html }});
        }});

        


        if (!mapDataLoaded) {{
            $('#geoLoader').show();
            geoChartInst.showLoading();

            $.getJSON('https://s3-us-west-2.amazonaws.com/s.cdpn.io/95368/world.json', function(mapJson) {{
                echarts.registerMap('world', mapJson);
                mapDataLoaded = true;
                geoChartInst.hideLoading();
                $('#geoLoader').hide();
                renderGeoMap(mapSeriesData);
            }}).fail(function() {{
                alert("Could not load map data (internet required).");
                geoChartInst.hideLoading();
                $('#geoLoader').hide();
            }});
        }} else {{
            renderGeoMap(mapSeriesData);
        }}
    }}

    



    function renderGeoMap(data) {{
        geoChartInst.setOption({{
            title: {{ text: 'Power Heatmap', left: 'center' }},
            tooltip: {{
                trigger: 'item',
                formatter: function(params) {{
                    if(!params.data) return params.name;
                    return `
                        <div style="font-size:14px; font-weight:bold; margin-bottom:5px;">${{params.name}}</div>
                        <div style="border-bottom:1px solid #555; padding-bottom:5px; margin-bottom:5px;">
                            Total Power: <span style="color:#3DCD58; font-weight:bold;">${{params.value}}</span>
                        </div>
                        <div style="font-size:12px; color:#ccc; margin-bottom:3px;">Top 5 Companies:</div>
                        ${{params.data.top5}}
                    `;
                }}
            }},
            visualMap: {{
                left: 20, bottom: 40, min: 0, max: geoHeatmapMax,
                text: ['High', 'Low'], calculable: true,
                inRange: {{ color: ['#e0f2f1', '#3DCD58', '#004d2e'] }}
            }},
            series: [{{
                type: 'map', mapType: 'world',
                roam: true,
                emphasis: {{ label: {{ show: true }} }},
                data: data
            }}]
        }}, true);

        window.addEventListener('resize', () => geoChartInst.resize());
    }}

    


    function calculateLandscape() {{
        if(!isStdMode) return;
        const idxs = getIndices();
        if (idxs.comp === -1 || idxs.grp === -1) {{ alert("Missing columns."); return; }}
        const grps = {{}};
        filteredIndices.forEach(i => {{
            const row = fullData[i]; const grp = (row[idxs.grp]||"").trim(); let comp = (row[idxs.comp]||"").trim(); const desc = idxs.desc !== -1 ? (row[idxs.desc]||"") : "";
            if(!grp) return;
            if(!grps[grp]) grps[grp] = {{ se:0, comps:{{}}, desc: desc }};
            let w = getWeight(row[idxs.func]);
            if(comp.toLowerCase().includes('schneider') || comp.toLowerCase() === 'se') grps[grp].se += w;
            else if(comp.length > 0) {{ if(!grps[grp].comps[comp]) grps[grp].comps[comp] = 0; grps[grp].comps[comp] += w; }}
        }});
        let items = [];
        Object.keys(grps).forEach(g => {{ 
            let maxC = 0, topC = "None"; Object.entries(grps[g].comps).forEach(([c, s]) => {{ if(s > maxC) {{ maxC = s; topC = c; }} }});
            let diff = grps[g].se - maxC;
            if (Math.abs(diff) >= SIGNIFICANT_GAP_THRESHOLD) items.push({{ grp: g, desc: grps[g].desc, competitor: topC, seScore: grps[g].se, cScore: maxC, diff: diff }});
        }});
        items.sort((a,b) => Math.abs(b.diff) - Math.abs(a.diff));
        landscapeExportData = items.map(item => [item.grp, item.desc, item.competitor, item.seScore, item.cScore, item.diff, (item.diff > 0 ? 'Dominating' : 'High Risk')]);
        let html = '<table id="landscapeTable" class="dataTable"><thead><tr><th>Group</th><th>Top Competitor</th><th>Score Comparison</th><th>Status</th></tr></thead><tbody>';
        items.forEach(item => {{ 
            let isWin = item.diff > 0; let statusStyle = isWin ? 'color:#3DCD58; font-weight:bold;' : 'color:#d63031; font-weight:bold;';
            html += `<tr><td><strong>${{item.grp}}</strong><br><span style="font-size:11px;color:#888">${{item.desc}}</span></td><td>${{item.competitor}}</td><td>SE: <strong>${{item.seScore}}</strong> vs ${{item.cScore}}</td><td style="${{statusStyle}}">${{isWin?'DOMINATING':'HIGH RISK'}} (${{(isWin?'+':'')+item.diff}})</td></tr>`; 
        }});
        let btnHtml = `<div style="text-align:right; margin-bottom:10px;"><button class="btn btn-secondary" onclick="exportDataToExcel(['Group', 'Description', 'Top Competitor', 'SE Score', 'Competitor Score', 'Gap', 'Status'], landscapeExportData, 'Competitive_Landscape')"><i class="fas fa-file-download"></i> Export Full List</button></div>`;
        $('#landscapeTableContainer').html(btnHtml + html + '</tbody></table>'); makeModalTable('landscapeTable'); $('#landscapeModal').show();
    }}

    function calculateBench() {{
    if(!isStdMode) return;
    const idxs = getIndices();
    if (idxs.lname === -1 && idxs.fname === -1) {{ alert("Missing Name columns."); return; }}
    if (idxs.nc === -1) {{ alert("Missing NC column."); return; }}
    if (idxs.comp === -1 || idxs.func === -1) {{ alert("Missing Company/Role columns."); return; }}

    function filterSchneiderAndCountry(code) {{
        addFilter('schneider', idxs.comp, 'Company');
       

        addFilter(code, idxs.nc, 'NC');
    }}

    const countryStats = {{}};
    let benchGeoMax = 0;

    filteredIndices.forEach(i => {{
        const row = fullData[i];
        const comp = (row[idxs.comp] || "").toLowerCase();
        if(!(comp.includes('schneider') || comp === 'se')) return;

        const code = (row[idxs.nc] || "").trim().toUpperCase();
        if(!code || code.length > 3) return;

        const mapName = isoMap[code] || code;
        const person = getFullName(row, idxs);
        const w = getWeight(row[idxs.func]);

        if(!countryStats[mapName]) countryStats[mapName] = {{ total: 0, people: {{}}, code: code }};
        countryStats[mapName].total += w;
        benchGeoMax = Math.max(benchGeoMax, countryStats[mapName].total);

        if(person && person !== "Unknown Name" && w > 0) {{
            if(!countryStats[mapName].people[person]) countryStats[mapName].people[person] = 0;
            countryStats[mapName].people[person] += w;
        }}
    }});

    


    const mapSeriesData = [];
    const personTotals = {{}};

    Object.keys(countryStats).forEach(countryName => {{
        const d = countryStats[countryName];

    
        
        const topPeople = Object.entries(d.people)
            .sort((a,b) => b[1] - a[1])
            .slice(0, 5);

        topPeople.forEach(([p, s]) => {{
            personTotals[p] = (personTotals[p] || 0) + s;
        }});

        let peopleHtml = topPeople.map(([p,s]) =>
            `<div style="display:flex; justify-content:space-between; gap:12px;">
                <span style="white-space:nowrap; overflow:hidden; text-overflow:ellipsis; max-width:260px;">${{p}}</span>
                <strong>${{s}}</strong>
             </div>`
        ).join("");

        if(!peopleHtml) peopleHtml = "<div>No people details</div>";

        
        mapSeriesData.push({{
            name: countryName,
            value: d.total,
            code: d.code,
            peopleHtml: peopleHtml
        }});
    }});

    const sortedPeople = Object.entries(personTotals).sort((a,b) => b[1] - a[1]).slice(0, 20);
    benchExportData = sortedPeople.map(([name, power]) => [name, power]);

    

    let html = `
      <table id="benchTable" class="dataTable">
        <thead><tr>
          <th>Expert Name</th>
          <th>Total Strength</th>
          <th>Action</th>
        </tr></thead>
        <tbody>
    `;

    sortedPeople.forEach(([name, power]) => {{
        const safe = name.replace(/'/g, "\\\\'");
        html += `
          <tr>
            <td><strong>${{name}}</strong></td>
            <td><strong>${{power}}</strong></td>
            <td>
              <button class="btn btn-link" onclick="addFilter('${{safe}}', 'all', 'All')">Search Person</button>
            </td>
          </tr>
        `;
    }});

    

    html += `</tbody></table>`;

    const btnHtml = `
      <div style="text-align:right; margin-bottom:10px;">
        <button class="btn btn-secondary"
          onclick="exportDataToExcel(['Expert Name','Total Strength'], benchExportData, 'Expert_Workload_SE')">
          <i class="fas fa-file-download"></i> Export Full List
        </button>
      </div>
    `;

    $('#benchTableContainer').html(btnHtml + html);
    makeModalTable('benchTable');

    


    $('#benchModal').show();

    if (benchGeoChartInst) benchGeoChartInst.dispose();
    benchGeoChartInst = echarts.init(document.getElementById('benchGeoContainer'));

    function renderBenchGeoMap() {{
        benchGeoChartInst.setOption({{
            title: {{ text: 'Schneider Expert Strength by Country (NC)', left: 'center' }},
            tooltip: {{
                trigger: 'item',
                formatter: function(params) {{
                    if(!params.data) return params.name;

                    return `
                        <div style="font-size:14px; font-weight:bold; margin-bottom:6px;">${{params.name}}</div>
                        <div style="margin-bottom:8px;">
                            Total Strength: <span style="color:#3DCD58; font-weight:bold;">${{params.value}}</span>
                        </div>
                        <div style="font-size:12px; color:#ccc; margin-bottom:4px;">Top People:</div>
                        ${{params.data.peopleHtml}}
                        <div style="margin-top:8px; font-size:11px; color:#aaa;"><i class="fas fa-mouse-pointer"></i> Double-click to filter</div>
                    `;
                }}
            }},
            visualMap: {{
                left: 20, bottom: 40,
                min: 0, max: benchGeoMax,
                text: ['High', 'Low'],
                calculable: true,
                inRange: {{ color: ['#e0f2f1', '#3DCD58', '#004d2e'] }}
            }},
            series: [{{
                type: 'map',
                map: 'world',
                roam: true,
                emphasis: {{ label: {{ show: true }} }},
                data: mapSeriesData
            }}]
        }}, true);

        benchGeoChartInst.off('dblclick');
        benchGeoChartInst.on('dblclick', function(params) {{
            if(!params.data) return;
            const code = params.data.code || (nameToIso[params.name.toLowerCase()] || "");
            if(code) filterSchneiderAndCountry(code);
        }});

        window.addEventListener('resize', () => benchGeoChartInst.resize());
    }}

    if (!mapDataLoaded) {{
        benchGeoChartInst.showLoading();
        $.getJSON('https://s3-us-west-2.amazonaws.com/s.cdpn.io/95368/world.json', function(mapJson) {{
            echarts.registerMap('world', mapJson);
            mapDataLoaded = true;
            benchGeoChartInst.hideLoading();
            renderBenchGeoMap();
        }}).fail(function() {{
            benchGeoChartInst.hideLoading();
            alert("Could not load map data (internet required).");
        }});
    }} else {{
        renderBenchGeoMap();
    }}
}}

    
    // VOTING PREDICTOR 
    const CENELEC_MEMBERS = [
        {{ name:"France",                      iso:"FR", weight:29, eea:true,  type:"Blue"   }},
        {{ name:"Germany",                     iso:"DE", weight:29, eea:true,  type:"Blue"   }},
        {{ name:"Italy",                       iso:"IT", weight:29, eea:true,  type:"Blue"   }},
        {{ name:"Türkiye",                     iso:"TR", weight:29, eea:false, type:"Red"    }},
        {{ name:"United Kingdom",              iso:"GB", weight:29, eea:false, type:"Yellow" }},
        {{ name:"Poland",                      iso:"PL", weight:27, eea:true,  type:"Blue"   }},
        {{ name:"Spain",                       iso:"ES", weight:27, eea:true,  type:"Blue"   }},
        {{ name:"Romania",                     iso:"RO", weight:14, eea:true,  type:"Blue"   }},
        {{ name:"Netherlands",                 iso:"NL", weight:13, eea:true,  type:"Blue"   }},
        {{ name:"Belgium",                     iso:"BE", weight:12, eea:true,  type:"Blue"   }},
        {{ name:"Czech Republic",              iso:"CZ", weight:12, eea:true,  type:"Blue"   }},
        {{ name:"Greece",                      iso:"GR", weight:12, eea:true,  type:"Blue"   }},
        {{ name:"Hungary",                     iso:"HU", weight:12, eea:true,  type:"Blue"   }},
        {{ name:"Portugal",                    iso:"PT", weight:12, eea:true,  type:"Blue"   }},
        {{ name:"Austria",                     iso:"AT", weight:10, eea:true,  type:"Blue"   }},
        {{ name:"Bulgaria",                    iso:"BG", weight:10, eea:true,  type:"Blue"   }},
        {{ name:"Sweden",                      iso:"SE", weight:10, eea:true,  type:"Blue"   }},
        {{ name:"Switzerland",                 iso:"CH", weight:10, eea:false, type:"Red"    }},
        {{ name:"Croatia",                     iso:"HR", weight:7,  eea:true,  type:"Blue"   }},
        {{ name:"Denmark",                     iso:"DK", weight:7,  eea:true,  type:"Blue"   }},
        {{ name:"Finland",                     iso:"FI", weight:7,  eea:true,  type:"Blue"   }},
        {{ name:"Ireland",                     iso:"IE", weight:7,  eea:true,  type:"Blue"   }},
        {{ name:"Lithuania",                   iso:"LT", weight:7,  eea:true,  type:"Blue"   }},
        {{ name:"Norway",                      iso:"NO", weight:7,  eea:true,  type:"Blue"   }},
        {{ name:"Slovakia",                    iso:"SK", weight:7,  eea:true,  type:"Blue"   }},
        {{ name:"Serbia",                      iso:"RS", weight:7,  eea:false, type:"Red"    }},
        {{ name:"Cyprus",                      iso:"CY", weight:4,  eea:true,  type:"Blue"   }},
        {{ name:"Estonia",                     iso:"EE", weight:4,  eea:true,  type:"Blue"   }},
        {{ name:"Republic of North Macedonia", iso:"MK", weight:4,  eea:false, type:"Red"    }},
        {{ name:"Latvia",                      iso:"LV", weight:4,  eea:true,  type:"Blue"   }},
        {{ name:"Luxembourg",                  iso:"LU", weight:4,  eea:true,  type:"Blue"   }},
        {{ name:"Slovenia",                    iso:"SI", weight:4,  eea:true,  type:"Blue"   }},
        {{ name:"Iceland",                     iso:"IS", weight:3,  eea:true,  type:"Blue"   }},
        {{ name:"Malta",                       iso:"MT", weight:3,  eea:true,  type:"Blue"   }},
    ];

    // Build ISO→member lookup (also handles "UK"→GB)
    const ISO_TO_MEMBER = {{}};
    CENELEC_MEMBERS.forEach(m => {{ ISO_TO_MEMBER[m.iso] = m; }});
    ISO_TO_MEMBER['UK'] = ISO_TO_MEMBER['GB'];

    // State
    let currentVotingMode = 'IEC';
    let countryVotes  = {{}};   // country name → 'yes'|'no'|'abstain'
    let ncInfluence   = {{}};   // country name → {{ sePct, sePts, totalPts, topCompetitor, topCompPct, companyBreakdown, participates }}
    let currentVotingCommittee = '';
    let iecExtraMembers = [];   // Non-CENELEC countries found in data (IEC-only)

    // Helpers 
    function isSECompany(name) {{
        const n = (name || '').toLowerCase();
        return n.includes('schneider') || n === 'se';
    }}

    function seThreshold() {{
        return 50;
    }}

    function autoPredictVote(inf) {{
        if (!inf || !inf.participates) return 'abstain';
        
        // 1. SE Domination -> Push for YES
        if (inf.sePct >= seThreshold()) return 'yes';
        
        // 2. No Consensus / Lack of Dominance -> Abstain
        return 'abstain';
    }}

    // Load data from selected committee 
    window.loadVotingFromCommittee = function() {{
        const committee = document.getElementById('votingCommitteeSelect').value;
        currentVotingCommittee = committee;
        ncInfluence = {{}};
        countryVotes = {{}};

        if (!committee) {{
            renderVotingCountryList();
            renderVotingResults();
            return;
        }}

        const idxs = getIndices();

        // Build per-NC power map for this committee
        const ncMap = {{}};
        filteredIndices.forEach(i => {{
            const row = fullData[i];
            const grp = (row[idxs.grp] || '').trim();
            if (grp !== committee) return;
            const ncCode = (row[idxs.nc] || '').trim().toUpperCase();
            const comp   = (row[idxs.comp] || '').trim();
            const w      = getWeight(row[idxs.func]);
            if (!ncCode || !comp || w === 0) return;
            const member = ISO_TO_MEMBER[ncCode];
            // For CENELEC members use the full country name; for IEC-only use the ISO code as key
            const key = member ? member.name : ncCode;

            if (!ncMap[key]) ncMap[key] = {{ total:0, se:0, companies:{{}}, iso: ncCode, isCENELEC: !!member }};
            ncMap[key].total += w;
            if (isSECompany(comp)) ncMap[key].se += w;
            ncMap[key].companies[comp] = (ncMap[key].companies[comp] || 0) + w;
        }});

        // Build ncInfluence for all CENELEC members
        CENELEC_MEMBERS.forEach(m => {{
            const d = ncMap[m.name];
            if (!d) {{
                ncInfluence[m.name] = {{ participates:false, sePct:0, sePts:0, totalPts:0, topCompetitor:'', topCompPct:0, companyBreakdown:[] }};
            }} else {{
                const sePct = d.total > 0 ? (d.se / d.total * 100) : 0;
                // Top non-SE competitor
                const competitors = Object.entries(d.companies)
                    .filter(([c]) => !isSECompany(c))
                    .sort((a,b) => b[1]-a[1]);
                const topComp    = competitors.length ? competitors[0][0] : '';
                const topCompPct = competitors.length && d.total > 0 ? (competitors[0][1] / d.total * 100) : 0;
                // Full breakdown sorted
                const breakdown = Object.entries(d.companies)
                    .sort((a,b) => b[1]-a[1])
                    .map(([c,pts]) => ({{ company:c, pts, pct: d.total>0?(pts/d.total*100):0, isSE: isSECompany(c) }}));
                ncInfluence[m.name] = {{ participates:true, sePct, sePts:d.se, totalPts:d.total, topCompetitor:topComp, topCompPct, companyBreakdown:breakdown }};
            }}
            countryVotes[m.name] = autoPredictVote(ncInfluence[m.name]);
        }});

        // Build ncInfluence for IEC-only (non-CENELEC) countries found in data
        // These appear in IEC votes but are outside the CENELEC weighted system
        iecExtraMembers = [];
        const cenelecNames = new Set(CENELEC_MEMBERS.map(m => m.name));
        Object.keys(ncMap).forEach(key => {{
            if (cenelecNames.has(key)) return;  // already handled above
            const d = ncMap[key];
            const sePct = d.total > 0 ? (d.se / d.total * 100) : 0;
            const competitors = Object.entries(d.companies)
                .filter(([c]) => !isSECompany(c))
                .sort((a,b) => b[1]-a[1]);
            const topComp    = competitors.length ? competitors[0][0] : '';
            const topCompPct = competitors.length && d.total > 0 ? (competitors[0][1] / d.total * 100) : 0;
            const breakdown  = Object.entries(d.companies)
                .sort((a,b) => b[1]-a[1])
                .map(([c,pts]) => ({{ company:c, pts, pct: d.total>0?(pts/d.total*100):0, isSE: isSECompany(c) }}));
            ncInfluence[key] = {{ participates:true, sePct, sePts:d.se, totalPts:d.total, topCompetitor:topComp, topCompPct, companyBreakdown:breakdown }};
            countryVotes[key] = autoPredictVote(ncInfluence[key]);
            iecExtraMembers.push({{ name:key, iso:d.iso }});
        }});

        renderVotingCountryList();
        renderVotingResults();
    }};

    window.reAutoPredict = function() {{
        CENELEC_MEMBERS.forEach(m => {{
            countryVotes[m.name] = autoPredictVote(ncInfluence[m.name]);
        }});
        iecExtraMembers.forEach(m => {{
            countryVotes[m.name] = autoPredictVote(ncInfluence[m.name]);
        }});
        renderVotingCountryList();
        renderVotingResults();
    }};

    // Mode toggle 
    window.setVotingMode = function(mode) {{
        currentVotingMode = mode;
        document.getElementById('modeIEC').style.background     = mode==='IEC'     ? '#3DCD58' : 'transparent';
        document.getElementById('modeIEC').style.color          = mode==='IEC'     ? 'white'   : '#636e72';
        document.getElementById('modeCENELEC').style.background = mode==='CENELEC' ? '#3DCD58' : 'transparent';
        document.getElementById('modeCENELEC').style.color      = mode==='CENELEC' ? 'white'   : '#636e72';
        renderVotingCountryList();
        renderVotingResults();
    }};

    // Set vote for one country 
    window.setVote = function(country, vote) {{
        countryVotes[country] = vote;
        const rowId = 'vrow-' + country.replace(/[^a-z0-9]/gi,'_');
        const row = document.getElementById(rowId);
        if (row) updateVoteButtons(row, vote);
        renderVotingResults();
    }};

    function updateVoteButtons(rowEl, vote) {{
        rowEl.querySelectorAll('.vote-btn').forEach(btn => {{
            const v = btn.getAttribute('data-vote');
            const active = v === vote;
            if (active) {{
                if (v==='yes')     {{ btn.style.background='#27ae60'; btn.style.color='white'; btn.style.borderColor='#27ae60'; btn.style.fontWeight='700'; btn.style.opacity='1'; }}
                else if (v==='no') {{ btn.style.background='#c0392b'; btn.style.color='white'; btn.style.borderColor='#c0392b'; btn.style.fontWeight='700'; btn.style.opacity='1'; }}
                else               {{ btn.style.background='#7f8c8d'; btn.style.color='white'; btn.style.borderColor='#7f8c8d'; btn.style.fontWeight='700'; btn.style.opacity='1'; }}
            }} else {{
                btn.style.background='white'; btn.style.color='#636e72'; btn.style.borderColor='#ddd'; btn.style.fontWeight='400'; btn.style.opacity='0.55';
            }}
        }});
    }}

    window.setAllVotingParticipants = function(vote) {{
        const applyTo = (m) => {{
            if (vote === 'abstain' && ncInfluence[m.name]) {{
                countryVotes[m.name] = autoPredictVote(ncInfluence[m.name]);
            }} else {{
                countryVotes[m.name] = vote;
            }}
        }};
        CENELEC_MEMBERS.forEach(applyTo);
        iecExtraMembers.forEach(applyTo);
        renderVotingCountryList();
        renderVotingResults();
    }};

    // Render country list
    function renderVotingCountryList() {{
        const container = document.getElementById('votingCountryList');
        const noMsg     = document.getElementById('votingNoDataMsg');
        if (!container) return;

        const typeColors = {{ Blue:'#2471a3', Red:'#c0392b', Yellow:'#d68910' }};
        const typeBg     = {{ Blue:'#EBF5FB', Red:'#FDEDEC', Yellow:'#FEF9E7' }};

        const hasData = Object.keys(ncInfluence).length > 0;
        noMsg.style.display    = hasData ? 'none'  : 'block';
        container.style.display = hasData ? 'block' : 'none';
        if (!hasData) return;

        // Sort: participating first, then by sePct desc
        const cenelecSorted = [...CENELEC_MEMBERS].sort((a, b) => {{
            const ia = ncInfluence[a.name] || {{}};
            const ib = ncInfluence[b.name] || {{}};
            if (ia.participates && !ib.participates) return -1;
            if (!ia.participates && ib.participates)  return  1;
            return (ib.sePct||0) - (ia.sePct||0);
        }});

        // In IEC mode, also include non-CENELEC members sorted by sePct desc
        const iecSorted = currentVotingMode === 'IEC'
            ? [...iecExtraMembers].sort((a, b) => (ncInfluence[b.name]?.sePct||0) - (ncInfluence[a.name]?.sePct||0))
            : [];

        const sorted = [...cenelecSorted, ...iecSorted];

        let html = '';
        sorted.forEach(m => {{
            const inf  = ncInfluence[m.name] || {{ participates:false, sePct:0, sePts:0, totalPts:0, topCompetitor:'', topCompPct:0, companyBreakdown:[] }};
            const vote = countryVotes[m.name] || 'abstain';
            const rowId = 'vrow-' + m.name.replace(/[^a-z0-9]/gi,'_');
            const dimmed = !inf.participates;

            // Type badge (CENELEC mode only)
            const typeBadge = currentVotingMode === 'CENELEC'
                ? `<span style="font-size:9px;padding:1px 5px;border-radius:8px;background:${{typeBg[m.type]}};color:${{typeColors[m.type]}};font-weight:700;flex-shrink:0;">${{m.type}}</span>`
                : '';
            const weightBadge = currentVotingMode === 'CENELEC'
                ? `<span style="font-size:10px;color:#aaa;flex-shrink:0;">w:${{m.weight}}</span>`
                : '';

            // SE influence bar
            let influenceHtml = '';
            if (inf.participates) {{
                const sePct  = inf.sePct.toFixed(0);
                const barColor = inf.sePct >= seThreshold() ? '#006039' : (inf.sePct > 0 ? '#e67e22' : '#e74c3c');
                const barWidth = Math.max(inf.sePct, 2).toFixed(1);
                const topCompHtml = inf.topCompetitor
                    ? `<span style="font-size:9.5px;color:#c0392b;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;max-width:110px;display:inline-block;vertical-align:bottom;" title="Top competitor: ${{inf.topCompetitor}} (${{inf.topCompPct.toFixed(0)}}%)">${{inf.topCompetitor.length > 14 ? inf.topCompetitor.slice(0,13)+'…' : inf.topCompetitor}} (${{inf.topCompPct.toFixed(0)}}%)</span>`
                    : `<span style="font-size:9.5px;color:#27ae60;">SE only</span>`;
                const seLabel = inf.sePts > 0
                    ? `<span style="font-size:9.5px;font-weight:700;color:${{barColor}};">SE ${{sePct}}%</span>`
                    : `<span style="font-size:9.5px;color:#aaa;">SE 0%</span>`;
                influenceHtml = `
                    <div style="display:flex;flex-direction:column;gap:2px;min-width:120px;max-width:120px;">
                        <div style="background:#eee;border-radius:3px;height:5px;overflow:hidden;">
                            <div style="width:${{barWidth}}%;height:100%;background:${{barColor}};border-radius:3px;"></div>
                        </div>
                        <div style="display:flex;justify-content:space-between;gap:4px;">${{seLabel}}${{topCompHtml}}</div>
                    </div>`;
            }} else {{
                influenceHtml = `<div style="min-width:120px;max-width:120px;font-size:10px;color:#ccc;font-style:italic;">not in committee</div>`;
            }}

            function bStyle(v) {{
                const active = vote === v;
                if (!active) return 'background:white;color:#636e72;border:1px solid #ddd;font-weight:400;opacity:0.55;';
                if (v==='yes')    return 'background:#27ae60;color:white;border:1px solid #27ae60;font-weight:700;opacity:1;';
                if (v==='no')     return 'background:#c0392b;color:white;border:1px solid #c0392b;font-weight:700;opacity:1;';
                return 'background:#7f8c8d;color:white;border:1px solid #7f8c8d;font-weight:700;opacity:1;';
            }}

            // Tooltip: full company breakdown
            const tooltipLines = inf.companyBreakdown
                ? inf.companyBreakdown.slice(0,6).map(d => `${{d.company}}: ${{d.pts}} pts (${{d.pct.toFixed(0)}}%)`).join('&#10;')
                : '';

            html += `<div id="${{rowId}}" title="${{tooltipLines}}" style="display:grid;grid-template-columns:1fr 120px 90px;gap:8px;align-items:center;padding:5px 12px;border-bottom:1px solid #f0f0f0;min-height:38px;opacity:${{dimmed?'0.45':'1'}};background:${{!dimmed && inf.sePct>=seThreshold()?'#f9fffe':'white'}};transition:background 0.2s;">
                <div style="display:flex;align-items:center;gap:5px;overflow:hidden;">
                    <span style="font-size:13px;font-weight:${{inf.participates?'600':'400'}};overflow:hidden;text-overflow:ellipsis;white-space:nowrap;">${{m.name}}</span>
                    ${{typeBadge}}${{weightBadge}}
                </div>
                ${{influenceHtml}}
                <div style="display:flex;gap:3px;justify-content:flex-end;flex-shrink:0;">
                    <button class="vote-btn" data-vote="yes"     onclick="setVote('${{m.name}}','yes')"     style="padding:2px 7px;border-radius:4px;cursor:pointer;font-size:11px;${{bStyle('yes')}}">✓</button>
                    <button class="vote-btn" data-vote="no"      onclick="setVote('${{m.name}}','no')"      style="padding:2px 7px;border-radius:4px;cursor:pointer;font-size:11px;${{bStyle('no')}}">✗</button>
                    <button class="vote-btn" data-vote="abstain" onclick="setVote('${{m.name}}','abstain')" style="padding:2px 7px;border-radius:4px;cursor:pointer;font-size:11px;${{bStyle('abstain')}}">—</button>
                </div>
            </div>`;
        }});
        container.innerHTML = html;
    }}

    // Results panel 
    function renderVotingResults() {{
        const container = document.getElementById('votingResultsPanel');
        if (!container) return;

        // Summary stats
        const participating = CENELEC_MEMBERS.filter(m => ncInfluence[m.name] && ncInfluence[m.name].participates);
        const seYes  = participating.filter(m => ncInfluence[m.name].sePct >= seThreshold()).length;
        const seRisk = participating.filter(m => ncInfluence[m.name].participates && ncInfluence[m.name].sePct === 0).length;

        function stepBadge(pass, pct, noData) {{
            if (noData) return '<span style="color:#888;font-size:11px;">No votes</span>';
            const bg = pass ? '#27ae60' : '#c0392b';
            return `<span style="background:${{bg}};color:white;padding:2px 8px;border-radius:10px;font-size:11px;font-weight:700;">${{pass?'PASS':'FAIL'}} (${{pct.toFixed(1)}}%)</span>`;
        }}

        function progressBar(pct, pass, threshold=50) {{
            const lineLeft = threshold;
            return `<div style="background:#eee;border-radius:4px;height:8px;margin:7px 0;position:relative;overflow:hidden;">
                <div style="width:${{Math.min(pct,100).toFixed(1)}}%;height:100%;background:${{pass?'#27ae60':'#e74c3c'}};border-radius:4px;transition:width 0.4s;"></div>
                <div style="position:absolute;top:0;left:${{lineLeft}}%;width:2px;height:100%;background:#2d3436;opacity:0.5;"></div>
            </div>
            <div style="display:flex;justify-content:space-between;font-size:10px;color:#aaa;">
                <span>0%</span><span style="color:#2d3436;font-weight:600;">${{lineLeft}}% threshold</span><span>100%</span>
            </div>`;
        }}

        function miniStats(yes, no, abs, label='') {{
            return `<div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:6px;margin-top:7px;text-align:center;">
                <div style="background:white;border-radius:6px;padding:6px;border:1px solid #eee;">
                    <div style="font-weight:700;font-size:15px;color:#27ae60;">${{yes}}</div>
                    <div style="font-size:9px;color:#888;">YES${{label}}</div>
                </div>
                <div style="background:white;border-radius:6px;padding:6px;border:1px solid #eee;">
                    <div style="font-weight:700;font-size:15px;color:#c0392b;">${{no}}</div>
                    <div style="font-size:9px;color:#888;">NO${{label}}</div>
                </div>
                <div style="background:white;border-radius:6px;padding:6px;border:1px solid #eee;">
                    <div style="font-weight:700;font-size:15px;color:#95a5a6;">${{abs}}</div>
                    <div style="font-size:9px;color:#888;">ABSTAIN${{label}}</div>
                </div>
            </div>`;
        }}

        // SE coverage summary block
        const hasCoverage = participating.length > 0;
        const coveragePct = hasCoverage ? (seYes / participating.length * 100) : 0;
        const coverageBlock = hasCoverage ? `
            <div style="background:#f8f9fa;border-radius:10px;padding:11px 13px;margin-bottom:12px;border:1px solid #edf2f7;">
                <div style="font-weight:700;font-size:12px;color:#2d3436;margin-bottom:6px;">🏢 SE Coverage — ${{currentVotingCommittee||'all'}}</div>
                <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:6px;text-align:center;">
                    <div style="background:white;border-radius:6px;padding:7px;border:1px solid #eee;">
                        <div style="font-size:16px;font-weight:700;color:#006039;">${{participating.length}}</div>
                        <div style="font-size:9px;color:#888;">NC participating</div>
                    </div>
                    <div style="background:white;border-radius:6px;padding:7px;border:1px solid #eee;">
                        <div style="font-size:16px;font-weight:700;color:#27ae60;">${{seYes}}</div>
                        <div style="font-size:9px;color:#888;">SE dominant</div>
                    </div>
                    <div style="background:white;border-radius:6px;padding:7px;border:1px solid #eee;">
                        <div style="font-size:16px;font-weight:700;color:#c0392b;">${{seRisk}}</div>
                        <div style="font-size:9px;color:#888;">SE absent</div>
                    </div>
                </div>
                <div style="margin-top:7px;background:#eee;border-radius:4px;height:6px;overflow:hidden;">
                    <div style="width:${{coveragePct.toFixed(1)}}%;height:100%;background:#006039;border-radius:4px;"></div>
                </div>
                <div style="font-size:10px;color:#888;margin-top:3px;">SE dominant in ${{coveragePct.toFixed(0)}}% of participating NCs</div>
            </div>` : '';

        if (currentVotingMode === 'IEC') {{
            let yes=0, no=0, abstain=0;
            // IEC = 1 country 1 vote — count all participants (CENELEC + IEC-only)
            const allIecMembers = [...CENELEC_MEMBERS, ...iecExtraMembers];
            allIecMembers.forEach(m => {{
                const v = countryVotes[m.name] || 'abstain';
                if (v==='yes') yes++; else if (v==='no') no++; else abstain++;
            }});
            const voting = yes+no;
            
            // IEC Math Update: >= 66.66% YES AND <= 25% NO
            const yesPct = voting>0 ? (yes/voting*100) : 0;
            const noPct = voting>0 ? (no/voting*100) : 0;
            const passed = (yesPct >= 66.66) && (noPct <= 25.00) && (voting > 0);
            
            const noVotes = voting === 0;
            const verdict = noVotes ? 'No votes cast' : (passed ? '✅ ADOPTED' : '❌ REJECTED');
            const vColor  = noVotes ? '#888' : (passed ? '#006039' : '#c0392b');
            const vBg     = noVotes ? '#f8f9fa' : (passed ? '#f0fdf4' : '#fff5f5');
            const vBorder = noVotes ? '#eee'   : (passed ? '#27ae60' : '#c0392b');

            container.innerHTML = `
                <div style="font-size:13px;font-weight:700;color:#2d3436;margin-bottom:10px;padding-bottom:7px;border-bottom:2px solid #edf2f7;">📊 IEC Result</div>
                ${{coverageBlock}}
                <div style="text-align:center;padding:16px;background:${{vBg}};border-radius:12px;margin-bottom:12px;border:2px solid ${{vBorder}};">
                    <div style="font-size:22px;font-weight:800;color:${{vColor}};">${{verdict}}</div>
                    ${{!noVotes?`<div style="font-size:11px;color:#888;margin-top:4px;">${{yes}} yes (${{yesPct.toFixed(1)}}%) vs ${{no}} no (${{noPct.toFixed(1)}}%) &nbsp;·&nbsp; ${{voting}} countries voted</div>`:''}}
                </div>
                ${{miniStats(yes,no,abstain,' countries')}}
                ${{progressBar(yesPct, passed, 66.66)}}
                <div style="margin-top:10px;background:#f8f9fa;border-radius:8px;padding:9px;font-size:11px;color:#636e72;line-height:1.5;">
                    <strong>IEC rule:</strong> ≥ 66.6% in favour AND ≤ 25% negative votes of countries voting.
                </div>`;

        }} else {{
            // CENELEC two-step
            let s1Yes=0, s1No=0, s1Abs=0, s2Yes=0, s2No=0, s2Abs=0;
            let s1CountryYes=0, s1CountryVoting=0, s2CountryYes=0, s2CountryVoting=0;

            CENELEC_MEMBERS.forEach(m => {{
                const v = countryVotes[m.name] || 'abstain';
                if (v==='yes') {{ 
                    s1Yes+=m.weight; s1CountryYes++; s1CountryVoting++;
                    if(m.type==='Blue') {{ s2Yes+=m.weight; s2CountryYes++; s2CountryVoting++; }} 
                }}
                else if (v==='no') {{ 
                    s1No +=m.weight; s1CountryVoting++;
                    if(m.type==='Blue') {{ s2No +=m.weight; s2CountryVoting++; }} 
                }}
                else {{ 
                    s1Abs+=m.weight; 
                    if(m.type==='Blue') s2Abs+=m.weight; 
                }}
            }});
            const s1Voting=s1Yes+s1No, s2Voting=s2Yes+s2No;
            const s1Pct = s1Voting>0 ? (s1Yes/s1Voting*100) : 0;
            const s2Pct = s2Voting>0 ? (s2Yes/s2Voting*100) : 0;
            
            // CENELEC Update: >= 71% points AND > 50% of countries voting
            const s1Pass = (s1Pct >= 71) && (s1CountryYes > s1CountryVoting / 2) && (s1Voting > 0);
            const s2Pass = (s2Pct >= 71) && (s2CountryYes > s2CountryVoting / 2) && (s2Voting > 0);

            let verdict, vColor, vBg, vDetail;
            if (s1Voting===0)     {{ verdict='No votes cast';             vColor='#888';    vBg='#f8f9fa'; vDetail=''; }}
            else if (s1Pass)      {{ verdict='✅ ADOPTED — All Members';  vColor='#006039'; vBg='#f0fdf4'; vDetail='Step 1 passed. Binding on all CENELEC members including non-EEA.'; }}
            else if (s2Pass)      {{ verdict='🔵 ADOPTED — EEA Only';    vColor='#2471a3'; vBg='#EBF5FB'; vDetail='Step 1 failed, EEA Safety Net passed. Binding on Blue (EEA) members only.'; }}
            else                  {{ verdict='❌ NOT ADOPTED';            vColor='#c0392b'; vBg='#fff5f5'; vDetail='Both Step 1 and the EEA Safety Net failed.'; }}

            const vBorderColor = s1Voting===0 ? '#eee' : vColor;

            container.innerHTML = `
                <div style="font-size:13px;font-weight:700;color:#2d3436;margin-bottom:10px;padding-bottom:7px;border-bottom:2px solid #edf2f7;">📊 CENELEC Result</div>
                ${{coverageBlock}}
                <div style="text-align:center;padding:14px;background:${{vBg}};border-radius:12px;margin-bottom:12px;border:2px solid ${{vBorderColor}};">
                    <div style="font-size:18px;font-weight:800;color:${{vColor}};">${{verdict}}</div>
                    ${{vDetail?`<div style="font-size:11px;color:#636e72;margin-top:5px;line-height:1.4;">${{vDetail}}</div>`:''}}
                </div>

                <div style="background:${{s1Pass?'#f0fdf4':'#f8f9fa'}};border-radius:10px;padding:11px;margin-bottom:9px;border:1px solid ${{s1Pass?'#a9dfbf':'#eee'}};">
                    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:3px;">
                        <span style="font-weight:700;font-size:12px;">Step 1 — All Members <span style="font-size:10px;color:#888;font-weight:400;">(6.2.1.1)</span></span>
                        ${{stepBadge(s1Pass, s1Pct, s1Voting===0)}}
                    </div>
                    <div style="font-size:10px;color:#636e72;margin-bottom:2px;">🔵 Blue + 🔴 Red + 🟡 Yellow — need ≥71% pts & >50% countries</div>
                    ${{progressBar(s1Pct, s1Pass, 71)}}
                    ${{miniStats(s1Yes, s1No, s1Abs, ' pts')}}
                </div>

                <div style="background:${{s1Pass?'#f8f9fa':(s2Pass?'#EBF5FB':'#f8f9fa')}};border-radius:10px;padding:11px;border:1px solid ${{s1Pass?'#eee':(s2Pass?'#aed6f1':'#eee')}};opacity:${{s1Pass?'0.4':'1'}};transition:opacity 0.3s;">
                    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:3px;">
                        <span style="font-weight:700;font-size:12px;">Step 2 — EEA Safety Net <span style="font-size:10px;color:#888;font-weight:400;">(6.2.1.2)</span></span>
                        ${{s1Pass
                            ? '<span style="color:#27ae60;font-size:11px;font-weight:600;">Not needed</span>'
                            : stepBadge(s2Pass, s2Pct, s2Voting===0)}}
                    </div>
                    <div style="font-size:10px;color:#636e72;margin-bottom:2px;">🔵 Blue only — need ≥71% pts & >50% countries</div>
                    ${{progressBar(s2Pct, s2Pass, 71)}}
                    ${{miniStats(s2Yes, s2No, s2Abs, ' pts')}}
                </div>

                <div style="margin-top:10px;background:#f8f9fa;border-radius:8px;padding:9px;font-size:10px;color:#636e72;line-height:1.5;">
                    <span style="display:inline-flex;align-items:center;background:#EBF5FB;color:#2471a3;padding:1px 7px;border-radius:8px;margin:0 2px;font-weight:600;">🔵 Blue</span> EU/EEA — both steps &nbsp;·&nbsp;
                    <span style="display:inline-flex;align-items:center;background:#FDEDEC;color:#c0392b;padding:1px 7px;border-radius:8px;margin:0 2px;font-weight:600;">🔴 Red</span> Non-EEA — Step 1 only &nbsp;·&nbsp;
                    <span style="display:inline-flex;align-items:center;background:#FEF9E7;color:#d68910;padding:1px 7px;border-radius:8px;margin:0 2px;font-weight:600;">🟡 Yellow</span> UK — Step 1 only
                </div>`;
        }}
    }}

    // entry
    window.calculateVoting = function() {{
        if (!isStdMode) return;
        const idxs = getIndices();
        if (idxs.nc === -1)   {{ alert("No 'NC' column found."); return; }}
        if (idxs.comp === -1) {{ alert("No 'Company' column found."); return; }}
        if (idxs.grp === -1)  {{ alert("No 'Group/Committee' column found."); return; }}

        // Populate committee dropdown
        const allGroups = new Set();
        filteredIndices.forEach(i => {{
            const grp = (fullData[i][idxs.grp] || '').trim();
            if (grp) allGroups.add(grp);
        }});
        const sel = document.getElementById('votingCommitteeSelect');
        sel.innerHTML = '<option value="">— select committee —</option>';
        Array.from(allGroups).sort().forEach(g => {{
            const opt = document.createElement('option');
            opt.value = g; opt.textContent = g;
            sel.appendChild(opt);
        }});

        // Reset state
        countryVotes = {{}};
        ncInfluence  = {{}};
        iecExtraMembers = [];
        currentVotingMode = 'IEC';
        CENELEC_MEMBERS.forEach(m => {{ countryVotes[m.name] = 'abstain'; }});

        $('#votingModal').show();
        setVotingMode('IEC');
        renderVotingCountryList();
        renderVotingResults();
    }};

    window.exportVotingResults = function() {{
        const cenelecRows = CENELEC_MEMBERS.map(m => {{
            const inf = ncInfluence[m.name] || {{}};
            return [
                m.name, m.iso, m.type, m.eea?'Yes':'No', m.weight,
                inf.participates?'Yes':'No',
                inf.participates ? inf.sePct.toFixed(1)+'%' : 'N/A',
                inf.participates ? inf.totalPts : 'N/A',
                inf.topCompetitor || '',
                countryVotes[m.name] || 'abstain'
            ];
        }});
        const iecRows = iecExtraMembers.map(m => {{
            const inf = ncInfluence[m.name] || {{}};
            return [
                m.name, m.iso, 'IEC-only', 'N/A', 1,
                'Yes',
                inf.sePct.toFixed(1)+'%',
                inf.totalPts,
                inf.topCompetitor || '',
                countryVotes[m.name] || 'abstain'
            ];
        }});
        exportDataToExcel(
            ['Country','ISO','Type','EEA','CENELEC Weight','In Committee','SE Influence','Total NC Pts','Top Competitor','Predicted Vote'],
            [...cenelecRows, ...iecRows], 'Voting_Prediction_' + (currentVotingCommittee||'all').replace(/[^a-z0-9]/gi,'_')
        );
    }};

    function closeModal(id) {{ document.getElementById(id).style.display = 'none'; }}
    window.onclick = (e) => {{ if($(e.target).hasClass('modal')) $('.modal').hide(); }};

    //INTERACTIVE CHART EXPORT
    window.exportInteractiveChart = function(engine, chartInstance, filename) {{
        if (!chartInstance) {{
            alert("Chart is not fully loaded yet.");
            return;
        }}

        let chartConfigStr = "";
        let libraryScript = "";
        let initScript = "";

        // ECharts
        if (engine === 'echarts') {{
            const currentOption = chartInstance.getOption();
            chartConfigStr = JSON.stringify(currentOption);
            libraryScript = '<script src="https://cdn.jsdelivr.net/npm/echarts@5.4.3/dist/echarts.min.js"></' + 'script>';
            
            
            let tooltipPatch = "";
            if (filename.includes('Network')) {{
                tooltipPatch = `option.tooltip = {{ trigger: 'item', formatter: function(params) {{ if (params.dataType === 'node') return "<div style='text-align:left;'><strong>" + params.data.name + "</strong><br/><br/>" + (params.data.desc || "") + "</div>"; return params.name; }} }};`;
            }} else if (filename.includes('Geo') || filename.includes('Workload')) {{
                tooltipPatch = `option.tooltip = {{ trigger: 'item', formatter: function(params) {{ if(!params.data) return params.name; return "<div style='font-size:14px; font-weight:bold; margin-bottom:5px;'>" + params.name + "</div><div style='border-bottom:1px solid #555; padding-bottom:5px; margin-bottom:5px;'>Total Score: <span style='color:#3DCD58; font-weight:bold;'>" + params.value + "</span></div>" + (params.data.top5 || params.data.peopleHtml ? "<div style='font-size:12px; color:#ccc; margin-bottom:3px;'>Top Entities:</div>" + (params.data.top5 || params.data.peopleHtml) : ""); }} }};`;
            }}

            const isMap = chartConfigStr.includes('"map":"world"') || chartConfigStr.includes('"mapType":"world"');
            
            if (isMap) {{
                initScript = "var myChart = echarts.init(document.getElementById('chart-container')); " +
                             "myChart.showLoading(); " +
                             "fetch('https://s3-us-west-2.amazonaws.com/s.cdpn.io/95368/world.json').then(r => r.json()).then(mapJson => {{ " +
                             "    echarts.registerMap('world', mapJson); " +
                             "    myChart.hideLoading(); " +
                             "    var option = " + chartConfigStr + "; " +
                             "    " + tooltipPatch + " " +
                             "    myChart.setOption(option, true); " +
                             "    window.addEventListener('resize', () => myChart.resize()); " +
                             "}});";
            }} else {{
                initScript = "var myChart = echarts.init(document.getElementById('chart-container')); " +
                             "var option = " + chartConfigStr + "; " +
                             "    " + tooltipPatch + " " +
                             "myChart.setOption(option, true); " +
                             "window.addEventListener('resize', () => myChart.resize());";
            }}
        }} 
        // Chart.js
        else if (engine === 'chartjs') {{
            const safeConfig = {{
                type: 'bar',
                data: chartInstance.data,
                options: {{ responsive: true, maintainAspectRatio: false, scales: {{ x: {{ stacked: true }}, y: {{ stacked: true }} }} }}
            }};
            chartConfigStr = JSON.stringify(safeConfig);
            libraryScript = '<script src="https://cdn.jsdelivr.net/npm/chart.js"></' + 'script>';
            

            initScript = "var container = document.getElementById('chart-container'); " +
                         "var wrapper = document.createElement('div'); " +
                         "wrapper.style.width = '90vw'; wrapper.style.height = '85vh'; wrapper.style.position = 'relative'; wrapper.style.marginTop = '80px'; " +
                         "var canvas = document.createElement('canvas'); " +
                         "wrapper.appendChild(canvas); " +
                         "container.appendChild(wrapper); " +
                         "var ctx = canvas.getContext('2d'); " +
                         "new Chart(ctx, " + chartConfigStr + ");";
        }}

        const displayTitle = filename.replace(/_/g, ' ');

        const htmlContent = "<!DOCTYPE html>\\n" +
        "<html lang='en'>\\n" +
        "<head>\\n" +
        "    <meta charset='UTF-8'>\\n" +
        "    <meta name='viewport' content='width=device-width, initial-scale=1.0'>\\n" +
        "    <title>" + displayTitle + "</title>\\n" +
        "    " + libraryScript + "\\n" +
        "    <link href='https://fonts.googleapis.com/css2?family=Inter:wght@400;600&display=swap' rel='stylesheet'>\\n" +
        "    <style>\\n" +
        "        body {{ margin: 0; padding: 0; font-family: 'Inter', sans-serif; background: #FBFCFD; overflow: hidden; }}\\n" +
        "        #chart-container {{ width: 100vw; height: 100vh; display: flex; justify-content: center; align-items: center; padding: 30px; box-sizing: border-box; }}\\n" +
        "        .header {{ position: absolute; top: 15px; left: 15px; z-index: 10; background: rgba(255,255,255,0.95); padding: 12px 20px; border-radius: 8px; box-shadow: 0 4px 12px rgba(0,0,0,0.1); border: 1px solid #EDF2F7; }}\\n" +
        "        .header h3 {{ margin: 0; color: #2D3436; font-size: 16px; }}\\n" +
        "        .header p {{ margin: 4px 0 0 0; color: #636E72; font-size: 12px; }}\\n" +
        "    </style>\\n" +
        "</head>\\n" +
        "<body>\\n" +
        "    <div class='header'>\\n" +
        "        <h3>" + displayTitle + "</h3>\\n" +
        "        <p>Interactive View - Data Visualization snapshot captured from Data Manager.</p>\\n" +
        "    </div>\\n" +
        "    <div id='chart-container'></div>\\n" +
        "    <script>\\n" +
        "        " + initScript + "\\n" +
        "    </" + "script>\\n" +
        "</body>\\n" +
        "</html>";

        const blob = new Blob([htmlContent], {{ type: 'text/html' }});
        const url = URL.createObjectURL(blob);
        const a = document.createElement('a');
        a.href = url;
        a.download = filename + '.html';
        document.body.appendChild(a);
        a.click();
        document.body.removeChild(a);
        URL.revokeObjectURL(url);
    }};

    $(document).ready(function() {{
        initTable(currentTableName);
        
        $('#addFilter').click(() => {{ const v = $('#customSearch').val(), i = $('#columnSelect').val(), n = $('#columnSelect option:selected').text(); i === 'all' ? addFilter(v, 'all') : addFilter(v, parseInt(i), n); }});
        $('#excludeBtn').click(() => {{
  const v = $('#customSearch').val();
  const i = $('#columnSelect').val();
  const n = $('#columnSelect option:selected').text();

  if (i === 'all') addFilter(v, 'all', 'All', false, true);
  else addFilter(v, parseInt(i), n, false, true);
}});

        $('#customSearch').keypress(e => {{ if(e.which === 13) $('#addFilter').click(); }});
        $('#clearAll').click(() => {{ activeFilters = []; renderFilters(); applyFilters(); }});
        
        $('#topExportBtn').click(() => {{
             exportDataToExcel(fullColumns.map(c=>c.title), filteredIndices.map(i=>fullData[i]), currentTableName.replace(/ /g, "_") + "_Export");
        }});
        
        if(isStdMode) {{
            $('#powerBtn').click(calculatePower); 
            $('#radarBtn').click(initRadar); 
            $('#landscapeBtn').click(calculateLandscape); 
            $('#benchBtn').click(calculateBench); 
            $('#networkBtn').click(showNetworkMap);
            $('#geoBtn').click(calculateGeo);
            $('#votingBtn').click(calculateVoting);
        }}
    }});

    // Tutorial
    const TUT_STEPS = [
        {{
            title: "👋 Welcome to the Data Manager Dashboard",
            body: "This interactive dashboard lets you explore, filter, and analyse standardisation committee data.\\n\\nUse the ◀ Previous / Next ▶ buttons to navigate through the tutorial, or click ✕ at any time to close it."
        }},
        {{
            title: "📋 Dataset Overview (Left Sidebar)",
            body: "The left sidebar shows key statistics at a glance:\\n\\n• Total Records — total rows across all loaded files.\\n• Visible — rows currently shown after applying filters.\\n• Active Files — number of source files contributing to the view.\\n\\nYou can toggle individual files on or off by clicking their names in the Active Files list."
        }},
        {{
            title: "🔍 Search & Filter Bar",
            body: "Use the search bar at the top of the main panel to filter the table:\\n\\n• Select a column from the dropdown (or use 'Global Search' to search all columns).\\n• Type one or more keywords and click Apply.\\n• Multiple keywords in the SAME column act as OR — multiple columns act as AND.\\n• Double-click any cell in the table to instantly filter by that value.\\n\\n Click the ✕ on any active filter chip to remove it. Click Reset to clear all filters."
        }},
        {{
            title: "🔘 Action Buttons (Filter Bar)",
            body: "Three action buttons sit next to the search input:\\n\\n• Apply — adds the current keyword as a filter chip.\\n• 🚫 Exclude — removes all rows that match the keyword (bulk data hygiene).\\n• ↺ Reset — clears every active filter and restores the full dataset.\\n\\nFilters update the table and all analytics in real time."
        }},
        {{
            title: "📈 Analytics Toolbar",
            body: "Seven analytical modules are available in Standard Mode (hover each button for a summary):\\n\\n⚡ Company Rankings — top 20 entities by weighted Power Score.\\n🎯 Competitor Radar — spider chart for up to 5 companies.\\n⚖ Gaps & Leads — committees where SE leads or trails by >15 pts.\\n🕸 Network Map — force-directed graph of company ↔ committee links.\\n🌍 Geo Heatmap — world map coloured by national influence.\\n🛡 Expert Workload — SE expert role count and geographic coverage.\\n🗳 Voting — simulate IEC (1 country = 1 vote) or CENELEC (weighted two-step) votes."
        }},
        {{
            title: "📤 Exporting Data",
            body: "You can export data at any time:\\n\\n• Export Table button (top-right of the table) — downloads the currently visible, filtered rows.\\n• Each analytics modal also has its own Export button to save that specific chart or list.\\n\\nSupported formats depend on your configuration in the app: .xlsx (default) or .csv."
        }},
        {{
            title: "✅ You're Ready!",
            body: "That covers all the key features of the dashboard.\\n\\nQuick tips:\\n• Hover over analytics buttons to read a one-line description.\\n• Analytics always reflect the currently filtered data — filter first, then analyse.\\n• Click Tutorial in the top navbar anytime to re-open this guide.\\n\\nQuestions? Contact: martial.patra@se.com or pauline.mourlon@se.com\\n\\nHappy exploring! 🎉"
        }}
    ];
    let tutIdx = 0;
    function openTutorial() {{
        tutIdx = 0;
        renderTutStep();
        document.getElementById('tutorialModal').style.display = 'flex';
    }}
    function closeTutorial() {{
        document.getElementById('tutorialModal').style.display = 'none';
    }}
    function tutNav(dir) {{
        if (dir === 1 && tutIdx === TUT_STEPS.length - 1) {{ closeTutorial(); return; }}
        tutIdx = Math.max(0, Math.min(TUT_STEPS.length - 1, tutIdx + dir));
        renderTutStep();
    }}
    function renderTutStep() {{
        const s = TUT_STEPS[tutIdx];
        const n = TUT_STEPS.length;
        document.getElementById('tut-title').textContent = s.title;
        document.getElementById('tut-body').textContent = s.body;
        document.getElementById('tut-step-label').textContent = 'Step ' + (tutIdx+1) + ' of ' + n;
        document.getElementById('tut-counter').textContent = (tutIdx+1) + ' / ' + n;
        document.getElementById('tut-progress').style.width = ((tutIdx+1)/n*100) + '%';
        const prevBtn = document.getElementById('tut-prev');
        const nextBtn = document.getElementById('tut-next');
        prevBtn.disabled = tutIdx === 0;
        prevBtn.style.opacity = tutIdx === 0 ? '0.4' : '1';
        const isLast = tutIdx === n - 1;
        nextBtn.textContent = isLast ? 'Close ✓' : 'Next ▶';
        nextBtn.style.background = isLast ? '#c0392b' : '#3DCD58';
    }}

    </script>
    """

if __name__ == "__main__":
    app = ctk.CTk()
    gui = App(app)
    app.mainloop()