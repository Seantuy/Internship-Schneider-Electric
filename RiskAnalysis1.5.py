"""
ERM Analyst Cockpit
Developed by: Sean 
Purpose: Visualize analyst-weighted risk trends as an interactive bubble chart.
"""

import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
import os
import json
import webbrowser
from datetime import datetime
import threading

# ── Logo (Schneider Electric) ──────────────────────────────────────────────────
LOGO_URL = "https://raw.githubusercontent.com/Seantuy/Internship-Schneider-Electric/main/Schneider-Electric-Logo.jpg"

# ── Theme (light) ──────────────────────────────────────────────────────────────
BRAND_GREEN  = "#2EB84B"
BRAND_BG     = "#F5F6FA"
BRAND_PANEL  = "#FFFFFF"
BRAND_BORDER = "#DDE1EA"
BRAND_TEXT   = "#1A1D23"
BRAND_MUTED  = "#6B7280"
BRAND_CARD   = "#F0F2F7"

ctk.set_appearance_mode("light")
ctk.set_default_color_theme("green")


# ── Excel parsing ──────────────────────────────────────────────────────────────

def parse_excel(path: str):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = wb.sheetnames
    if len(sheets) < 2:
        raise ValueError(f"Expected at least 2 sheets, found {len(sheets)}: {sheets}")

    def read_sheet(ws, required_cols):
        rows_data, headers, header_map = [], None, {}
        for row in ws.iter_rows(values_only=True):
            if headers is None:
                if any(c is not None for c in row):
                    headers = [str(c).strip() if c is not None else "" for c in row]
                    for i, h in enumerate(headers):
                        header_map[h.lower()] = i
                    missing = [rc for rc in required_cols if rc.lower() not in header_map]
                    if missing:
                        raise ValueError(
                            f"Sheet '{ws.title}' missing columns: {missing}. Found: {headers}"
                        )
                continue
            if all(c is None for c in row):
                continue
            record = {}
            for rc in required_cols:
                idx = header_map[rc.lower()]
                record[rc] = row[idx] if idx < len(row) else None
            rows_data.append(record)
        return rows_data

    analyst_rows = read_sheet(wb[sheets[0]], ["Trend", "Poids", "Risk", "Analyst", "Sector"])
    se_rows      = read_sheet(wb[sheets[1]], ["Trend", "Poids"])
    wb.close()
    return analyst_rows, se_rows


def build_bubble_data(analyst_rows, se_rows):
    se_lookup = {}
    for r in se_rows:
        t = str(r["Trend"]).strip() if r["Trend"] else ""
        try:
            p = float(r["Poids"])
        except (TypeError, ValueError):
            p = None
        if t:
            se_lookup[t.lower()] = (t, p)

    groups = {}
    for r in analyst_rows:
        t = str(r["Trend"]).strip() if r["Trend"] else ""
        if not t:
            continue
        key = t.lower()
        if key not in groups:
            groups[key] = {"trend": t, "poids_values": [], "analysts": [],
                           "risks": [], "sectors": []}
        try:
            p = float(r["Poids"]) if r["Poids"] is not None else None
        except (TypeError, ValueError):
            p = None
        groups[key]["poids_values"].append(p)
        groups[key]["analysts"].append(str(r["Analyst"]) if r["Analyst"] else "—")
        groups[key]["risks"].append(str(r["Risk"])    if r["Risk"]    else "—")
        # Strip spaces and uppercase for cleaner matching later
        groups[key]["sectors"].append(str(r["Sector"]).strip().upper() if r["Sector"] else "—")

    bubbles, invalid = [], []
    for key, g in groups.items():
        se_poids = se_lookup[key][1] if key in se_lookup else None
        valid_p  = [v for v in g["poids_values"] if v is not None]
        # CHANGE: Calculate Gross Sum instead of Average
        gross_p  = round(sum(valid_p), 2) if valid_p else None
        occ      = len(g["poids_values"])
        detail   = [{"analyst": a, "poids": p, "risk": r, "sector": s}
                    for a, p, r, s in zip(g["analysts"], g["poids_values"],
                                          g["risks"], g["sectors"])]
        b = {"trend": g["trend"], "se_poids": se_poids,
             "gross_analyst_poids": gross_p, "occurrences": occ, "analysts": detail}
        (bubbles if se_poids is not None and gross_p is not None else invalid).append(b)

    return bubbles, invalid


# ── HTML ───────────────────────────────────────────────────────────────────────

HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>ERM Analyst Cockpit</title>
<script src="https://cdn.plot.ly/plotly-2.32.0.min.js"></script>
<style>
  @import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600&family=Space+Grotesk:wght@500;600;700&display=swap');

  *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

  :root {
    --green:     #2EB84B;
    --green-lt:  #E6F9EC;
    --green-dk:  #1d8a35;
    --bg:        #F5F6FA;
    --panel:     #FFFFFF;
    --card:      #F0F2F7;
    --border:    #DDE1EA;
    --text:      #1A1D23;
    --muted:     #6B7280;
    --accent:    #3B82F6;
    --shadow-sm: 0 1px 3px rgba(0,0,0,.07), 0 1px 2px rgba(0,0,0,.04);
    --shadow-md: 0 4px 16px rgba(0,0,0,.09);
    --shadow-lg: 0 12px 36px rgba(0,0,0,.13), 0 4px 10px rgba(0,0,0,.07);
  }

  html, body {
    height: 100%;
    background: var(--bg);
    color: var(--text);
    font-family: 'DM Sans', sans-serif;
    overflow-x: hidden;
    overflow-y: auto;
  }

  /* ── Topbar ── */
  .topbar {
    display: flex; align-items: center; justify-content: space-between;
    padding: 11px 22px;
    background: var(--panel);
    border-bottom: 1px solid var(--border);
    box-shadow: var(--shadow-sm);
  }
  .topbar-left { display: flex; align-items: center; gap: 12px; }
  .logo {
    height: 38px; width: auto; max-width: 120px;
    object-fit: contain;
    flex-shrink: 0;
  }
  .topbar-title {
    font-family: 'Space Grotesk', sans-serif; font-size: 17px;
    font-weight: 700; color: var(--text); letter-spacing: -.3px;
  }
  .topbar-sub { font-size: 11px; color: var(--muted); margin-top: 2px; }
  .topbar-meta { font-size: 11px; color: var(--muted); text-align: right; line-height: 1.7; }
  .topbar-meta strong { color: var(--green); font-weight: 600; }

  /* ── Year tabs ── */
  .year-tabs {
    display: flex; align-items: center; gap: 6px;
    padding: 9px 22px;
    background: var(--panel);
    border-bottom: 1px solid var(--border);
    flex-wrap: wrap;
  }
  .year-tabs-label {
    font-size: 11px; font-weight: 600; color: var(--muted);
    text-transform: uppercase; letter-spacing: .6px;
    margin-right: 4px; white-space: nowrap;
  }
  .year-tab {
    padding: 5px 15px; border-radius: 20px;
    border: 1.5px solid var(--border);
    background: var(--card); color: var(--muted);
    font-family: 'DM Sans', sans-serif; font-size: 12px; font-weight: 500;
    cursor: pointer; transition: all .16s ease; white-space: nowrap; outline: none;
  }
  .year-tab:hover { border-color: var(--green); color: var(--green); background: var(--green-lt); }
  .year-tab.active {
    background: var(--green); color: #fff; border-color: var(--green);
    box-shadow: 0 2px 8px rgba(46,184,75,.28);
  }

  /* ── Stats bar ── */
  .statsbar {
    display: flex; background: var(--panel);
    border-bottom: 1px solid var(--border);
  }
  .stat {
    padding: 8px 20px; border-right: 1px solid var(--border);
    display: flex; align-items: center; gap: 10px;
  }
  .stat:last-child { border-right: none; }
  .stat-ico { font-size: 19px; flex-shrink: 0; }
  .stat-val {
    font-family: 'Space Grotesk', sans-serif; font-size: 20px;
    font-weight: 700; color: var(--text); line-height: 1;
  }
  .stat-lbl { font-size: 10px; color: var(--muted); text-transform: uppercase;
    letter-spacing: .6px; margin-top: 3px; }

  /* ── Legend ── */
  .legend-bar {
    padding: 6px 22px; background: var(--bg);
    border-bottom: 1px solid var(--border);
    display: flex; align-items: center; gap: 18px; flex-wrap: wrap;
  }
  .legend-bar > span { font-size: 11px; color: var(--muted); }
  .legend-item { display: flex; align-items: center; gap: 6px; font-size: 11px; color: var(--muted); }
  .legend-bubble { border-radius: 50%; background: var(--green); opacity: .7; display: inline-block; }
  .axis-hint { font-size: 10px; color: var(--muted); margin-left: auto; }

  /* ── Chart ── */
  #chart { width: 100%; overflow: visible; }

  /* ── Detail panel ── */
  #detail-panel {
    display: none;
    position: fixed; top: 0; right: 0;
    width: 315px;
    background: var(--panel);
    border-left: 1px solid var(--border);
    border-bottom: 1px solid var(--border);
    border-bottom-left-radius: 14px;
    box-shadow: var(--shadow-lg);
    z-index: 9999;
    flex-direction: column;
    overflow: hidden;
    max-height: 70vh;
  }
  #detail-panel.open { display: flex; }

  .dp-header {
    padding: 13px 14px 10px;
    border-bottom: 1px solid var(--border);
    display: flex; align-items: flex-start; justify-content: space-between;
    flex-shrink: 0; background: var(--panel);
  }
  .dp-title {
    font-family: 'Space Grotesk', sans-serif; font-size: 14px; font-weight: 700;
    color: var(--text); word-break: break-word; padding-right: 8px; line-height: 1.3;
  }
  .dp-occ { font-size: 11px; color: var(--muted); margin-top: 3px; }
  .dp-close {
    background: var(--card); border: 1px solid var(--border); border-radius: 6px;
    width: 26px; height: 26px; cursor: pointer; display: flex;
    align-items: center; justify-content: center;
    color: var(--muted); font-size: 13px; flex-shrink: 0;
    transition: all .14s; font-family: sans-serif; line-height: 1;
  }
  .dp-close:hover { background: #e0e4ec; color: var(--text); }

  .dp-kpis {
    display: grid; grid-template-columns: 1fr 1fr;
    gap: 8px; padding: 10px 14px; background: var(--bg); flex-shrink: 0;
  }
  .dp-kpi {
    background: var(--panel); border: 1px solid var(--border);
    border-radius: 9px; padding: 9px 10px; text-align: center;
  }
  .dp-kpi-val {
    font-family: 'Space Grotesk', sans-serif; font-size: 20px; font-weight: 700;
  }
  .dp-kpi-lbl { font-size: 9px; color: var(--muted); text-transform: uppercase;
    letter-spacing: .5px; margin-top: 3px; }

  .dp-section-label {
    padding: 8px 14px 4px; font-size: 10px; font-weight: 600; color: var(--muted);
    text-transform: uppercase; letter-spacing: .6px; flex-shrink: 0;
  }

  .dp-table-wrap { overflow-y: auto; flex: 1; }
  .dp-table { width: 100%; border-collapse: collapse; }
  .dp-table thead th {
    position: sticky; top: 0; background: var(--bg);
    padding: 5px 10px; text-align: left;
    font-size: 10px; font-weight: 600; color: var(--muted);
    text-transform: uppercase; letter-spacing: .5px;
    border-bottom: 1px solid var(--border);
  }
  .dp-table td { padding: 6px 10px; font-size: 12px; color: var(--text);
    border-bottom: 1px solid var(--border); }
  .dp-table tbody tr:last-child td { border-bottom: none; }
  .dp-table tbody tr:hover td { background: var(--card); }
  .dp-poids { font-weight: 700; color: var(--green); text-align: center; }

  /* ── Trend changes notice bar ── */
  #trend-notice {
    display: none;
    padding: 7px 22px;
    background: var(--panel);
    border-bottom: 1px solid var(--border);
    align-items: flex-start;
    gap: 10px;
    flex-wrap: wrap;
  }
  #trend-notice.visible { display: flex; }
  .notice-group { display: flex; align-items: center; gap: 6px; flex-wrap: wrap; }
  .notice-label {
    font-size: 10px; font-weight: 700; text-transform: uppercase;
    letter-spacing: .6px; white-space: nowrap; padding: 3px 8px;
    border-radius: 20px; flex-shrink: 0;
  }
  .notice-label.new  { background: #fee2e2; color: #b91c1c; } /* Now Red */
  .notice-label.gone { background: #dcfce7; color: #15803d; } /* Now Green */
  .notice-divider { width: 1px; height: 20px; background: var(--border); flex-shrink: 0; align-self: center; }
  .notice-pill {
    font-size: 11px; font-weight: 500; padding: 3px 9px; border-radius: 20px;
    white-space: nowrap; cursor: default; border: 1px solid transparent;
  }
  .notice-pill.new  { background: #fff1f2; color: #9f1239; border-color: #fca5a5; } /* Now Red */
  .notice-pill.gone { background: #f0fdf4; color: #166534; border-color: #86efac; } /* Now Green */
  .notice-toggle {
    font-size: 10px; color: var(--muted); cursor: pointer; text-decoration: underline;
    white-space: nowrap; align-self: center; background: none; border: none;
    padding: 0; font-family: inherit;
  }
  .notice-toggle:hover { color: var(--text); }

  /* ── Clickable stat ── */
  .stat-clickable {
    cursor: pointer; transition: background .15s;
  }
  .stat-clickable:hover { background: var(--green-lt); }

  /* ── List modal ── */
  #list-modal-overlay {
    display: none; position: fixed; inset: 0;
    background: rgba(26,29,35,.35); z-index: 99999;
    align-items: center; justify-content: center;
  }
  #list-modal-overlay.open { display: flex; }
  #list-modal {
    background: var(--panel); border-radius: 14px;
    box-shadow: var(--shadow-lg); width: 360px; max-height: 70vh;
    display: flex; flex-direction: column; overflow: hidden;
    border: 1px solid var(--border);
  }
  .lm-header {
    padding: 14px 16px 11px;
    border-bottom: 1px solid var(--border);
    display: flex; align-items: center; justify-content: space-between;
    flex-shrink: 0;
  }
  .lm-title {
    font-family: 'Space Grotesk', sans-serif; font-size: 14px;
    font-weight: 700; color: var(--text);
  }
  .lm-count {
    font-size: 11px; color: var(--muted); margin-top: 2px;
  }
  .lm-close {
    background: var(--card); border: 1px solid var(--border); border-radius: 6px;
    width: 26px; height: 26px; cursor: pointer; display: flex;
    align-items: center; justify-content: center;
    color: var(--muted); font-size: 13px; flex-shrink: 0;
    transition: all .14s; font-family: sans-serif; line-height: 1; outline: none;
  }
  .lm-close:hover { background: #e0e4ec; color: var(--text); }
  .lm-search-wrap {
    padding: 10px 14px; border-bottom: 1px solid var(--border); flex-shrink: 0;
  }
  .lm-search {
    width: 100%; padding: 7px 10px; border-radius: 8px;
    border: 1.5px solid var(--border); font-family: 'DM Sans', sans-serif;
    font-size: 12px; color: var(--text); background: var(--bg); outline: none;
    box-sizing: border-box;
  }
  .lm-search:focus { border-color: var(--green); }
  .lm-list { overflow-y: auto; flex: 1; padding: 8px 0; }
  .lm-item {
    padding: 8px 16px; font-size: 13px; color: var(--text);
    border-bottom: 1px solid var(--border); display: flex;
    align-items: center; gap: 8px;
  }
  .lm-item:last-child { border-bottom: none; }
  .lm-item-bullet {
    width: 7px; height: 7px; border-radius: 50%;
    background: var(--green); flex-shrink: 0;
  }
  .lm-item.hidden { display: none; }
</style>
</head>
<body>

<div class="topbar" id="topbar">
  <div class="topbar-left">
    <img class="logo" src="__LOGO_URL__" alt="Logo">
    <div>
      <div class="topbar-title">ERM Analyst Cockpit</div>
      <div class="topbar-sub">Trend bubble map — Analyst Gross Risk (X) × SE Possible Impact (Y)</div>
    </div>
  </div>
  <div class="topbar-meta" id="top-meta"></div>
</div>

<div class="year-tabs" id="year-tabs-bar"></div>

<div class="statsbar" id="statsbar">
  <div class="stat stat-clickable" onclick="openListModal('trends')" title="Click to see all trends"><span class="stat-ico">🔵</span><div>
    <div class="stat-val" id="s-trends">—</div><div class="stat-lbl">Total Number of Trends</div>
  </div></div>
  <div class="stat stat-clickable" onclick="openListModal('analysts')" title="Click to see all analysts"><span class="stat-ico">👤</span><div>
    <div class="stat-val" id="s-analysts">—</div><div class="stat-lbl">E Analysts</div>
  </div></div>
  <div class="stat"><span class="stat-ico">📋</span><div>
    <div class="stat-val" id="s-entries">—</div><div class="stat-lbl">Total Entries</div>
  </div></div>
  <div class="stat"><span class="stat-ico">📊</span><div>
    <div class="stat-val" id="s-maxocc">—</div><div class="stat-lbl">Max Occurrences</div>
  </div></div>
</div>

<div class="legend-bar" id="legendbar">
  <span>Bubble size = number of analyst occurrences</span>
  <div class="axis-hint">X: Analyst Gross Risk &nbsp;|&nbsp; Y: SE Possible Impact &nbsp;|&nbsp; Click bubble for details</div>
</div>

<div id="trend-notice"></div>

<div id="chart"></div>

<div id="detail-panel">
  <div class="dp-header">
    <div>
      <div class="dp-title" id="dp-title">—</div>
      <div class="dp-occ"   id="dp-occ">—</div>
    </div>
    <button class="dp-close" id="dp-close-btn">✕</button>
  </div>
  <div class="dp-kpis">
    <div class="dp-kpi">
      <div class="dp-kpi-val" id="dp-avg" style="color:var(--accent)">—</div>
      <div class="dp-kpi-lbl">Analyst Gross Risk</div>
    </div>
    <div class="dp-kpi">
      <div class="dp-kpi-val" id="dp-se" style="color:var(--green)">—</div>
      <div class="dp-kpi-lbl">SE Possible Impact</div>
    </div>
  </div>
  <div class="dp-section-label">Analyst Breakdown</div>
  <div class="dp-table-wrap">
    <table class="dp-table">
      <thead><tr>
        <th>Analyst</th>
        <th style="text-align:center">Poids</th>
        <th>Risk</th>
        <th>Sector</th>
      </tr></thead>
      <tbody id="dp-tbody"></tbody>
    </table>
  </div>
</div>

<div id="list-modal-overlay">
  <div id="list-modal">
    <div class="lm-header">
      <div>
        <div class="lm-title" id="lm-title">—</div>
        <div class="lm-count" id="lm-count"></div>
      </div>
      <button class="lm-close" onclick="closeListModal()">✕</button>
    </div>
    <div class="lm-search-wrap">
      <input class="lm-search" id="lm-search" type="text" placeholder="Search…" oninput="filterListModal(this.value)">
    </div>
    <div class="lm-list" id="lm-list"></div>
  </div>
</div>

<script>
// ── Injected data ─────────────────────────────────────────────────────────────
const ALL_DATASETS = __ALL_DATASETS_JSON__;

// ── Sector Mapping ────────────────────────────────────────────────────────────
// Maps the exact abbreviations from the image to their names and hex colors
const SECTOR_MAP = {
  'S':  { name: 'Societal', color: '#FF0000' },                 // Red
  'T':  { name: 'Technology disruption', color: '#B000B0' },    // Purple
  'EC': { name: 'Economic', color: '#0070C0' },                 // Blue
  'EN': { name: 'Environment', color: '#00B050' },              // Green
  'P':  { name: 'Geopolitics & Regulation', color: '#FF9900' }  // Orange
};

// ── Panel helpers ─────────────────────────────────────────────────────────────
const panel = document.getElementById('detail-panel');

function openPanel(b, grossP, seP, prevB) {
  document.getElementById('dp-title').textContent = b.trend;
  document.getElementById('dp-occ').textContent =
    b.occurrences + ' occurrence' + (b.occurrences !== 1 ? 's' : '');

  function deltaHtml(current, previous, isRisk) {
    if (previous == null || previous == undefined) return current;
    const diff = parseFloat((current - previous).toFixed(2));
    if (diff === 0) {
      return `${current} <span style="font-size:13px;color:#6B7280;font-weight:500">→ 0</span>`;
    }
    // For risk (X axis): up = worse (red), down = better (green)
    // For SE impact (Y axis): up = more impact (red), down = less (green)
    const up    = diff > 0;
    const arrow = up ? '↑' : '↓';
    const color = up ? '#ef4444' : '#22c55e';
    return `${current} <span style="font-size:13px;color:${color};font-weight:600">${arrow} ${Math.abs(diff)}</span>`;
  }

  const prevGross = prevB ? prevB.gross_analyst_poids : null;
  const prevSe    = prevB ? prevB.se_poids            : null;

  document.getElementById('dp-avg').innerHTML = deltaHtml(grossP, prevGross, true);
  document.getElementById('dp-se').innerHTML  = deltaHtml(seP,    prevSe,    false);

  document.getElementById('dp-tbody').innerHTML = b.analysts.map(a => `
    <tr>
      <td>${a.analyst}</td>
      <td class="dp-poids">${a.poids ?? '—'}</td>
      <td>${a.risk ?? '—'}</td>
      <td>${a.sector ?? '—'}</td>
    </tr>`).join('');

  const headerBottom = document.getElementById('legendbar').getBoundingClientRect().bottom;
  panel.style.top = headerBottom + 'px';
  panel.style.maxHeight = (window.innerHeight - headerBottom - 20) + 'px';
  panel.classList.add('open');
}

function closePanel() {
  panel.classList.remove('open');
}

// ── List modal ────────────────────────────────────────────────────────────────
const listOverlay = document.getElementById('list-modal-overlay');

function openListModal(type) {
  const ds      = ALL_DATASETS[currentKey];
  const bubbles = ds.bubbles;
  let items = [];
  let title = '';

  if (type === 'trends') {
    title = 'All Trends';
    items = bubbles.map(b => b.trend).sort((a, b) => a.localeCompare(b));
  } else {
    title = 'Unique Analysts';
    const set = new Set();
    bubbles.forEach(b => b.analysts.forEach(a => { if (a.analyst && a.analyst !== '—') set.add(a.analyst); }));
    items = [...set].sort((a, b) => a.localeCompare(b));
  }

  document.getElementById('lm-title').textContent  = title;
  document.getElementById('lm-count').textContent  = items.length + ' ' + (items.length === 1 ? 'item' : 'items');
  document.getElementById('lm-search').value       = '';

  const list = document.getElementById('lm-list');
  list.innerHTML = items.map(item => `
    <div class="lm-item" data-name="${item.toLowerCase()}">
      <span class="lm-item-bullet"></span>${item}
    </div>`).join('');

  listOverlay.classList.add('open');
}

function closeListModal() {
  listOverlay.classList.remove('open');
}

function filterListModal(q) {
  const lower = q.toLowerCase().trim();
  document.querySelectorAll('#lm-list .lm-item').forEach(el => {
    el.classList.toggle('hidden', lower && !el.dataset.name.includes(lower));
  });
}

listOverlay.addEventListener('click', e => { if (e.target === listOverlay) closeListModal(); });
document.addEventListener('keydown', e => { if (e.key === 'Escape') closeListModal(); });

document.getElementById('dp-close-btn').addEventListener('click', closePanel);

// ── Chart rendering ───────────────────────────────────────────────────────────
function showDataset(key) {
  closePanel();

  document.querySelectorAll('.year-tab').forEach(b => {
    b.classList.toggle('active', b.dataset.key === key);
  });

  const ds      = ALL_DATASETS[key];
  const bubbles = ds.bubbles;

  // Stats
  const analysts = new Set();
  let totalEntries = 0, maxOcc = 0;
  bubbles.forEach(b => {
    b.analysts.forEach(a => analysts.add(a.analyst));
    totalEntries += b.occurrences;
    if (b.occurrences > maxOcc) maxOcc = b.occurrences;
  });
  document.getElementById('s-trends').textContent   = bubbles.length;
  document.getElementById('s-analysts').textContent = analysts.size;
  document.getElementById('s-entries').textContent  = totalEntries;
  document.getElementById('s-maxocc').textContent   = maxOcc;
  document.getElementById('top-meta').innerHTML =
    `Generated <strong>${ds.date}</strong><br>Source: <strong>${ds.filename}</strong>`;

  // ── Previous-year lookup for arrow badges ─────────────────────────────────
  const allKeys   = Object.keys(ALL_DATASETS);
  const currIdx   = allKeys.indexOf(key);
  const prevBubbles = currIdx > 0
    ? ALL_DATASETS[allKeys[currIdx - 1]].bubbles
    : [];
  const prevMap = {};
  prevBubbles.forEach(b => { prevMap[b.trend.toLowerCase()] = b; });

  // ── New / Gone trend notice bar ───────────────────────────────────────────
  const noticeEl = document.getElementById('trend-notice');
  const currKeys = new Set(bubbles.map(b => b.trend.toLowerCase()));
  const prevKeys = new Set(prevBubbles.map(b => b.trend.toLowerCase()));

  const newTrends  = bubbles.filter(b => !prevKeys.has(b.trend.toLowerCase())).map(b => b.trend);
  const goneTrends = prevBubbles.filter(b => !currKeys.has(b.trend.toLowerCase())).map(b => b.trend);

  const PILL_LIMIT = 6;

  function buildGroup(trends, type, icon, labelText) {
    if (!trends.length) return '';
    const visible  = trends.slice(0, PILL_LIMIT);
    const overflow = trends.slice(PILL_LIMIT);
    const pills    = visible.map(t =>
      `<span class="notice-pill ${type}" title="${t}">${icon} ${t}</span>`
    ).join('');
    const more = overflow.length
      ? `<button class="notice-toggle" onclick="this.parentElement.querySelectorAll('.notice-pill.hidden-pill').forEach(p=>p.style.display='inline');this.remove()">+${overflow.length} more</button>
         ${overflow.map(t => `<span class="notice-pill ${type} hidden-pill" style="display:none" title="${t}">${icon} ${t}</span>`).join('')}`
      : '';
    return `<div class="notice-group">
      <span class="notice-label ${type}">${labelText}</span>
      ${pills}${more}
    </div>`;
  }

  if (newTrends.length || goneTrends.length) {
    const newHtml  = buildGroup(newTrends,  'new',  '⚠️', '⚠️ New Risk');
    const goneHtml = buildGroup(goneTrends, 'gone', '✓', '✓ Risk Neutralized');
    const divider  = (newTrends.length && goneTrends.length) ? '<div class="notice-divider"></div>' : '';
    noticeEl.innerHTML = newHtml + divider + goneHtml;
    noticeEl.classList.add('visible');
  } else {
    noticeEl.innerHTML = '';
    noticeEl.classList.remove('visible');
  }

  function dirArrow(b) {
    const prev = prevMap[b.trend.toLowerCase()];
    if (!prev) return null;
    const dX = parseFloat((b.gross_analyst_poids - prev.gross_analyst_poids).toFixed(4));
    const dY = parseFloat((b.se_poids - prev.se_poids).toFixed(4));
    const EPS = 0.001;
    const xUp = dX > EPS, xDn = dX < -EPS;
    const yUp = dY > EPS, yDn = dY < -EPS;
    if (!xUp && !xDn && !yUp && !yDn) return null; // no change
    if  (xUp && yUp)  return { sym: '⬈', color: '#000000' };
    if  (xUp && yDn)  return { sym: '⬊', color: '#000000' };
    if  (xDn && yUp)  return { sym: '⬉', color: '#000000' };
    if  (xDn && yDn)  return { sym: '⬋', color: '#000000' };
    if  (xUp)         return { sym: '➡', color: '#000000' };
    if  (xDn)         return { sym: '⬅', color: '#000000' };
    if  (yUp)         return { sym: '⬆', color: '#000000' };
                      return { sym: '⬇', color: '#000000' };
  }

  // Build arrow overlay data (only bubbles with a change)
  const arrowX = [], arrowY = [], arrowText = [], arrowColors = [];
  bubbles.forEach(b => {
    const a = dirArrow(b);
    if (!a) return;
    arrowX.push(b.gross_analyst_poids);
    arrowY.push(b.se_poids);
    arrowText.push(a.sym);
    arrowColors.push(a.color);
  });

  const arrowTrace = {
    name: 'YoY change',
    type: 'scatter',
    mode: 'text',
    x: arrowX,
    y: arrowY,
    text: arrowText,
    textposition: 'top right',
    textfont: { size: 22, color: '#000000', family: 'DM Sans, sans-serif', weight: 'bold' },
    hoverinfo: 'skip',
    showlegend: false,
  };

  // Group by dominant sector for Legend mapping
  const sectorGroups = {};
  bubbles.forEach(b => {
    const secs = [...new Set(b.analysts.map(a => a.sector))];
    const top  = secs.length ? secs[0] : 'Unknown';
    if (!sectorGroups[top]) sectorGroups[top] = [];
    sectorGroups[top].push(b);
  });

  const traces = Object.keys(sectorGroups).map((sectorKey) => {
    const items = sectorGroups[sectorKey];
    // Grab color and proper name from our dictionary, fallback to grey if unmapped
    const secData = SECTOR_MAP[sectorKey] || { name: sectorKey, color: '#888888' };
    
    return {
      name: secData.name,
      type: 'scatter',
      mode: 'markers+text',
      x: items.map(b => b.gross_analyst_poids),  // X = Sum of weights
      y: items.map(b => b.se_poids),             // Y = SE Poids
      text: items.map(b => b.trend),
      
      // CHANGE: Centers the text inside the bubble and makes it black
      textposition: 'middle center',
      textfont: { size: 10, color: '#000000', family: 'DM Sans, sans-serif', weight: 'bold' },
      
      marker: {
        // CHANGE: Scale the bubbles proportionally so text fits inside
        size: items.map(b => Math.max(30, Math.sqrt(b.occurrences) * 22)),
        sizemode: 'diameter',
        color: secData.color, // CHANGE: Uses hardcoded exact color
        opacity: 0.95,
        line: { color: 'rgba(255,255,255,0.4)', width: 1 }
      },
      customdata: items,
      hovertemplate:
        '<b>%{text}</b><br>' +
        '<span style="color:#6B7280">Analyst Gross Risk:</span> <b>%{x}</b><br>' +
        '<span style="color:#6B7280">SE Possible Impact:</span> <b>%{y}</b><br>' +
        '<span style="color:#6B7280">Occurrences:</span> <b>%{customdata.occurrences}</b><br>' +
        '<extra></extra>',
    };
  });

  const usedPx =
    document.getElementById('topbar').offsetHeight +
    document.getElementById('year-tabs-bar').offsetHeight +
    document.getElementById('statsbar').offsetHeight +
    document.getElementById('legendbar').offsetHeight +
    document.getElementById('trend-notice').offsetHeight;
  const chartHeight = Math.max(window.innerHeight - usedPx - 30, 520);
  document.getElementById('chart').style.height = chartHeight + 'px';

  const layout = {
    paper_bgcolor: '#F5F6FA',
    plot_bgcolor:  '#FFFFFF',
    font: { family: 'DM Sans', color: '#6B7280', size: 12 },
    xaxis: {
      title: { text: 'Analyst Gross Risk', font: { color: '#1A1D23', size: 13 }, standoff: 12 },
      gridcolor: '#EEF0F5', zerolinecolor: '#DDE1EA',
      tickfont: { color: '#6B7280' },
      automargin: true,
    },
    yaxis: {
      title: { text: 'SE possible impact', font: { color: '#1A1D23', size: 13 }, standoff: 12 },
      gridcolor: '#EEF0F5', zerolinecolor: '#DDE1EA',
      tickfont: { color: '#6B7280' },
      automargin: true,
    },
    legend: {
      bgcolor: '#FFFFFF', bordercolor: '#DDE1EA', borderwidth: 1,
      font: { color: '#1A1D23', size: 11 },
      title: { text: 'Sector', font: { color: '#6B7280', size: 11 } },
      orientation: 'v', x: 1.01, y: 1,
    },
    margin: { t: 24, b: 110, l: 68, r: 175 },
    hoverlabel: {
      bgcolor: '#FFFFFF', bordercolor: '#2EB84B',
      font: { family: 'DM Sans', color: '#1A1D23', size: 12 }
    },
    hoverdistance: 20,
  };

  const config = {
    responsive: true, displaylogo: false,
    modeBarButtonsToRemove: ['select2d', 'lasso2d', 'autoScale2d'],
    toImageButtonOptions: { format: 'png', filename: 'risk_analysis_graph', scale: 2 }
  };

  Plotly.react('chart', [...traces, arrowTrace], layout, config);

  const chartDiv = document.getElementById('chart');
  chartDiv.removeAllListeners && chartDiv.removeAllListeners('plotly_click');
  chartDiv.on('plotly_click', function(data) {
    if (!data.points.length) return;
    const pt = data.points[0];

    // Find the same trend in the immediately preceding dataset (if any)
    const allKeys = Object.keys(ALL_DATASETS);
    const currIdx = allKeys.indexOf(currentKey);
    let prevB = null;
    if (currIdx > 0) {
      const prevKey     = allKeys[currIdx - 1];
      const prevBubbles = ALL_DATASETS[prevKey].bubbles;
      prevB = prevBubbles.find(
        bb => bb.trend.toLowerCase() === pt.customdata.trend.toLowerCase()
      ) || null;
    }

    openPanel(pt.customdata, pt.x, pt.y, prevB);
  });
}

(function init() {
  const keys   = Object.keys(ALL_DATASETS);
  const tabBar = document.getElementById('year-tabs-bar');

  if (keys.length <= 1) {
    tabBar.style.display = 'none';
  } else {
    const label = document.createElement('span');
    label.className   = 'year-tabs-label';
    label.textContent = 'Dataset:';
    tabBar.appendChild(label);
    keys.forEach(key => {
      const btn         = document.createElement('button');
      btn.className     = 'year-tab';
      btn.dataset.key   = key;
      btn.textContent   = ALL_DATASETS[key].label;
      btn.onclick       = () => showDataset(key);
      tabBar.appendChild(btn);
    });
  }

  showDataset(keys[0]);
  window.addEventListener('resize', () => showDataset(currentKey));
})();

let currentKey = Object.keys(ALL_DATASETS)[0];
const _origShow = showDataset;
window.showDataset = function(key) { currentKey = key; _origShow(key); };
document.querySelectorAll('.year-tab').length &&
  document.querySelectorAll('.year-tab')[0] &&
  (currentKey = document.querySelectorAll('.year-tab')[0]?.dataset.key || currentKey);
</script>
</body>
</html>
"""


def generate_html(datasets: list, source_filenames: list) -> str:
    all_ds = {}
    for bubbles, path in zip(datasets, source_filenames):
        key = os.path.splitext(os.path.basename(path))[0]
        all_ds[key] = {
            "label":    key,
            "filename": os.path.basename(path),
            "date":     datetime.now().strftime("%d %b %Y %H:%M"),
            "bubbles":  bubbles,
        }
    html = HTML_TEMPLATE.replace("__ALL_DATASETS_JSON__", json.dumps(all_ds, ensure_ascii=False))
    html = html.replace("__LOGO_URL__", LOGO_URL)
    return html


# ── GUI ────────────────────────────────────────────────────────────────────────

class FileRow(ctk.CTkFrame):
    def __init__(self, parent, path: str, on_remove, **kwargs):
        super().__init__(parent, fg_color=BRAND_CARD, corner_radius=8, **kwargs)
        self.path = path
        ctk.CTkLabel(
            self, text=os.path.basename(path),
            text_color=BRAND_TEXT, font=ctk.CTkFont("Segoe UI", 11), anchor="w"
        ).pack(side="left", fill="x", expand=True, padx=(10, 6), pady=7)
        ctk.CTkButton(
            self, text="✕", width=28, height=28, corner_radius=7,
            fg_color=BRAND_BORDER, hover_color="#CBD0DA",
            text_color=BRAND_MUTED, font=ctk.CTkFont("Segoe UI", 12),
            command=on_remove
        ).pack(side="right", padx=(0, 6), pady=4)


class App:
    def __init__(self, root):
        self.root = root
        self.root.title("ERM Analyst Cockpit")
        self.root.geometry("680x580")
        self.root.resizable(False, False)
        self.root.configure(fg_color=BRAND_BG)
        self._files: list = []
        self._file_rows: list = []
        self._output_path = None
        self._last_output = None
        self._build_ui()

    def _card(self, parent, title):
        f = ctk.CTkFrame(parent, fg_color=BRAND_PANEL, corner_radius=12,
                         border_width=1, border_color=BRAND_BORDER)
        ctk.CTkLabel(f, text=title, text_color=BRAND_TEXT,
                     font=ctk.CTkFont("Segoe UI", 12, "bold"), anchor="w"
                     ).pack(fill="x", padx=14, pady=(10, 6))
        return f

    def _build_ui(self):
        # Header
        header = ctk.CTkFrame(self.root, fg_color=BRAND_PANEL, corner_radius=0, height=64)
        header.pack(fill="x")
        header.pack_propagate(False)

        # ── Logo (Simplified) ─────────────────────────────────────────────────
        ctk.CTkLabel(header, text="SE", fg_color=BRAND_GREEN, text_color="#fff",
                     font=ctk.CTkFont("Segoe UI", 22, "bold"),
                     width=35, height=35, corner_radius=5).place(x=18, y=11)
        txt_x = 72

        ctk.CTkLabel(header, text="ERM Analyst Cockpit", text_color=BRAND_TEXT,
                     font=ctk.CTkFont("Segoe UI", 16, "bold")).place(x=txt_x, y=10)
        ctk.CTkLabel(header,
                     text="Multi-year bubble chart  ·  X: Analyst Gross Risk  ·  Y: SE Possible Impact",
                     text_color=BRAND_MUTED,
                     font=ctk.CTkFont("Segoe UI", 11)).place(x=txt_x, y=36)

        content = ctk.CTkFrame(self.root, fg_color="transparent")
        content.pack(fill="both", expand=True, padx=22, pady=14)

        # ── Step 1 ────────────────────────────────────────────────────────────
        s1 = self._card(content, "① Select Excel File(s)")
        s1.pack(fill="x", pady=(0, 10))

        br = ctk.CTkFrame(s1, fg_color="transparent")
        br.pack(fill="x", padx=14, pady=(0, 8))
        ctk.CTkButton(br, text="+ Add File…",
                      fg_color=BRAND_GREEN, hover_color="#259e3e",
                      text_color="#fff", font=ctk.CTkFont("Segoe UI", 12, "bold"),
                      width=110, height=32, corner_radius=8,
                      command=self._add_file).pack(side="left")
        ctk.CTkLabel(br,
                     text="Each file becomes one tab in the HTML graph.",
                     text_color=BRAND_MUTED, font=ctk.CTkFont("Segoe UI", 10),
                     wraplength=430, anchor="w").pack(side="left", padx=(12, 0))

        # File list — wrapper enforces height, inner scrollable frame fills it
        _list_wrap = ctk.CTkFrame(
            s1, fg_color=BRAND_PANEL, corner_radius=8,
            border_width=1, border_color=BRAND_BORDER,
            height=90
        )
        _list_wrap.pack(fill="x", padx=14, pady=(0, 8))
        _list_wrap.pack_propagate(False)   # <-- enforces the height

        self.file_list = ctk.CTkScrollableFrame(
            _list_wrap, fg_color="transparent", corner_radius=0,
            border_width=0,
            scrollbar_button_color=BRAND_BORDER,
            scrollbar_button_hover_color="#CBD0DA"
        )
        self.file_list.pack(fill="both", expand=True)

        self.empty_lbl = ctk.CTkLabel(
            self.file_list,
            text="No files added yet — click '+ Add File…'",
            text_color=BRAND_MUTED, font=ctk.CTkFont("Segoe UI", 11)
        )
        self.empty_lbl.pack(pady=10)

        # Hint
        hint = ctk.CTkFrame(s1, fg_color=BRAND_CARD, corner_radius=8)
        hint.pack(fill="x", padx=14, pady=(0, 12))
        for tag, lbl, cols in [
            ("Sheet 1", "Analyst data", "Trend · Poids · Risk · Analyst · Sector"),
            ("Sheet 2", "SE reference",  "Trend · Poids  (no duplicates)"),
        ]:
            r = ctk.CTkFrame(hint, fg_color="transparent")
            r.pack(fill="x", padx=10, pady=3)
            ctk.CTkLabel(r, text=tag, text_color=BRAND_GREEN,
                         font=ctk.CTkFont("Segoe UI", 10, "bold"), width=50, anchor="w").pack(side="left")
            ctk.CTkLabel(r, text=lbl, text_color=BRAND_TEXT,
                         font=ctk.CTkFont("Segoe UI", 10), width=88, anchor="w").pack(side="left")
            ctk.CTkLabel(r, text=cols, text_color=BRAND_MUTED,
                         font=ctk.CTkFont("Segoe UI", 10), anchor="w").pack(side="left")

        # ── Step 2 ────────────────────────────────────────────────────────────
        s2 = self._card(content, "② Output Location")
        s2.pack(fill="x", pady=(0, 10))
        or_ = ctk.CTkFrame(s2, fg_color="transparent")
        or_.pack(fill="x", padx=14, pady=(0, 12))
        self.out_label = ctk.CTkLabel(or_, text="Same folder as first file",
                                      text_color=BRAND_MUTED,
                                      font=ctk.CTkFont("Segoe UI", 11), anchor="w")
        self.out_label.pack(side="left", fill="x", expand=True)
        ctk.CTkButton(or_, text="Change…",
                      fg_color=BRAND_CARD, hover_color=BRAND_BORDER,
                      text_color=BRAND_TEXT, font=ctk.CTkFont("Segoe UI", 11),
                      width=88, height=30, corner_radius=8,
                      command=self._browse_out).pack(side="right")

        # ── Step 3 ────────────────────────────────────────────────────────────
        s3 = self._card(content, "③ Generate")
        s3.pack(fill="x", pady=(0, 10))
        gr = ctk.CTkFrame(s3, fg_color="transparent")
        gr.pack(fill="x", padx=14, pady=(0, 14))
        self.gen_btn = ctk.CTkButton(gr, text="⚡  Generate HTML Graph",
                                     fg_color=BRAND_GREEN, hover_color="#259e3e",
                                     text_color="#fff", font=ctk.CTkFont("Segoe UI", 13, "bold"),
                                     height=40, corner_radius=10, command=self._generate)
        self.gen_btn.pack(side="left", fill="x", expand=True, padx=(0, 10))
        self.open_btn = ctk.CTkButton(gr, text="🌐  Open",
                                      fg_color=BRAND_CARD, hover_color=BRAND_BORDER,
                                      text_color=BRAND_TEXT, font=ctk.CTkFont("Segoe UI", 13),
                                      width=90, height=40, corner_radius=10, state="disabled",
                                      command=self._open_output)
        self.open_btn.pack(side="right")

    # ── File management ───────────────────────────────────────────────────────

    def _add_file(self):
        paths = filedialog.askopenfilenames(
            title="Select Excel file(s)",
            filetypes=[("Excel files", "*.xlsx *.xlsm *.xls"), ("All files", "*.*")]
        )
        for p in paths:
            if p not in self._files:
                self._files.append(p)
                self._add_row(p)
                self._log(f"Added: {os.path.basename(p)}")

    def _add_row(self, path):
        self.empty_lbl.pack_forget()
        row = FileRow(self.file_list, path, on_remove=lambda p=path: self._remove(p))
        row.pack(fill="x", padx=6, pady=3)
        self._file_rows.append(row)

    def _remove(self, path):
        if path in self._files:
            self._files.remove(path)
        for row in list(self._file_rows):
            if row.path == path:
                row.destroy()
                self._file_rows.remove(row)
                break
        if not self._files:
            self.empty_lbl.pack(pady=10)
        self._log(f"Removed: {os.path.basename(path)}")

    def _browse_out(self):
        folder = filedialog.askdirectory(title="Select output folder")
        if folder:
            self._output_path = folder
            self.out_label.configure(text=folder, text_color=BRAND_TEXT)

    # ── Generation ────────────────────────────────────────────────────────────

    def _generate(self):
        if not self._files:
            messagebox.showwarning("No files", "Please add at least one Excel file.")
            return
        self.gen_btn.configure(state="disabled", text="Processing…")
        self.open_btn.configure(state="disabled")
        threading.Thread(target=self._run, daemon=True).start()

    def _run(self):
        try:
            all_bubbles, all_paths = [], []
            for path in self._files:
                name = os.path.basename(path)
                self._log(f"Reading {name}…")
                analyst_rows, se_rows = parse_excel(path)
                bubbles, invalid = build_bubble_data(analyst_rows, se_rows)
                if invalid:
                    self._log(f"⚠️ Skipped {len(invalid)} trends missing X/Y data: {', '.join([b['trend'] for b in invalid])}")
                if not bubbles:
                    self._log(f"  ❌ Nothing to plot in {name}, skipping.")
                    continue
                all_bubbles.append(bubbles)
                all_paths.append(path)

            if not all_bubbles:
                self._log("❌ No plottable data found.")
                self.root.after(0, self._reset_btn)
                return

            self._log("Generating HTML…")
            html = generate_html(all_bubbles, all_paths)

            out_dir = self._output_path or os.path.dirname(all_paths[0])
            base = "risk_analysis_graph" if len(all_paths) > 1 else \
                   os.path.splitext(os.path.basename(all_paths[0]))[0] + "_risk_graph"
            out_path = os.path.join(out_dir, f"{base}.html")
            with open(out_path, "w", encoding="utf-8") as f:
                f.write(html)

            self._last_output = out_path
            self._log(f"✅ Saved → {out_path}")
            self.root.after(0, self._done)

        except Exception as e:
            self._log(f"❌ Error: {e}")
            self.root.after(0, self._reset_btn)

    def _done(self):
        self.gen_btn.configure(state="normal", text="⚡  Generate HTML Graph")
        self.open_btn.configure(state="normal")

    def _reset_btn(self):
        self.gen_btn.configure(state="normal", text="⚡  Generate HTML Graph")

    def _open_output(self):
        if self._last_output and os.path.exists(self._last_output):
            webbrowser.open(f"file:///{self._last_output.replace(os.sep, '/')}")
        else:
            messagebox.showinfo("Not found", "Generate the graph first.")

    def _log(self, msg):
        print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")


if __name__ == "__main__":
    root = ctk.CTk()
    app = App(root)
    root.mainloop()